############       Copyright 2024        #############
############ Developed By Furkan ARINCI  #############
############       BioDNA Biosphere S1-JP ############
###########     Biosphere                  ###########
# 2024.2.25 - 03:00 AM
# -*- coding: utf-8 -*-



################################## Yazım Stili İşlemleri ###############################################
import playsound
import requests





################################## Kütüphane İşlemleri ###############################################

import os
import sys
import time
import openpyxl
import matplotlib
import numpy as np
import pandas as pd
from pathlib import Path
from playsound import playsound
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from matplotlib.backends.backend_pdf import PdfPages
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from matplotlib.dates import DateFormatter, AutoDateLocator,  ConciseDateFormatter





################################## Kütüphane İşlemleri Tamamlandı ###############################################

################################## Fonksiyon İşlemleri ###############################################


def print_islem_baslat(Kategori, islem_adi):
    print(f"[ {Kategori} ] ( {islem_adi} ) - Başladı!")

def print_islem_bitir(Kategori, islem_adi):
    print(f"[ {Kategori} ] ( {islem_adi} ) - Tamamlandı!")

def bio_control_indir():
    file_id = "14959N1kL8956eEkZXH3jIodNbtDG1tyc"
    while True:
        url = f"https://drive.google.com/u/0/uc?id={file_id}&export=download"
        session = requests.Session()
        response = session.get(url, stream=True)
        if response.status_code == 404:
            print("[ Lütfen Geliştiricinizden Version Kontrol Dosyası Anahtar KOD Talep Edin ]")
            file_id = input("[ Anahtar KOD'u giriniz: ")
        else:
            # Dosyayı indirme
            with open("BioControl.xlsx", "wb") as f:
                for chunk in response.iter_content(chunk_size=1024):
                    if chunk:
                        f.write(chunk)
            break
def bio_control_et():
    print('')
    print("[ Sürüm Kontrol Ediliyor ]")

    try:
        with open("BioV.txt", "r") as f:
            current_version = f.read().strip()
    except:
        pass

    version_bilgilendirme = "[ Güncelleme hizmetine ulaşılamadı! ] "

    try:
        bio_control_indir()
        # Dosya yolu ve adı
        file_path = "BioControl.xlsx"
        # Dosyayı aç
        wb = openpyxl.load_workbook(file_path)
        # Aktif sayfayı al
        sheet = wb.active

        # Excel'deki tüm BioCode'ları okuyup bir listeye ekleyin
        BioCodes = []
        row = 2  # Excel'de verilerin başladığı satır
        while True:
            BioCode = sheet[f"B{row}"].value
            if BioCode is None:
                break
            BioCodes.append(BioCode)
            row += 1

        # İstenen bilgileri al
        BioV = sheet["A2"].value
        BioT = sheet["C2"].value
        BioD = sheet["D2"].value

        # Dosyayı sil
        os.remove(file_path)

        if current_version != BioV:
            version_bilgilendirme = "[ BioV No: " + current_version + "." + " Biosphere DNA Güncel Değil. Güncel BioV No: " + str(
                BioV)
            # Fonksiyonu çağırarak dosyaları indirme
            print("[ GÜNCELLEME MEVCUT! \nYeni Biosphere DNA özelliklerinden faydalanmak için lütfen geliştiricinizle iletişime geçiniz ve size verdiği bağlantıdan yeni sürümü indiriniz ]")

        else:
            version_bilgilendirme = "[ BioV No: " + current_version + "." + " Biosphere DNA Güncel!"

        print(version_bilgilendirme)

        return BioV, BioCodes, BioT, BioD

    except Exception as e:
        print(f"[ Hata: {e} ]")
        print("[ Güncelleme hizmetine ulaşılamadı! ]")
        try:
            # "BioControl.xlsx" dosyasını silmeye çalış
            os.remove("BioControl.xlsx")
        except Exception as e:
            pass
        time.sleep(1000)
        sys.exit()
        return None

def bio_ses(file_path):
    try:
        playsound(file_path)
    except:
        pass

def copy_font(original_font):
    return Font(name=original_font.name, size=original_font.size, bold=original_font.bold, italic=original_font.italic,
                vertAlign=original_font.vertAlign, underline=original_font.underline, strike=original_font.strike,
                color=original_font.color)

def copy_fill(original_fill):
    return PatternFill(fill_type=original_fill.fill_type, start_color=original_fill.start_color,
                       end_color=original_fill.end_color)

def copy_border(original_border):
    sides = ['left', 'right', 'top', 'bottom', 'diagonal', 'vertical', 'horizontal']
    new_border_args = {}

    for side in sides:
        original_side = getattr(original_border, side)
        if original_side is None or original_side.border_style is None:
            new_border_args[side] = Side(border_style=None)
        else:
            new_border_args[side] = Side(border_style=original_side.border_style, color=original_side.color)

    return Border(**new_border_args)

def copy_alignment(original_alignment):
    return Alignment(horizontal=original_alignment.horizontal, vertical=original_alignment.vertical,
                     text_rotation=original_alignment.text_rotation, wrap_text=original_alignment.wrap_text,
                     shrink_to_fit=original_alignment.shrink_to_fit, indent=original_alignment.indent)

# 'Devre No' sütununu ayrıştırma fonksiyonunu genişletme
def ayir_genis(devre_no):
    if pd.isnull(devre_no) or not isinstance(devre_no, str):
        return ["-"] * 6  # Yeterli bölüm olmadığında
    parcalar = devre_no.split(".")
    # Gerekli bölümler eksikse, eksik bölümleri 'Tespit edilemedi' ile doldur
    while len(parcalar) < 6:
        parcalar.append("-")
    return parcalar[:6]
def get_max_day(df, col):
    return df.loc[df[col].idxmax(), 'Gün']

def get_min_day(df, col):
    return df.loc[df[col].idxmin(), 'Gün']

def get_mode(series):
    return series.mode()[0] if not series.mode().empty else None

# Kategorik metrikler için mod değerlerini hesaplayalım.
# Kategorik metrikler için 'safe_mode' fonksiyonunu kullanarak hata yönetimini iyileştireceğiz.
def safe_mode(series):
    if series.empty:
        return None
    mode_result = series.mode()
    return mode_result[0] if not mode_result.empty else None

if not os.path.exists("BioV.txt"):
    with open("BioV.txt", "w") as f:
        f.write("Biosphere S1-JP")

# Dosyanın var olup olmadığını kontrol et
if not os.path.exists('BioControl.txt'):
    # Dosya yoksa, oluştur ve içine "Girilmez" yaz
    with open('BioControl.txt', 'w') as file:
        file.write("Girilmez")

################################## Fonksiyon İşlemleri Tamamlandı ###############################################

def Orjinal_Veri_Optimizasyon():
    print('')

    ################################## Orjinal Veri İşlemleri ###############################################

    print_islem_baslat('Orjinal Veri', 'Optimizasyon')

    print('')

    veri_girisi_klasoru = 'Veri Girişi'
    data_frames = []

    # 'Veri Girişi' klasörünün varlığını kontrol et, yoksa oluştur
    if not os.path.exists(veri_girisi_klasoru):
        os.makedirs(veri_girisi_klasoru)
        print(f"[ [{veri_girisi_klasoru}]  klasörü oluşturuldu ]")

    def xlsx_dosyasi_var_mi(klasor):
        """Klasör içinde .xlsx uzantılı dosya varlığını kontrol et"""
        return any(dosya.endswith('.xlsx') for dosya in os.listdir(klasor))

    while not xlsx_dosyasi_var_mi(veri_girisi_klasoru):
        print("[ Veri Girişi klasörüne uygun bir .xlsx dosyası yüklemenizi bekliyorum...")
        time.sleep(5)  # 5 saniye beklet

    # .xlsx dosyalarını bul ve işle, 'Veri.xlsx' hariç
    xlsx_dosyalari = [f for f in os.listdir(veri_girisi_klasoru) if f.endswith('.xlsx') and f != 'BioVeri.xlsx']

    initial_rows = 0
    initial_columns = None

    for dosya in xlsx_dosyalari:
        dosya_yolu = os.path.join(veri_girisi_klasoru, dosya)
        veri = pd.read_excel(dosya_yolu)
        initial_rows += len(veri)
        if initial_columns is None:
            initial_columns = len(veri.columns)
        data_frames.append(veri)
        print(
            f"[ [{dosya}] dosyası başarıyla işlendi. Satır sayısı: {len(veri)}, Sütun sayısı: {len(veri.columns)} ]")

    # DataFrame'leri birleştir
    data_combined = pd.concat(data_frames, ignore_index=True)

    # Tekrar eden satırları sil
    before_dropping_rows = len(data_combined)
    data_cleaned = data_combined.drop_duplicates()
    after_dropping_rows = len(data_cleaned)

    # Sütunları karşılaştırarak tamamen aynı olanları bul ve sil
    columns_to_remove = []
    for i in range(len(data_cleaned.columns)):
        for j in range(i + 1, len(data_cleaned.columns)):
            if all(data_cleaned.iloc[:, i] == data_cleaned.iloc[:, j]):
                columns_to_remove.append(data_cleaned.columns[j])

    data_original_dosya = data_cleaned.drop(columns=set(columns_to_remove))

    # Verileri 'Veri.xlsx' olarak kaydet
    veri_xlsx_yolu = os.path.join(veri_girisi_klasoru, 'BioVeri.xlsx')
    data_original_dosya.to_excel(veri_xlsx_yolu, index=False)
    print(f"[ [{veri_xlsx_yolu}] olarak veri başarıyla kaydedildi ]")

    # Sonuçları yazdır
    print(f"[ Birleştirilmeden önceki toplam satır sayısı: {initial_rows} ]")
    print(f"[ Birleştirildikten sonraki satır sayısı: {before_dropping_rows} ]")
    if before_dropping_rows == after_dropping_rows:
        print("[ Hiçbir satır silinmedi ]")
    else:
        print(f"[ Tekrar eden satırlar silindikten sonraki satır sayısı: {after_dropping_rows} ]")
    print(f"[ İlk dosyadaki sütun sayısı: {initial_columns} ]")
    print(
        f"[ Tekrar eden sütunlar silindikten sonraki sütun sayısı: {len(data_original_dosya.columns)} ]")
    if columns_to_remove:
        print(f"[ Silinen sütun sayısı: {len(set(columns_to_remove))} ]")
        print(f"[ Silinen sütunlar: {set(columns_to_remove)}")
    else:
        print("[ Hiçbir sütun silinmedi ]")

    data_original = pd.read_excel(veri_xlsx_yolu)

    print('')

    print_islem_bitir('Orjinal Veri', 'Optimizasyon')
    bio_ses('BS05.mp3')

    return data_original

    ################################## Orjinal Veri İşlemleri Tamamlandı ###############################################

def Takım_Degerlendirme(data_original, BioT):
    print('')

    ################################## Takım Veri Ön Hazırlık İşlemleri ###############################################

    print_islem_baslat('Takım Değerlendirme', 'Veri Optimizasyon')

    data = data_original

    # Yıl, Ay ve Gün sütunlarını ekleyin
    data['Yıl'] = data['Çözüm Tarihi'].dt.year
    data['Ay'] = data['Çözüm Tarihi'].dt.month
    data['Gün'] = data['Çözüm Tarihi'].dt.day

    # Veri setindeki tarih/saat sütununu datetime türüne dönüştürme
    data['Çözüm Tarihi'] = pd.to_datetime(data['Çözüm Tarihi'], format='%d/%m/%Y %H:%M:%S')

    # Onay ve Ret durumlarını sayısal değerlere dönüştürme
    data['Onay'] = (data['Durum'] == 'Onaylandı').astype(int)

    # Eğer 'Ret' sütunu eklemek ve 'Var' değerleri için 1, diğer durumlar için 0 atamak istiyorsanız:
    data['Ret'] = (data['Ret Geçmişi'] == 'Var').astype(int)

    # Benzersiz çalışma günleri hesaplaması yapılıyor ve mevcut DataFrame üzerine yazılıyor
    data['Çalışma Günü (Ay)'] = data.groupby(['Takım', 'Yıl', 'Ay'])['Gün'].transform('nunique')


    # 'Kayıt Tarihi' ve 'Çözüm Tarihi' sütunlarını datetime formatına çevir
    data['Kayıt Tarihi'] = pd.to_datetime(data['Kayıt Tarihi'], format='%d/%m/%Y %H:%M:%S')
    data['Çözüm Tarihi'] = pd.to_datetime(data['Çözüm Tarihi'], format='%d/%m/%Y %H:%M:%S')

    # Her ekip ve her gün için başlangıç ve bitiş zamanlarını bulma
    data['Baslangic_Zamani'] = data.groupby(['Takım', data['Çözüm Tarihi'].dt.date])['Çözüm Tarihi'].transform('min')
    data['Bitis_Zamani'] = data.groupby(['Takım', data['Çözüm Tarihi'].dt.date])['Çözüm Tarihi'].transform('max')

    # Tekrar eden arızalar için veriyi düzenleme
    data_sorted = data.sort_values(by=['Hizmet No', 'Çözüm Tarihi'])
    data['Sonraki Arıza Tarihi'] = data_sorted.groupby('Hizmet No')['Çözüm Tarihi'].shift(-1)
    data['Arıza Tekrar'] = (
                (data['Sonraki Arıza Tarihi'] - data['Çözüm Tarihi']).dt.total_seconds() / (3600 * 24)).between(1,
                                                                                                                BioT)

    # Arıza tekrar adedini hesapla ve yeni bir sütuna ata
    data['Arıza Tekrar Adet (Gün)'] = data['Arıza Tekrar'].astype(int)

    # Çalışma süresini hesaplama (saat cinsinden) ve yuvarlama
    data['Çalışma Süresi (Saat)'] = round((data['Bitis_Zamani'] - data['Baslangic_Zamani']).dt.total_seconds() / 3600,
                                          1)

    # Çalışma saat aralığını hesaplayalım ve yeni bir sütun olarak ekleyelim
    data['Çalışma Aralığı (Saat)'] = data.apply(lambda x:
                                                f"{x['Baslangic_Zamani'].strftime('%H:%M') if not pd.isnull(x['Baslangic_Zamani']) else 'NaN'} - {x['Bitis_Zamani'].strftime('%H:%M') if not pd.isnull(x['Bitis_Zamani']) else 'NaN'}",
                                                axis=1)

    # 'Yıl', 'Ay' ve 'Gün' sütunlarında eksik değerleri içeren satırları bulun ve silin
    data.dropna(subset=['Yıl', 'Ay', 'Gün'], inplace=True)

    # 'Yıl', 'Ay' ve 'Gün' sütunlarını tamsayıya dönüştür
    data[['Yıl', 'Ay', 'Gün']] = data[['Yıl', 'Ay', 'Gün']].astype(int)

    # Çözüm süresini saat cinsinden hesaplayıp orijinal veri setine yazın
    data['Çözüm Süresi (Saat)'] = (data['Çözüm Tarihi'] - data['Kayıt Tarihi']).dt.total_seconds() / 3600

    ################################## Takım Veri Ön Hazırlık 3 İşlemleri ###############################################

    # İlk gruplama ve en çok gidilen Hizmet No'nun eklenmesi
    grouped = data.groupby(['Takım', 'Yıl', 'Ay', 'Gün'])['Hizmet No'].agg(
        lambda x: x.value_counts().index[0] if not x.empty else None).reset_index(name='En Çok Gidilen Hizmet No (Gün)')
    data = pd.merge(data, grouped, on=['Takım', 'Yıl', 'Ay', 'Gün'], how='left')

    # Arıza tekrarı olanlar için gruplama ve en sık tekrar eden Hizmet No'nun eklenmesi
    # Hizmet numarasına göre gruplama ve Arıza Tekrar Adet (Gün) sütununun oluşturulması
    grouped_repeat_count = data.groupby(['Takım', 'Yıl', 'Ay', 'Gün', 'Hizmet No'])[
        'Arıza Tekrar Adet (Gün)'].transform('first')

    # Arıza tekrarı olanlar için gruplama ve en sık tekrar eden Hizmet No'nun eklenmesi
    grouped = data[data['Arıza Tekrar'] == True].groupby(['Takım', 'Yıl', 'Ay', 'Gün'])['Hizmet No'].agg(
        lambda x: x.mode()[0] if not x.empty else None).reset_index(name='En Sık Tekrar Eden Hizmet No (Gün)')

    # Ana veri çerçevesine 'En Sık Tekrar Eden Hizmet No (Gün)' eklenmesi
    data = pd.merge(data, grouped, on=['Takım', 'Yıl', 'Ay', 'Gün'], how='left')

    # Ana veri çerçevesine 'Arıza Tekrar Adet (Gün)' sütununun eklenmesi
    data['Arıza Tekrar Adet (Gün)'] = grouped_repeat_count

    # Her ekip için en sık kullanılan çözüm açıklaması ve çözüm nedenlerinin eklenmesi
    grouped_solution = data.groupby(['Takım'])['Çözüm Açıklaması'].agg(
        lambda x: x.value_counts().index[0] if not x.empty else None).reset_index(name='En Sık Çözüm Açıklaması (Gün)')
    grouped_reason = data.groupby(['Takım'])['Çözüm Nedenleri'].agg(
        lambda x: x.value_counts().index[0] if not x.empty else None).reset_index(name='En Sık Çözüm Nedenleri (Gün)')
    data = pd.merge(data, grouped_solution, on=['Takım'], how='left')
    data = pd.merge(data, grouped_reason, on=['Takım'], how='left')

    # Günlük takım sayısının eklenmesi
    team_count_per_day = data.groupby(['Yıl', 'Ay', 'Gün']).apply(lambda x: x['Takım'].nunique()).reset_index(
        name='Takım Sayısı (Gün)')
    data = pd.merge(data, team_count_per_day, on=['Yıl', 'Ay', 'Gün'], how='left')

    # Arıza tekrarı olanlar için 'En Sık Kullanılan Çözüm Açıklaması (Gün)'
    grouped_repeat_solution = data[data['Arıza Tekrar'] == True].groupby(['Takım', 'Yıl', 'Ay', 'Gün'])[
        'Çözüm Açıklaması'].agg(lambda x: x.value_counts().index[0] if x.value_counts().size > 0 else None).reset_index(
        name='Tekrar Eden Arıza En Sık Çözüm Açıklaması (Gün)')
    data = pd.merge(data, grouped_repeat_solution, on=['Takım', 'Yıl', 'Ay', 'Gün'], how='left')

    # Arıza tekrarı olanlar için 'En Sık Kullanılan Çözüm Nedenleri (Gün)'
    grouped_repeat_reason = data[data['Arıza Tekrar'] == True].groupby(['Takım', 'Yıl', 'Ay', 'Gün'])[
        'Çözüm Nedenleri'].agg(lambda x: x.value_counts().index[0] if x.value_counts().size > 0 else None).reset_index(
        name='Tekrar Eden Arıza En Sık Çözüm Nedenleri (Gün)')
    data = pd.merge(data, grouped_repeat_reason, on=['Takım', 'Yıl', 'Ay', 'Gün'], how='left')

    print_islem_bitir('Takım Değerlendirme', 'Veri Optimizasyon')
    bio_ses('BS05.mp3')

    ################################## Takım Günlük Detaylı Değerlendirme İşlemleri ###############################################

    print_islem_baslat('Takım Değerlendirme', 'Günlük Detaylı Değerlendirme')

    # Günlük gruplandırılmış veriler için önceden tanımlanmış kısım
    ekip_daily_grouped = data.groupby(
        ['Yıl', 'Ay', 'Gün', 'Takım', 'Çalışma Süresi (Saat)', 'Çalışma Aralığı (Saat)', 'Çalışma Günü (Ay)',
         'En Çok Gidilen Hizmet No (Gün)', 'En Sık Tekrar Eden Hizmet No (Gün)', 'En Sık Çözüm Açıklaması (Gün)',
         'En Sık Çözüm Nedenleri (Gün)', 'Tekrar Eden Arıza En Sık Çözüm Açıklaması (Gün)',
         'Tekrar Eden Arıza En Sık Çözüm Nedenleri (Gün)', 'Takım Sayısı (Gün)']).agg(
        Günlük_Arıza=('Çağrı No', 'size'),
        Günlük_Onay=('Onay', 'sum'),
        Günlük_Ret=('Ret', 'sum'),
        Günlük_Tekrar=('Arıza Tekrar', 'sum'),
        Arıza_Tekrar_adet=('Arıza Tekrar Adet (Gün)', 'size'),
        En_Yuksek_Cozum_Sure=('Çözüm Süresi (Saat)', 'max'),
        Ortalama_Cozum_Sure=('Çözüm Süresi (Saat)', 'mean'),
    ).reset_index()
    # Ekip bazında Günlük Ret ve Tekrar Yüzdesi hesaplamaları
    ekip_daily_grouped['Günlük_Ret_Yüzdesi'] = (
                (ekip_daily_grouped['Günlük_Ret'] / ekip_daily_grouped['Günlük_Arıza']) * 100).round(2)
    ekip_daily_grouped['Günlük_Ret_Yüzdesi'] = ekip_daily_grouped['Günlük_Ret_Yüzdesi'].fillna(0)

    ekip_daily_grouped['Günlük_Tekrar_Yüzdesi'] = (
                (ekip_daily_grouped['Günlük_Tekrar'] / ekip_daily_grouped['Günlük_Arıza']) * 100).round(2)
    ekip_daily_grouped['Günlük_Tekrar_Yüzdesi'] = ekip_daily_grouped['Günlük_Tekrar_Yüzdesi'].fillna(0)

    # "Genel Değerlendirme" klasörünü kontrol edelim ve yoksa oluşturalım.
    general_evaluation_folder_path = Path('Günlük Değerlendirme')
    general_evaluation_folder_path.mkdir(parents=True, exist_ok=True)

    # "Genel Değerlendirme" klasörü içinde "Rapor" klasörünü kontrol edelim ve yoksa oluşturalım.
    report_folder_path = general_evaluation_folder_path / 'Takım Değerlendirme'
    report_folder_path.mkdir(parents=True, exist_ok=True)

    ekip_daily_grouped = ekip_daily_grouped.sort_values(by='Günlük_Arıza', ascending=False)

    # Excel dosyasının kaydedileceği tam yol
    final_output_path = report_folder_path / 'Takım Günlük Detaylı Değerlendirme Raporu.xlsx'

    ekip_daily_grouped.to_excel(final_output_path, index=False)

    # Ayarlanacak boyutlar: Başlık satırı yüksekliği 143 piksel, diğer satırlar 66 piksel, tüm sütunların genişliği 10 piksel.
    # Excel'de piksel yerine nokta ve karakter genişliği kullanıldığı için, bu değerler yaklaşık olarak dönüştürülecektir.

    # Pikseli Excel'e uygun birime dönüştürme yaklaşık değerleri
    # Satır yüksekliği doğrudan piksel olarak ayarlanabilir (yaklaşık olarak)
    # Sütun genişliği için Excel genişliği, piksel değerine göre ayarlanmalıdır. Excel genişliği ve piksel arasında doğrudan bir oran yoktur,
    # ancak genel olarak 1 karakter genişliği yaklaşık 7 piksele eşdeğerdir. Bu yüzden, 10 piksel yaklaşık 1.43 karakter genişliğine eşittir.

    # Başlık satırı ve diğer satırlar için yükseklik ayarları
    header_row_height = 143  # Excel'de doğrudan piksel olarak ayarlanabilir
    other_row_height = 66  # Excel'de doğrudan piksel olarak ayarlanabilir

    # Sütun genişliği için pikseli karakter genişliğine çevirme
    column_width_in_pixels = 66
    # Excel'de genişlik karakter cinsindendir, bu yüzden dönüşüm yapılır
    column_width_in_chars = column_width_in_pixels / 7  # Yaklaşık bir dönüşüm

    wb = openpyxl.load_workbook(final_output_path)
    ws = wb.active

    # Biçimlendirme ayarları
    header_fill = PatternFill(start_color='0D47A1', end_color='0D47A1', fill_type='solid')  # Lacivert arka plan
    white_font = Font(bold=True, color='FFFFFF')  # Başlık için beyaz, kalın yazı tipi
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                         bottom=Side(style='thin'))  # İnce kenarlık
    centered_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)  # Metni ortala ve kaydır

    # Başlık satırının yüksekliğini ayarla
    ws.row_dimensions[1].height = 143

    # Diğer tüm satırlar için yüksekliği ayarla
    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].height = 35

    # Tüm sütunlar için genişliği ayarla
    for col in ws.iter_cols(min_col=1, max_col=ws.max_column):
        ws.column_dimensions[col[0].column_letter].width = 73 / 7  # Piksel değerini Excel genişliğine çevirme

    # Tüm hücrelere biçimlendirme uygula
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
            cell.alignment = centered_alignment
            if cell.row == 1:  # Başlık satırı için ek biçimlendirme
                cell.fill = header_fill
                cell.font = white_font

    # Yuvarlanacak başlıkların listesi
    headers_to_round = [
        "Ortalama Ret Yüzdesi (Gün)",
        "Ortalama Tekrar Yüzdesi (Gün)",
        "En Yüksek Tekrar Yüzdesi (Gün)",
        "Ortalama Çözüm Süresi Saat (Gün)",
        "En Yüksek Çözüm Süresi Saat (Gün)",
        "Ortalama Çalışma Süresi Saati (Gün)",
        "En Yüksek Çalışma Süresi Saat (Gün)"
    ]

    # Sütun harflerini ve başlıklarını bul
    columns_to_round = {}
    for idx, col in enumerate(ws.iter_cols(1, ws.max_column, 1, 1, True), start=1):
        if col[0] in headers_to_round:
            columns_to_round[get_column_letter(idx)] = col[0]

    # Her belirtilen sütun için, sayısal hücreleri yuvarla ve formatı '0.0' olarak ayarla
    for col_letter, header in columns_to_round.items():
        for cell in ws[col_letter][1:]:  # Üstbilgiyi atla, 2. satırdan başla
            if isinstance(cell.value, (int, float)):
                # Sayısal değeri yuvarla ve tek ondalık basamakla formatla
                cell.value = round(cell.value, 1)
                cell.number_format = '0.0'

    # Save the workbook with all the changes
    wb.save(final_output_path)

    print_islem_bitir('Takım Değerlendirme', 'Günlük Detaylı Değerlendirme')

    ################################## Takım Günlük Değerlendirme İşlemleri ###############################################

    print_islem_baslat('Takım Değerlendirme', 'Günlük Değerlendirme')

    # 'Aylık Ortalama Takım Sayısı (Gün)' hesaplama ve doğrudan mevcut DataFrame'e eklenmesi
    ekip_daily_grouped['Aylık Ortalama Takım Sayısı (Gün)'] = ekip_daily_grouped.groupby(['Yıl', 'Ay'])[
        'Takım Sayısı (Gün)'].transform('mean')

    # Perform the groupings and calculations
    grouped = ekip_daily_grouped.groupby(['Yıl', 'Ay', 'Takım'])

    gün_bazli_aylik_degerlendirme = grouped.apply(lambda x: pd.Series({

        'Ortalama Arıza Adedi (Gün)': int(x['Günlük_Arıza'].mean()),
        'En Yüksek Arıza Adedi (Gün)': x['Günlük_Arıza'].max(),
        'En Yüksek Arıza Günü (Gün)': get_max_day(x, 'Günlük_Arıza'),

        'Ortalama Ret Yüzdesi (Gün)': x['Günlük_Ret_Yüzdesi'].mean(),
        'En Yüksek Ret Yüzdesi (Gün)': x['Günlük_Ret_Yüzdesi'].max(),
        'En Yüksek Ret Günü (Gün)': get_max_day(x, 'Günlük_Ret_Yüzdesi'),

        'Ortalama Tekrar Yüzdesi (Gün)': x['Günlük_Tekrar_Yüzdesi'].mean(),
        'En Yüksek Tekrar Yüzdesi (Gün)': x['Günlük_Tekrar_Yüzdesi'].max(),
        'En Yüksek Tekrar Günü (Gün)': get_max_day(x, 'Günlük_Tekrar_Yüzdesi'),

        'Ortalama Çözüm Süresi Saat (Gün)': x['Ortalama_Cozum_Sure'].mean(),
        'En Yüksek Çözüm Süresi Saat (Gün)': x['En_Yuksek_Cozum_Sure'].max(),
        'En Yüksek Çözüm Süresi Günü (Gün)': get_max_day(x, 'En_Yuksek_Cozum_Sure'),

        'Ortalama Çalışma Süresi Saati (Gün)': x['Çalışma Süresi (Saat)'].mean(),
        'En Yüksek Çalışma Süresi Saat (Gün)': x['Çalışma Süresi (Saat)'].max(),
        'En Yüksek Çalışma Süresi Günü (Gün)': get_max_day(x, 'Çalışma Süresi (Saat)'),

        'En Tercih Edilen Çalışma Aralığı (Saat)': get_mode(x['Çalışma Aralığı (Saat)']),

        'Çalışma Günü (Ay)': x['Çalışma Günü (Ay)'].iloc[0],
        'Ortalama Takım Sayısı (Gün)': int(x['Aylık Ortalama Takım Sayısı (Gün)'].mean()),

        'En Çok Gidilen Hizmet No (Gün)': get_mode(x['En Çok Gidilen Hizmet No (Gün)']),
        'En Sık Tekrar Eden Hizmet No (Gün)': get_mode(x['En Sık Tekrar Eden Hizmet No (Gün)']),

        'En Sık Çözüm Açıklaması (Gün)': get_mode(x['En Sık Çözüm Açıklaması (Gün)']),
        'En Sık Çözüm Nedenleri (Gün)': get_mode(x['En Sık Çözüm Nedenleri (Gün)']),

        'Tekrar Eden Arıza En Sık Çözüm Açıklaması (Gün)': get_mode(
            x['Tekrar Eden Arıza En Sık Çözüm Açıklaması (Gün)']),
        'Tekrar Eden Arıza En Sık Çözüm Nedenleri (Gün)': get_mode(x['Tekrar Eden Arıza En Sık Çözüm Nedenleri (Gün)']),

    })).reset_index()

    # "Genel Değerlendirme" klasörünü kontrol edelim ve yoksa oluşturalım.
    general_evaluation_folder_path = Path('Günlük Değerlendirme')
    general_evaluation_folder_path.mkdir(parents=True, exist_ok=True)

    # "Genel Değerlendirme" klasörü içinde "Rapor" klasörünü kontrol edelim ve yoksa oluşturalım.
    report_folder_path = general_evaluation_folder_path / 'Takım Değerlendirme'
    report_folder_path.mkdir(parents=True, exist_ok=True)

    gün_bazli_aylik_degerlendirme = gün_bazli_aylik_degerlendirme.sort_values(by='Ortalama Arıza Adedi (Gün)',
                                                                              ascending=False)

    # Excel dosyasının kaydedileceği tam yol
    final_output_path = report_folder_path / 'Takım Günlük Değerlendirme Raporu.xlsx'

    gün_bazli_aylik_degerlendirme.to_excel(final_output_path, index=False)

    # Ayarlanacak boyutlar: Başlık satırı yüksekliği 143 piksel, diğer satırlar 66 piksel, tüm sütunların genişliği 10 piksel.
    # Excel'de piksel yerine nokta ve karakter genişliği kullanıldığı için, bu değerler yaklaşık olarak dönüştürülecektir.

    # Pikseli Excel'e uygun birime dönüştürme yaklaşık değerleri
    # Satır yüksekliği doğrudan piksel olarak ayarlanabilir (yaklaşık olarak)
    # Sütun genişliği için Excel genişliği, piksel değerine göre ayarlanmalıdır. Excel genişliği ve piksel arasında doğrudan bir oran yoktur,
    # ancak genel olarak 1 karakter genişliği yaklaşık 7 piksele eşdeğerdir. Bu yüzden, 10 piksel yaklaşık 1.43 karakter genişliğine eşittir.

    # Başlık satırı ve diğer satırlar için yükseklik ayarları
    header_row_height = 143  # Excel'de doğrudan piksel olarak ayarlanabilir
    other_row_height = 66  # Excel'de doğrudan piksel olarak ayarlanabilir

    # Sütun genişliği için pikseli karakter genişliğine çevirme
    column_width_in_pixels = 66
    # Excel'de genişlik karakter cinsindendir, bu yüzden dönüşüm yapılır
    column_width_in_chars = column_width_in_pixels / 7  # Yaklaşık bir dönüşüm

    wb = openpyxl.load_workbook(final_output_path)
    ws = wb.active

    # Biçimlendirme ayarları
    header_fill = PatternFill(start_color='0D47A1', end_color='0D47A1', fill_type='solid')  # Lacivert arka plan
    white_font = Font(bold=True, color='FFFFFF')  # Başlık için beyaz, kalın yazı tipi
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                         bottom=Side(style='thin'))  # İnce kenarlık
    centered_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)  # Metni ortala ve kaydır

    # Başlık satırının yüksekliğini ayarla
    ws.row_dimensions[1].height = 143

    # Diğer tüm satırlar için yüksekliği ayarla
    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].height = 35

    # Tüm sütunlar için genişliği ayarla
    for col in ws.iter_cols(min_col=1, max_col=ws.max_column):
        ws.column_dimensions[col[0].column_letter].width = 73 / 7  # Piksel değerini Excel genişliğine çevirme

    # Tüm hücrelere biçimlendirme uygula
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
            cell.alignment = centered_alignment
            if cell.row == 1:  # Başlık satırı için ek biçimlendirme
                cell.fill = header_fill
                cell.font = white_font

    # Yuvarlanacak başlıkların listesi
    headers_to_round = [
        "Ortalama Ret Yüzdesi (Gün)",
        "Ortalama Tekrar Yüzdesi (Gün)",
        "En Yüksek Tekrar Yüzdesi (Gün)",
        "Ortalama Çözüm Süresi Saat (Gün)",
        "En Yüksek Çözüm Süresi Saat (Gün)",
        "Ortalama Çalışma Süresi Saati (Gün)",
        "En Yüksek Çalışma Süresi Saat (Gün)"
    ]

    # Sütun harflerini ve başlıklarını bul
    columns_to_round = {}
    for idx, col in enumerate(ws.iter_cols(1, ws.max_column, 1, 1, True), start=1):
        if col[0] in headers_to_round:
            columns_to_round[get_column_letter(idx)] = col[0]

    # Her belirtilen sütun için, sayısal hücreleri yuvarla ve formatı '0.0' olarak ayarla
    for col_letter, header in columns_to_round.items():
        for cell in ws[col_letter][1:]:  # Üstbilgiyi atla, 2. satırdan başla
            if isinstance(cell.value, (int, float)):
                # Sayısal değeri yuvarla ve tek ondalık basamakla formatla
                cell.value = round(cell.value, 1)
                cell.number_format = '0.0'

    # Save the workbook with all the changes
    wb.save(final_output_path)

    print_islem_bitir('Takım Değerlendirme', 'Günlük Değerlendirme')
    bio_ses('BS05.mp3')

    ################################## Takım Aylık Değerlendirme İşlemleri ###############################################

    print_islem_baslat('Takım Değerlendirme', 'Aylık Değerlendirme')

    # Veri setini yükleme
    data = pd.read_excel(final_output_path)

    # Sayısal metrikler listesi
    numerical_metrics = [

        "Ortalama Arıza Adedi (Gün)",
        "En Yüksek Arıza Adedi (Gün)",
        "En Yüksek Arıza Günü (Gün)",

        'Ortalama Ret Yüzdesi (Gün)',
        "En Yüksek Ret Yüzdesi (Gün)",
        "En Yüksek Ret Günü (Gün)",

        'Ortalama Tekrar Yüzdesi (Gün)',
        "En Yüksek Tekrar Yüzdesi (Gün)",
        "En Yüksek Tekrar Günü (Gün)",

        'Ortalama Çözüm Süresi Saat (Gün)',
        'En Yüksek Çözüm Süresi Saat (Gün)',
        'En Yüksek Çözüm Süresi Günü (Gün)',

        'Ortalama Çalışma Süresi Saati (Gün)',
        "En Yüksek Çalışma Süresi Saat (Gün)",
        "En Yüksek Çalışma Süresi Günü (Gün)",

        "Çalışma Günü (Ay)",
        'Ortalama Takım Sayısı (Gün)'
    ]

    # Kategorik metrikler listesi
    categorical_metrics = [

        'En Tercih Edilen Çalışma Aralığı (Saat)',

        "En Çok Gidilen Hizmet No (Gün)",
        "En Sık Tekrar Eden Hizmet No (Gün)",

        'En Sık Çözüm Açıklaması (Gün)',
        'En Sık Çözüm Nedenleri (Gün)',

        'Tekrar Eden Arıza En Sık Çözüm Açıklaması (Gün)',
        'Tekrar Eden Arıza En Sık Çözüm Nedenleri (Gün)',

    ]

    # 'Yıl' ve 'Ay' bazında sayısal metrikler için ortalama değerleri hesaplayalım ve sonuçları tam sayı olarak ayarlayalım.
    average_values_by_year_month = data.groupby(["Yıl", "Ay", 'Takım'])[numerical_metrics].mean().reset_index()
    average_values_by_year_month[numerical_metrics] = average_values_by_year_month[numerical_metrics].round(0).astype(
        int)

    # 'Yıl' ve 'Ay' bazında kategorik metrikler için mod değerlerini hesaplayalım.
    mode_values_by_year_month = data.groupby(["Yıl", "Ay", 'Takım']).agg(
        {metric: safe_mode for metric in categorical_metrics}).reset_index()

    # İki DataFrame'i "Yıl" ve "Ay" sütunlarına göre birleştir
    # Bu durumda her iki DataFrame'de de ortak olan "Yıl" ve "Ay" sütunları birleştirilmiş DataFrame'de bir kez gösterilir
    ay_bazli_yillik_degerlendirme = pd.merge(average_values_by_year_month, mode_values_by_year_month,
                                             on=["Yıl", "Ay", 'Takım'])

    # "Genel Değerlendirme" klasörünü kontrol edelim ve yoksa oluşturalım.
    general_evaluation_folder_path = Path('Aylık Değerlendirme')
    general_evaluation_folder_path.mkdir(parents=True, exist_ok=True)

    # "Genel Değerlendirme" klasörü içinde "Rapor" klasörünü kontrol edelim ve yoksa oluşturalım.
    report_folder_path = general_evaluation_folder_path / 'Takım Değerlendirme'
    report_folder_path.mkdir(parents=True, exist_ok=True)

    ay_bazli_yillik_degerlendirme = ay_bazli_yillik_degerlendirme.sort_values(by='Ortalama Arıza Adedi (Gün)',
                                                                              ascending=False)

    # Excel dosyasının kaydedileceği tam yol
    final_output_path = report_folder_path / 'Takım Aylık Değerlendirme Raporu.xlsx'

    ay_bazli_yillik_degerlendirme.to_excel(final_output_path, index=False)

    # Ayarlanacak boyutlar: Başlık satırı yüksekliği 143 piksel, diğer satırlar 66 piksel, tüm sütunların genişliği 10 piksel.
    # Excel'de piksel yerine nokta ve karakter genişliği kullanıldığı için, bu değerler yaklaşık olarak dönüştürülecektir.

    # Pikseli Excel'e uygun birime dönüştürme yaklaşık değerleri
    # Satır yüksekliği doğrudan piksel olarak ayarlanabilir (yaklaşık olarak)
    # Sütun genişliği için Excel genişliği, piksel değerine göre ayarlanmalıdır. Excel genişliği ve piksel arasında doğrudan bir oran yoktur,
    # ancak genel olarak 1 karakter genişliği yaklaşık 7 piksele eşdeğerdir. Bu yüzden, 10 piksel yaklaşık 1.43 karakter genişliğine eşittir.

    # Başlık satırı ve diğer satırlar için yükseklik ayarları
    header_row_height = 143  # Excel'de doğrudan piksel olarak ayarlanabilir
    other_row_height = 66  # Excel'de doğrudan piksel olarak ayarlanabilir

    # Sütun genişliği için pikseli karakter genişliğine çevirme
    column_width_in_pixels = 66
    # Excel'de genişlik karakter cinsindendir, bu yüzden dönüşüm yapılır
    column_width_in_chars = column_width_in_pixels / 7  # Yaklaşık bir dönüşüm

    wb = openpyxl.load_workbook(final_output_path)
    ws = wb.active

    # Biçimlendirme ayarları
    header_fill = PatternFill(start_color='0D47A1', end_color='0D47A1', fill_type='solid')  # Lacivert arka plan
    white_font = Font(bold=True, color='FFFFFF')  # Başlık için beyaz, kalın yazı tipi
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                         bottom=Side(style='thin'))  # İnce kenarlık
    centered_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)  # Metni ortala ve kaydır

    # Başlık satırının yüksekliğini ayarla
    ws.row_dimensions[1].height = 143

    # Diğer tüm satırlar için yüksekliği ayarla
    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].height = 35

    # Tüm sütunlar için genişliği ayarla
    for col in ws.iter_cols(min_col=1, max_col=ws.max_column):
        ws.column_dimensions[col[0].column_letter].width = 73 / 7  # Piksel değerini Excel genişliğine çevirme

    # Tüm hücrelere biçimlendirme uygula
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
            cell.alignment = centered_alignment
            if cell.row == 1:  # Başlık satırı için ek biçimlendirme
                cell.fill = header_fill
                cell.font = white_font

    # Yuvarlanacak başlıkların listesi
    headers_to_round = [
        "Ortalama Ret Yüzdesi (Gün)",
        "Ortalama Tekrar Yüzdesi (Gün)",
        "En Yüksek Tekrar Yüzdesi (Gün)",
        "Ortalama Çözüm Süresi Saat (Gün)",
        "En Yüksek Çözüm Süresi Saat (Gün)",
        "Ortalama Çalışma Süresi Saati (Gün)",
        "En Yüksek Çalışma Süresi Saat (Gün)"
    ]

    # Sütun harflerini ve başlıklarını bul
    columns_to_round = {}
    for idx, col in enumerate(ws.iter_cols(1, ws.max_column, 1, 1, True), start=1):
        if col[0] in headers_to_round:
            columns_to_round[get_column_letter(idx)] = col[0]

    # Her belirtilen sütun için, sayısal hücreleri yuvarla ve formatı '0.0' olarak ayarla
    for col_letter, header in columns_to_round.items():
        for cell in ws[col_letter][1:]:  # Üstbilgiyi atla, 2. satırdan başla
            if isinstance(cell.value, (int, float)):
                # Sayısal değeri yuvarla ve tek ondalık basamakla formatla
                cell.value = round(cell.value, 1)
                cell.number_format = '0.0'

    # Save the workbook with all the changes
    wb.save(final_output_path)

    print_islem_bitir('Takım Değerlendirme', 'Aylık Değerlendirme')
    bio_ses('BS05.mp3')

    ################################## Takım Yıllık Değerlendirme İşlemleri ###############################################

    print_islem_baslat('Takım Değerlendirme', 'Yıllık Değerlendirme')

    # Veri setini yükleme
    data = pd.read_excel(final_output_path)

    # Sayısal metrikler listesi
    numerical_metrics = [

        "Ortalama Arıza Adedi (Gün)",
        "En Yüksek Arıza Adedi (Gün)",
        "En Yüksek Arıza Günü (Gün)",

        'Ortalama Ret Yüzdesi (Gün)',
        "En Yüksek Ret Yüzdesi (Gün)",
        "En Yüksek Ret Günü (Gün)",

        'Ortalama Tekrar Yüzdesi (Gün)',
        "En Yüksek Tekrar Yüzdesi (Gün)",
        "En Yüksek Tekrar Günü (Gün)",

        'Ortalama Çözüm Süresi Saat (Gün)',
        'En Yüksek Çözüm Süresi Saat (Gün)',
        'En Yüksek Çözüm Süresi Günü (Gün)',

        'Ortalama Çalışma Süresi Saati (Gün)',
        "En Yüksek Çalışma Süresi Saat (Gün)",
        "En Yüksek Çalışma Süresi Günü (Gün)",

        "Çalışma Günü (Ay)",
        'Ortalama Takım Sayısı (Gün)'
    ]

    # Kategorik metrikler listesi
    categorical_metrics = [

        'En Tercih Edilen Çalışma Aralığı (Saat)',

        "En Çok Gidilen Hizmet No (Gün)",
        "En Sık Tekrar Eden Hizmet No (Gün)",

        'En Sık Çözüm Açıklaması (Gün)',
        'En Sık Çözüm Nedenleri (Gün)',

        'Tekrar Eden Arıza En Sık Çözüm Açıklaması (Gün)',
        'Tekrar Eden Arıza En Sık Çözüm Nedenleri (Gün)'
    ]

    # 'Yıl' ve 'Ay' bazında sayısal metrikler için ortalama değerleri hesaplayalım ve sonuçları tam sayı olarak ayarlayalım.
    average_values_by_year_month = data.groupby(["Yıl", 'Takım'])[numerical_metrics].mean().reset_index()
    average_values_by_year_month[numerical_metrics] = average_values_by_year_month[numerical_metrics].round(0).astype(
        int)

    # 'Yıl' ve 'Ay' bazında kategorik metrikler için mod değerlerini hesaplayalım.
    mode_values_by_year_month = data.groupby(["Yıl", 'Takım']).agg(
        {metric: safe_mode for metric in categorical_metrics}).reset_index()

    # İki DataFrame'i "Yıl" ve "Ay" sütunlarına göre birleştir
    # Bu durumda her iki DataFrame'de de ortak olan "Yıl" ve "Ay" sütunları birleştirilmiş DataFrame'de bir kez gösterilir
    yillik_bazli_genel_degerlendirme = pd.merge(average_values_by_year_month, mode_values_by_year_month,
                                                on=["Yıl", 'Takım'])

    # "Genel Değerlendirme" klasörünü kontrol edelim ve yoksa oluşturalım.
    general_evaluation_folder_path = Path('Yıllık Değerlendirme')
    general_evaluation_folder_path.mkdir(parents=True, exist_ok=True)

    # "Genel Değerlendirme" klasörü içinde "Rapor" klasörünü kontrol edelim ve yoksa oluşturalım.
    report_folder_path = general_evaluation_folder_path / 'Takım Değerlendirme'
    report_folder_path.mkdir(parents=True, exist_ok=True)


    yillik_bazli_genel_degerlendirme = yillik_bazli_genel_degerlendirme.sort_values(by='Ortalama Arıza Adedi (Gün)',
                                                                                    ascending=False)

    # Excel dosyasının kaydedileceği tam yol
    final_output_path = report_folder_path / 'Takım Yıllık Değerlendirme.xlsx'

    yillik_bazli_genel_degerlendirme.to_excel(final_output_path, index=False)

    # Ayarlanacak boyutlar: Başlık satırı yüksekliği 143 piksel, diğer satırlar 66 piksel, tüm sütunların genişliği 10 piksel.
    # Excel'de piksel yerine nokta ve karakter genişliği kullanıldığı için, bu değerler yaklaşık olarak dönüştürülecektir.

    # Pikseli Excel'e uygun birime dönüştürme yaklaşık değerleri
    # Satır yüksekliği doğrudan piksel olarak ayarlanabilir (yaklaşık olarak)
    # Sütun genişliği için Excel genişliği, piksel değerine göre ayarlanmalıdır. Excel genişliği ve piksel arasında doğrudan bir oran yoktur,
    # ancak genel olarak 1 karakter genişliği yaklaşık 7 piksele eşdeğerdir. Bu yüzden, 10 piksel yaklaşık 1.43 karakter genişliğine eşittir.

    # Başlık satırı ve diğer satırlar için yükseklik ayarları
    header_row_height = 143  # Excel'de doğrudan piksel olarak ayarlanabilir
    other_row_height = 66  # Excel'de doğrudan piksel olarak ayarlanabilir

    # Sütun genişliği için pikseli karakter genişliğine çevirme
    column_width_in_pixels = 66
    # Excel'de genişlik karakter cinsindendir, bu yüzden dönüşüm yapılır
    column_width_in_chars = column_width_in_pixels / 7  # Yaklaşık bir dönüşüm

    wb = openpyxl.load_workbook(final_output_path)
    ws = wb.active

    # Biçimlendirme ayarları
    header_fill = PatternFill(start_color='0D47A1', end_color='0D47A1', fill_type='solid')  # Lacivert arka plan
    white_font = Font(bold=True, color='FFFFFF')  # Başlık için beyaz, kalın yazı tipi
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                         bottom=Side(style='thin'))  # İnce kenarlık
    centered_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)  # Metni ortala ve kaydır

    # Başlık satırının yüksekliğini ayarla
    ws.row_dimensions[1].height = 143

    # Diğer tüm satırlar için yüksekliği ayarla
    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].height = 35

    # Tüm sütunlar için genişliği ayarla
    for col in ws.iter_cols(min_col=1, max_col=ws.max_column):
        ws.column_dimensions[col[0].column_letter].width = 73 / 7  # Piksel değerini Excel genişliğine çevirme

    # Tüm hücrelere biçimlendirme uygula
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
            cell.alignment = centered_alignment
            if cell.row == 1:  # Başlık satırı için ek biçimlendirme
                cell.fill = header_fill
                cell.font = white_font

    # Yuvarlanacak başlıkların listesi
    headers_to_round = [
        "Ortalama Ret Yüzdesi (Gün)",
        "Ortalama Tekrar Yüzdesi (Gün)",
        "En Yüksek Tekrar Yüzdesi (Gün)",
        "Ortalama Çözüm Süresi Saat (Gün)",
        "En Yüksek Çözüm Süresi Saat (Gün)",
        "Ortalama Çalışma Süresi Saati (Gün)",
        "En Yüksek Çalışma Süresi Saat (Gün)"
    ]

    # Sütun harflerini ve başlıklarını bul
    columns_to_round = {}
    for idx, col in enumerate(ws.iter_cols(1, ws.max_column, 1, 1, True), start=1):
        if col[0] in headers_to_round:
            columns_to_round[get_column_letter(idx)] = col[0]

    # Her belirtilen sütun için, sayısal hücreleri yuvarla ve formatı '0.0' olarak ayarla
    for col_letter, header in columns_to_round.items():
        for cell in ws[col_letter][1:]:  # Üstbilgiyi atla, 2. satırdan başla
            if isinstance(cell.value, (int, float)):
                # Sayısal değeri yuvarla ve tek ondalık basamakla formatla
                cell.value = round(cell.value, 1)
                cell.number_format = '0.0'

    # Save the workbook with all the changes
    wb.save(final_output_path)

    print_islem_bitir('Takım Değerlendirme', 'Yıllık Değerlendirme')
    bio_ses('BS05.mp3')

    ################################## Takım Genel Değerlendirme İşlemleri ###############################################

    print_islem_baslat('Takım Değerlendirme', 'Genel Değerlendirme')

    data = pd.read_excel(final_output_path)

    # Sayısal metrikler listesi
    numerical_metrics = [

        "Ortalama Arıza Adedi (Gün)",
        "En Yüksek Arıza Adedi (Gün)",
        "En Yüksek Arıza Günü (Gün)",

        'Ortalama Ret Yüzdesi (Gün)',
        "En Yüksek Ret Yüzdesi (Gün)",
        "En Yüksek Ret Günü (Gün)",

        'Ortalama Tekrar Yüzdesi (Gün)',
        "En Yüksek Tekrar Yüzdesi (Gün)",
        "En Yüksek Tekrar Günü (Gün)",

        'Ortalama Çözüm Süresi Saat (Gün)',
        'En Yüksek Çözüm Süresi Saat (Gün)',
        'En Yüksek Çözüm Süresi Günü (Gün)',

        'Ortalama Çalışma Süresi Saati (Gün)',
        "En Yüksek Çalışma Süresi Saat (Gün)",
        "En Yüksek Çalışma Süresi Günü (Gün)",

        "Çalışma Günü (Ay)",
        'Ortalama Takım Sayısı (Gün)'
    ]

    # Kategorik metrikler listesi
    categorical_metrics = [

        'En Tercih Edilen Çalışma Aralığı (Saat)',

        "En Çok Gidilen Hizmet No (Gün)",
        "En Sık Tekrar Eden Hizmet No (Gün)",

        'En Sık Çözüm Açıklaması (Gün)',
        'En Sık Çözüm Nedenleri (Gün)',

        'Tekrar Eden Arıza En Sık Çözüm Açıklaması (Gün)',
        'Tekrar Eden Arıza En Sık Çözüm Nedenleri (Gün)'
    ]

    # 'Yıl' ve 'Ay' bazında sayısal metrikler için ortalama değerleri hesaplayalım ve sonuçları tam sayı olarak ayarlayalım.
    average_values_by_year_month = data.groupby(['Takım'])[numerical_metrics].mean().reset_index()
    average_values_by_year_month[numerical_metrics] = average_values_by_year_month[numerical_metrics].round(0).astype(
        int)

    # 'Yıl' ve 'Ay' bazında kategorik metrikler için mod değerlerini hesaplayalım.
    mode_values_by_year_month = data.groupby(['Takım']).agg(
        {metric: safe_mode for metric in categorical_metrics}).reset_index()

    # İki DataFrame'i "Yıl" ve "Ay" sütunlarına göre birleştir
    # Bu durumda her iki DataFrame'de de ortak olan "Yıl" ve "Ay" sütunları birleştirilmiş DataFrame'de bir kez gösterilir
    genel_degerlendirme = pd.merge(average_values_by_year_month, mode_values_by_year_month, on=['Takım'])

    # "Genel Değerlendirme" klasörünü kontrol edelim ve yoksa oluşturalım.
    general_evaluation_folder_path = Path('Genel Değerlendirme')
    general_evaluation_folder_path.mkdir(parents=True, exist_ok=True)

    # "Genel Değerlendirme" klasörü içinde "Rapor" klasörünü kontrol edelim ve yoksa oluşturalım.
    report_folder_path = general_evaluation_folder_path / 'Takım Değerlendirme'
    report_folder_path.mkdir(parents=True, exist_ok=True)


    genel_degerlendirme = genel_degerlendirme.sort_values(by='Ortalama Arıza Adedi (Gün)', ascending=False)

    # Excel dosyasının kaydedileceği tam yol
    final_output_path = report_folder_path / 'Takım Genel Değerlendirme Raporu.xlsx'

    genel_degerlendirme.to_excel(final_output_path, index=False)

    # Ayarlanacak boyutlar: Başlık satırı yüksekliği 143 piksel, diğer satırlar 66 piksel, tüm sütunların genişliği 10 piksel.
    # Excel'de piksel yerine nokta ve karakter genişliği kullanıldığı için, bu değerler yaklaşık olarak dönüştürülecektir.

    # Pikseli Excel'e uygun birime dönüştürme yaklaşık değerleri
    # Satır yüksekliği doğrudan piksel olarak ayarlanabilir (yaklaşık olarak)
    # Sütun genişliği için Excel genişliği, piksel değerine göre ayarlanmalıdır. Excel genişliği ve piksel arasında doğrudan bir oran yoktur,
    # ancak genel olarak 1 karakter genişliği yaklaşık 7 piksele eşdeğerdir. Bu yüzden, 10 piksel yaklaşık 1.43 karakter genişliğine eşittir.

    # Başlık satırı ve diğer satırlar için yükseklik ayarları
    header_row_height = 143  # Excel'de doğrudan piksel olarak ayarlanabilir
    other_row_height = 66  # Excel'de doğrudan piksel olarak ayarlanabilir

    # Sütun genişliği için pikseli karakter genişliğine çevirme
    column_width_in_pixels = 66
    # Excel'de genişlik karakter cinsindendir, bu yüzden dönüşüm yapılır
    column_width_in_chars = column_width_in_pixels / 7  # Yaklaşık bir dönüşüm

    wb = openpyxl.load_workbook(final_output_path)
    ws = wb.active

    # Biçimlendirme ayarları
    header_fill = PatternFill(start_color='0D47A1', end_color='0D47A1', fill_type='solid')  # Lacivert arka plan
    white_font = Font(bold=True, color='FFFFFF')  # Başlık için beyaz, kalın yazı tipi
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                         bottom=Side(style='thin'))  # İnce kenarlık
    centered_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)  # Metni ortala ve kaydır

    # Başlık satırının yüksekliğini ayarla
    ws.row_dimensions[1].height = 143

    # Diğer tüm satırlar için yüksekliği ayarla
    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].height = 35

    # Tüm sütunlar için genişliği ayarla
    for col in ws.iter_cols(min_col=1, max_col=ws.max_column):
        ws.column_dimensions[col[0].column_letter].width = 73 / 7  # Piksel değerini Excel genişliğine çevirme

    # Tüm hücrelere biçimlendirme uygula
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
            cell.alignment = centered_alignment
            if cell.row == 1:  # Başlık satırı için ek biçimlendirme
                cell.fill = header_fill
                cell.font = white_font

    # Yuvarlanacak başlıkların listesi
    headers_to_round = [
        "Ortalama Ret Yüzdesi (Gün)",
        "Ortalama Tekrar Yüzdesi (Gün)",
        "En Yüksek Tekrar Yüzdesi (Gün)",
        "Ortalama Çözüm Süresi Saat (Gün)",
        "En Yüksek Çözüm Süresi Saat (Gün)",
        "Ortalama Çalışma Süresi Saati (Gün)",
        "En Yüksek Çalışma Süresi Saat (Gün)"
    ]

    # Sütun harflerini ve başlıklarını bul
    columns_to_round = {}
    for idx, col in enumerate(ws.iter_cols(1, ws.max_column, 1, 1, True), start=1):
        if col[0] in headers_to_round:
            columns_to_round[get_column_letter(idx)] = col[0]

    # Her belirtilen sütun için, sayısal hücreleri yuvarla ve formatı '0.0' olarak ayarla
    for col_letter, header in columns_to_round.items():
        for cell in ws[col_letter][1:]:  # Üstbilgiyi atla, 2. satırdan başla
            if isinstance(cell.value, (int, float)):
                # Sayısal değeri yuvarla ve tek ondalık basamakla formatla
                cell.value = round(cell.value, 1)
                cell.number_format = '0.0'

    # Save the workbook with all the changes
    wb.save(final_output_path)

    print_islem_bitir('Takım Değerlendirme', 'Genel Değerlendirme')

    ################################## Takım İşlemleri Tamamlandı ###############################################

    print_islem_bitir('Takım Değerlendirme', 'Tüm Değerlendirme')
    bio_ses('BS05.mp3')

def Abone_Degerlendirme(data_original, BioT):
    print('')

    ################################## Abone Veri Optimizasyon İşlemleri ###############################################

    print_islem_baslat('Abone Değerlendirme', 'Veri Optimizasyon')

    data = data_original

    # Yıl, Ay ve Gün sütunlarını ekleyin
    data['Yıl'] = data['Çözüm Tarihi'].dt.year
    data['Ay'] = data['Çözüm Tarihi'].dt.month
    data['Gün'] = data['Çözüm Tarihi'].dt.day

    # Veri setindeki tarih/saat sütununu datetime türüne dönüştürme
    data['Çözüm Tarihi'] = pd.to_datetime(data['Çözüm Tarihi'], format='%d/%m/%Y %H:%M:%S')

    # Onay ve Ret durumlarını sayısal değerlere dönüştürme
    data['Onay'] = (data['Durum'] == 'Onaylandı').astype(int)

    # Eğer 'Ret' sütunu eklemek ve 'Var' değerleri için 1, diğer durumlar için 0 atamak istiyorsanız:
    data['Ret'] = (data['Ret Geçmişi'] == 'Var').astype(int)

    # Benzersiz çalışma günleri hesaplaması yapılıyor ve mevcut DataFrame üzerine yazılıyor
    data['Arıza Yaşanan Gün (Ay)'] = data.groupby(['Hizmet No', 'Yıl', 'Ay'])['Gün'].transform('nunique')

    ################################## Abone Veri Optimizasyon İşlemleri ###############################################

    # Her ekip ve her gün için başlangıç ve bitiş zamanlarını bulma
    data['Baslangic_Zamani'] = data.groupby(['Hizmet No', data['Kayıt Tarihi'].dt.date])['Kayıt Tarihi'].transform(
        'min')
    data['Bitis_Zamani'] = data.groupby(['Hizmet No', data['Kayıt Tarihi'].dt.date])['Kayıt Tarihi'].transform('max')

    # 'Kayıt Tarihi' ve 'Çözüm Tarihi' sütunlarını datetime türüne dönüştürme
    data['Kayıt Tarihi'] = pd.to_datetime(data['Kayıt Tarihi'], format='%d/%m/%Y %H:%M:%S')
    data['Çözüm Tarihi'] = pd.to_datetime(data['Çözüm Tarihi'], format='%d/%m/%Y %H:%M:%S')

    # Arıza süresini saat cinsinden hesaplama
    data['Arıza Süresi Saat (Gün)'] = (data['Çözüm Tarihi'] - data['Kayıt Tarihi']).dt.total_seconds() / 3600

    # Çalışma süresini hesaplama (saat cinsinden) ve yuvarlama
    data['Yoğunluk Süresi (Saat)'] = round((data['Bitis_Zamani'] - data['Baslangic_Zamani']).dt.total_seconds() / 3600,
                                           1)

    # Çalışma saat aralığını hesaplayalım ve yeni bir sütun olarak ekleyelim
    data['Müsait Zaman (Saat)'] = data.apply(lambda x:
                                             f"{x['Baslangic_Zamani'].strftime('%H:%M') if not pd.isnull(x['Baslangic_Zamani']) else 'NaN'} - {x['Bitis_Zamani'].strftime('%H:%M') if not pd.isnull(x['Bitis_Zamani']) else 'NaN'}",
                                             axis=1)

    # Her abone ve her gün için toplam arıza süresini hesaplama
    data['Kayıt Tarihi (Gün)'] = data['Kayıt Tarihi'].dt.date
    toplam_ariza_suresi = data.groupby(['Hizmet No', 'Kayıt Tarihi (Gün)'])[
        'Arıza Süresi Saat (Gün)'].sum().reset_index(name='Toplam Arıza Süresi Saat (Gün)')

    # Tekrar eden arızalar için veriyi düzenleme
    data_sorted = data.sort_values(by=['Hizmet No', 'Çözüm Tarihi'])
    data['Sonraki Arıza Tarihi'] = data_sorted.groupby('Hizmet No')['Çözüm Tarihi'].shift(-1)
    data['Arıza Tekrar'] = (
                (data['Sonraki Arıza Tarihi'] - data['Çözüm Tarihi']).dt.total_seconds() / (3600 * 24)).between(1,
                                                                                                                BioT)

    ################################## Abone Veri Optimizasyon İşlemleri ###############################################

    # Hizmet numarasına göre gruplama ve Arıza Tekrar Adet (Gün) sütununun oluşturulması
    grouped_repeat_count = data.groupby(['Hizmet No', 'Yıl', 'Ay', 'Gün'])['Arıza Tekrar Adet (Gün)'].transform('first')

    # Gruplama işlemi ve en sık tekrar eden arıza numarasını bulma
    grouped = data[data['Arıza Tekrar'] == True].groupby(['Hizmet No', 'Yıl', 'Ay', 'Gün'])['Takım'].agg(
        lambda x: x.value_counts().index[0]).reset_index(name='En Sık Tekrar Eden Takım (Gün)')

    # Ana veri çerçevesine 'Arıza Tekrar Adet (Gün)' sütununun eklenmesi
    data['Arıza Tekrar Adet (Gün)'] = grouped_repeat_count

    # Sonucu orijinal DataFrame'e ekleyin
    data = pd.merge(data, grouped, on=['Hizmet No', 'Yıl', 'Ay', 'Gün'], how='left')

    # Her hizmet numarası için en sık karşılaşılan Arıza Tipini bulma
    en_sik_ariza_tipi = data.groupby('Hizmet No')['Arıza Tipi'].agg(lambda x: x.value_counts().index[0]).reset_index(
        name='En Sık Arıza Tipi')
    en_sik_ariza_aciklamasi = data.groupby('Hizmet No')['Açıklama'].agg(
        lambda x: x.value_counts().index[0]).reset_index(name='En Sık Arıza Açıklama')

    # Bulunan en sık Arıza Tipini 'Arıza Nedeni' olarak mevcut DataFrame'e eklemek
    # Bu adımda, öncelikle 'Hizmet No' sütunu üzerinden birleştirme işlemi yapacağız.
    data = pd.merge(data, en_sik_ariza_tipi, on='Hizmet No', how='left')
    data = pd.merge(data, en_sik_ariza_aciklamasi, on='Hizmet No', how='left')

    # Şimdi, 'En Sık Arıza Tipi' bilgisini 'Arıza Nedeni' sütununa kopyalayabiliriz.
    data['En Sık Arıza Nedeni (Gün)'] = data['En Sık Arıza Tipi']
    data['En Sık Arıza Açıklaması (Gün)'] = data['En Sık Arıza Açıklama']

    # 'En Sık Arıza Tipi' sütununu kaldırabiliriz, eğer artık bu sütuna ihtiyacımız yoksa
    data.drop('En Sık Arıza Tipi', axis=1, inplace=True)
    data.drop('En Sık Arıza Açıklama', axis=1, inplace=True)

    ################################## Abone Veri Optimizasyon İşlemleri ###############################################

    # 'Devre No' sütununu kullanarak 'İş Ortağı', 'Bölge' ve 'FN' bilgilerini ayırma
    def ayir(devre_no):
        # Eğer devre_no boş ise veya uygun formatı taşımıyorsa, 'Tespit edilemedi' döndür
        if pd.isnull(devre_no) or not isinstance(devre_no, str) or len(devre_no.split(".")) < 3:
            return ["Tespit edilemedi", "Tespit edilemedi", "Tespit edilemedi"]

        # Devre No'yu "." ile ayır ve gerekli bölümleri al
        parcalar = devre_no.split(".")
        return parcalar[0:3]  # İlk üç parçayı döndür

    # Hizmet No bazında en sık kullanılan çözüm açıklaması ve çözüm nedenlerini bulma
    grouped_solution_hizmetno = data.groupby(['Hizmet No'])['Çözüm Açıklaması'].agg(
        lambda x: x.value_counts().index[0] if not x.value_counts().empty else "Bilinmiyor").reset_index(
        name='En Sık Çözüm Açıklaması (Gün)')
    grouped_reason_hizmetno = data.groupby(['Hizmet No'])['Çözüm Nedenleri'].agg(
        lambda x: x.value_counts().index[0] if not x.value_counts().empty else "Bilinmiyor").reset_index(
        name='En Sık Çözüm Nedenleri (Gün)')

    # Sonuçları orijinal DataFrame'e ekleyin
    data = pd.merge(data, grouped_solution_hizmetno, on=['Hizmet No'], how='left')
    data = pd.merge(data, grouped_reason_hizmetno, on=['Hizmet No'], how='left')

    # Yeni sütunları hesapla ve ekle
    data[['İş Ortağı', 'Bölge', 'FN']] = pd.DataFrame(data['Devre No'].apply(ayir).tolist(), index=data.index)

    # Boş veya uyumsuz değerleri 'Tespit edilemedi' ile doldurma
    data[['İş Ortağı', 'Bölge', 'FN']] = data[['İş Ortağı', 'Bölge', 'FN']].fillna("Tespit edilemedi")

    print_islem_bitir('Abone Değerlendirme', 'Veri Optimizasyon')
    bio_ses('BS05.mp3')

    ################################## Abone Günlük Detaylı Değerlendirme İşlemleri ###############################################

    print_islem_baslat('Abone Değerlendirme', 'Günlük Detaylı Değerlendirme')

    # Yıl, Ay, Gün sütunlarını Nullable Integer türüne dönüştür
    data['Yıl'] = data['Yıl'].astype('Int64')
    data['Ay'] = data['Ay'].astype('Int64')
    data['Gün'] = data['Gün'].astype('Int64')

    abone_daily_grouped = data.groupby(
        ['Yıl', 'Ay', 'Gün', 'Hizmet No', 'Arıza Yaşanan Gün (Ay)', 'En Sık Tekrar Eden Takım (Gün)',
         'En Sık Arıza Nedeni (Gün)', 'En Sık Arıza Açıklaması (Gün)', 'En Sık Çözüm Açıklaması (Gün)',
         'En Sık Çözüm Nedenleri (Gün)', 'Yoğunluk Süresi (Saat)', 'Müsait Zaman (Saat)', 'İş Ortağı', 'Bölge',
         'FN']).agg(
        Günlük_Arıza=('Hizmet No', 'size'),
        Günlük_Onay=('Onay', 'sum'),
        Günlük_Ret=('Ret', 'sum'),
        Günlük_Tekrar=('Arıza Tekrar', 'size'),
        En_Yüksek_Arıza_Süresi=('Arıza Süresi Saat (Gün)', 'max'),
        Ortalama_Arıza_Süresi=('Arıza Süresi Saat (Gün)', 'mean')

    ).reset_index()
    # Ekip bazında Günlük Ret ve Tekrar Yüzdesi hesaplamaları
    abone_daily_grouped['Günlük_Ret_Yüzdesi'] = (
                (abone_daily_grouped['Günlük_Ret'] / abone_daily_grouped['Günlük_Arıza']) * 100).round(2)
    abone_daily_grouped['Günlük_Ret_Yüzdesi'] = abone_daily_grouped['Günlük_Ret_Yüzdesi'].fillna(0)

    abone_daily_grouped['Günlük_Tekrar_Yüzdesi'] = (
                (abone_daily_grouped['Günlük_Tekrar'] / abone_daily_grouped['Günlük_Arıza']) * 100).round(2)
    abone_daily_grouped['Günlük_Tekrar_Yüzdesi'] = abone_daily_grouped['Günlük_Tekrar_Yüzdesi'].fillna(0)

    # "Genel Değerlendirme" klasörünü kontrol edelim ve yoksa oluşturalım.
    general_evaluation_folder_path = Path('Günlük Değerlendirme')
    general_evaluation_folder_path.mkdir(parents=True, exist_ok=True)

    # "Genel Değerlendirme" klasörü içinde "Rapor" klasörünü kontrol edelim ve yoksa oluşturalım.
    report_folder_path = general_evaluation_folder_path / 'Abone Değerlendirme'
    report_folder_path.mkdir(parents=True, exist_ok=True)

    abone_daily_grouped = abone_daily_grouped.sort_values(by='Arıza Yaşanan Gün (Ay)', ascending=False)

    # Excel dosyasının kaydedileceği tam yol
    final_output_path = report_folder_path / 'Abone Günlük Detaylı Değerlendirme Raporu.xlsx'

    abone_daily_grouped.to_excel(final_output_path, index=False)

    # Ayarlanacak boyutlar: Başlık satırı yüksekliği 143 piksel, diğer satırlar 66 piksel, tüm sütunların genişliği 10 piksel.
    # Excel'de piksel yerine nokta ve karakter genişliği kullanıldığı için, bu değerler yaklaşık olarak dönüştürülecektir.

    # Pikseli Excel'e uygun birime dönüştürme yaklaşık değerleri
    # Satır yüksekliği doğrudan piksel olarak ayarlanabilir (yaklaşık olarak)
    # Sütun genişliği için Excel genişliği, piksel değerine göre ayarlanmalıdır. Excel genişliği ve piksel arasında doğrudan bir oran yoktur,
    # ancak genel olarak 1 karakter genişliği yaklaşık 7 piksele eşdeğerdir. Bu yüzden, 10 piksel yaklaşık 1.43 karakter genişliğine eşittir.

    # Başlık satırı ve diğer satırlar için yükseklik ayarları
    header_row_height = 143  # Excel'de doğrudan piksel olarak ayarlanabilir
    other_row_height = 66  # Excel'de doğrudan piksel olarak ayarlanabilir

    # Sütun genişliği için pikseli karakter genişliğine çevirme
    column_width_in_pixels = 66
    # Excel'de genişlik karakter cinsindendir, bu yüzden dönüşüm yapılır
    column_width_in_chars = column_width_in_pixels / 7  # Yaklaşık bir dönüşüm

    wb = openpyxl.load_workbook(final_output_path)
    ws = wb.active

    # Biçimlendirme ayarları
    header_fill = PatternFill(start_color='60b516', end_color='60b516', fill_type='solid')  # Lacivert arka plan
    white_font = Font(bold=True, color='FFFFFF')  # Başlık için beyaz, kalın yazı tipi
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                         bottom=Side(style='thin'))  # İnce kenarlık
    centered_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)  # Metni ortala ve kaydır

    # Başlık satırının yüksekliğini ayarla
    ws.row_dimensions[1].height = 143

    # Diğer tüm satırlar için yüksekliği ayarla
    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].height = 35

    # Tüm sütunlar için genişliği ayarla
    for col in ws.iter_cols(min_col=1, max_col=ws.max_column):
        ws.column_dimensions[col[0].column_letter].width = 73 / 7  # Piksel değerini Excel genişliğine çevirme

    # Tüm hücrelere biçimlendirme uygula
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
            cell.alignment = centered_alignment
            if cell.row == 1:  # Başlık satırı için ek biçimlendirme
                cell.fill = header_fill
                cell.font = white_font

    # Yuvarlanacak başlıkların listesi
    headers_to_round = [
        "En Yüksek Çözüm Süresi Saat (Gün)",
        "Ortalama Arıza Süresi Saat (Gün)",
    ]

    # Sütun harflerini ve başlıklarını bul
    columns_to_round = {}
    for idx, col in enumerate(ws.iter_cols(1, ws.max_column, 1, 1, True), start=1):
        if col[0] in headers_to_round:
            columns_to_round[get_column_letter(idx)] = col[0]

    # Her belirtilen sütun için, sayısal hücreleri yuvarla ve formatı '0.0' olarak ayarla
    for col_letter, header in columns_to_round.items():
        for cell in ws[col_letter][1:]:  # Üstbilgiyi atla, 2. satırdan başla
            if isinstance(cell.value, (int, float)):
                # Sayısal değeri yuvarla ve tek ondalık basamakla formatla
                cell.value = round(cell.value, 1)
                cell.number_format = '0.0'

    # Save the workbook with all the changes
    wb.save(final_output_path)

    print_islem_bitir('Abone Değerlendirme', 'Günlük Detaylı Değerlendirme')

    ################################## Abone Günlük Değerlendirme İşlemleri ###############################################

    print_islem_baslat('Abone Değerlendirme', 'Günlük Değerlendirme')

    # Perform the groupings and calculations
    grouped = abone_daily_grouped.groupby(['Yıl', 'Ay', 'Hizmet No'])

    gün_bazli_aylik_degerlendirme = grouped.apply(lambda x: pd.Series({

        'Arıza Yaşanan Gün (Ay)': int(x['Arıza Yaşanan Gün (Ay)'].iloc[0]),

        'Ortalama Arıza Adedi (Gün)': int(x['Günlük_Arıza'].mean()),
        'En Yüksek Arıza Adedi (Gün)': x['Günlük_Arıza'].max(),
        'En Yüksek Arıza Günü (Gün)': get_max_day(x, 'Günlük_Arıza'),

        'Ortalama Tekrar Yüzdesi (Gün)': x['Günlük_Tekrar_Yüzdesi'].mean(),
        'En Yüksek Çözüm Süresi Saat (Gün)': x['En_Yüksek_Arıza_Süresi'].max(),
        'En Yüksek Çözüm Süresi Günü (Gün)': get_max_day(x, 'En_Yüksek_Arıza_Süresi'),

        'Ortalama Arıza Süresi Saat (Gün)': x['Ortalama_Arıza_Süresi'].mean(),
        'Müsait Olduğu Zaman (Saat)': get_mode(x['Müsait Zaman (Saat)']),

        'Ortalama Yoğunluk Süresi Saati (Gün)': x['Yoğunluk Süresi (Saat)'].mean(),
        'Arıza Yaşanan Gün (Ay)': x['Arıza Yaşanan Gün (Ay)'].iloc[0],

        'En Sık Tekrar Eden Takım (Gün)': get_mode(x['En Sık Tekrar Eden Takım (Gün)']),
        'En Sık Arıza Nedeni (Gün)': get_mode(x['En Sık Arıza Nedeni (Gün)']),
        'En Sık Arıza Açıklaması (Gün)': get_mode(x['En Sık Arıza Açıklaması (Gün)']),
        'En Sık Çözüm Açıklaması (Gün)': get_mode(x['En Sık Çözüm Açıklaması (Gün)']),
        'En Sık Çözüm Nedenleri (Gün)': get_mode(x['En Sık Çözüm Nedenleri (Gün)']),
        'Arıza Tekrar Adet (Gün)': int(x['Günlük_Tekrar'].iloc[0]),

        'İş Ortağı': get_mode(x['İş Ortağı']),
        'Bölge': get_mode(x['Bölge']),
        'FN': get_mode(x['FN']),

    })).reset_index()

    # "Genel Değerlendirme" klasörünü kontrol edelim ve yoksa oluşturalım.
    general_evaluation_folder_path = Path('Günlük Değerlendirme')
    general_evaluation_folder_path.mkdir(parents=True, exist_ok=True)

    # "Genel Değerlendirme" klasörü içinde "Rapor" klasörünü kontrol edelim ve yoksa oluşturalım.
    report_folder_path = general_evaluation_folder_path / 'Abone Değerlendirme'
    report_folder_path.mkdir(parents=True, exist_ok=True)

    # Excel dosyasının kaydedileceği tam yol
    final_output_path = report_folder_path / 'Abone Günlük Değerlendirme Raporu.xlsx'

    gün_bazli_aylik_degerlendirme = gün_bazli_aylik_degerlendirme.sort_values(by='Arıza Yaşanan Gün (Ay)',
                                                                              ascending=False)

    gün_bazli_aylik_degerlendirme.to_excel(final_output_path, index=False)

    # Ayarlanacak boyutlar: Başlık satırı yüksekliği 143 piksel, diğer satırlar 66 piksel, tüm sütunların genişliği 10 piksel.
    # Excel'de piksel yerine nokta ve karakter genişliği kullanıldığı için, bu değerler yaklaşık olarak dönüştürülecektir.

    # Pikseli Excel'e uygun birime dönüştürme yaklaşık değerleri
    # Satır yüksekliği doğrudan piksel olarak ayarlanabilir (yaklaşık olarak)
    # Sütun genişliği için Excel genişliği, piksel değerine göre ayarlanmalıdır. Excel genişliği ve piksel arasında doğrudan bir oran yoktur,
    # ancak genel olarak 1 karakter genişliği yaklaşık 7 piksele eşdeğerdir. Bu yüzden, 10 piksel yaklaşık 1.43 karakter genişliğine eşittir.

    # Başlık satırı ve diğer satırlar için yükseklik ayarları
    header_row_height = 143  # Excel'de doğrudan piksel olarak ayarlanabilir
    other_row_height = 66  # Excel'de doğrudan piksel olarak ayarlanabilir

    # Sütun genişliği için pikseli karakter genişliğine çevirme
    column_width_in_pixels = 66
    # Excel'de genişlik karakter cinsindendir, bu yüzden dönüşüm yapılır
    column_width_in_chars = column_width_in_pixels / 7  # Yaklaşık bir dönüşüm

    wb = openpyxl.load_workbook(final_output_path)
    ws = wb.active

    # Biçimlendirme ayarları
    header_fill = PatternFill(start_color='60b516', end_color='60b516', fill_type='solid')  # Lacivert arka plan
    white_font = Font(bold=True, color='FFFFFF')  # Başlık için beyaz, kalın yazı tipi
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                         bottom=Side(style='thin'))  # İnce kenarlık
    centered_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)  # Metni ortala ve kaydır

    # Başlık satırının yüksekliğini ayarla
    ws.row_dimensions[1].height = 143

    # Diğer tüm satırlar için yüksekliği ayarla
    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].height = 35

    # Tüm sütunlar için genişliği ayarla
    for col in ws.iter_cols(min_col=1, max_col=ws.max_column):
        ws.column_dimensions[col[0].column_letter].width = 73 / 7  # Piksel değerini Excel genişliğine çevirme

    # Tüm hücrelere biçimlendirme uygula
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
            cell.alignment = centered_alignment
            if cell.row == 1:  # Başlık satırı için ek biçimlendirme
                cell.fill = header_fill
                cell.font = white_font

    # Yuvarlanacak başlıkların listesi
    headers_to_round = [
        "En Yüksek Çözüm Süresi Saat (Gün)",
        "Ortalama Arıza Süresi Saat (Gün)",
    ]

    # Sütun harflerini ve başlıklarını bul
    columns_to_round = {}
    for idx, col in enumerate(ws.iter_cols(1, ws.max_column, 1, 1, True), start=1):
        if col[0] in headers_to_round:
            columns_to_round[get_column_letter(idx)] = col[0]

    # Her belirtilen sütun için, sayısal hücreleri yuvarla ve formatı '0.0' olarak ayarla
    for col_letter, header in columns_to_round.items():
        for cell in ws[col_letter][1:]:  # Üstbilgiyi atla, 2. satırdan başla
            if isinstance(cell.value, (int, float)):
                # Sayısal değeri yuvarla ve tek ondalık basamakla formatla
                cell.value = round(cell.value, 1)
                cell.number_format = '0.0'

    # Save the workbook with all the changes
    wb.save(final_output_path)

    print_islem_bitir('Abone Değerlendirme', 'Günlük Değerlendirme')
    bio_ses('BS05.mp3')

    ################################## Abone Aylık Değerlendirme İşlemleri ###############################################

    print_islem_baslat('Abone Değerlendirme', 'Aylık Değerlendirme')

    # Veri setini yükleme
    data = pd.read_excel(final_output_path)

    # Sayısal metrikler listesi
    numerical_metrics = [

        'Arıza Yaşanan Gün (Ay)',

        'Ortalama Arıza Adedi (Gün)',
        'En Yüksek Arıza Adedi (Gün)',
        'En Yüksek Arıza Günü (Gün)',

        'Ortalama Tekrar Yüzdesi (Gün)',
        'En Yüksek Çözüm Süresi Saat (Gün)',
        'En Yüksek Çözüm Süresi Günü (Gün)',

        'Ortalama Arıza Süresi Saat (Gün)',

        'Ortalama Yoğunluk Süresi Saati (Gün)',

        'Arıza Tekrar Adet (Gün)',

    ]

    # Kategorik metrikler listesi
    categorical_metrics = [

        'Müsait Olduğu Zaman (Saat)',
        'En Sık Tekrar Eden Takım (Gün)',

        'En Sık Arıza Nedeni (Gün)',
        'En Sık Arıza Açıklaması (Gün)',
        'En Sık Çözüm Açıklaması (Gün)',
        'En Sık Çözüm Nedenleri (Gün)',

        'İş Ortağı',
        'Bölge',
        'FN'

    ]

    # 'Yıl' ve 'Ay' bazında sayısal metrikler için ortalama değerleri hesaplayalım ve sonuçları tam sayı olarak ayarlayalım.
    average_values_by_year_month = data.groupby(["Yıl", "Ay", 'Hizmet No'])[numerical_metrics].mean().reset_index()
    average_values_by_year_month[numerical_metrics] = average_values_by_year_month[numerical_metrics].round(0).astype(
        int)

    # Kategorik metrikler için mod değerlerini hesaplayalım.
    # Kategorik metrikler için 'safe_mode' fonksiyonunu kullanarak hata yönetimini iyileştireceğiz.
    def safe_mode(series):
        if series.empty:
            return None
        mode_result = series.mode()
        return mode_result[0] if not mode_result.empty else None

    # 'Yıl' ve 'Ay' bazında kategorik metrikler için mod değerlerini hesaplayalım.
    mode_values_by_year_month = data.groupby(["Yıl", "Ay", 'Hizmet No']).agg(
        {metric: safe_mode for metric in categorical_metrics}).reset_index()

    # İki DataFrame'i "Yıl" ve "Ay" sütunlarına göre birleştir
    # Bu durumda her iki DataFrame'de de ortak olan "Yıl" ve "Ay" sütunları birleştirilmiş DataFrame'de bir kez gösterilir
    ay_bazli_yillik_degerlendirme = pd.merge(average_values_by_year_month, mode_values_by_year_month,
                                             on=["Yıl", "Ay", 'Hizmet No'])

    # "Genel Değerlendirme" klasörünü kontrol edelim ve yoksa oluşturalım.
    general_evaluation_folder_path = Path('Aylık Değerlendirme')
    general_evaluation_folder_path.mkdir(parents=True, exist_ok=True)

    # "Genel Değerlendirme" klasörü içinde "Rapor" klasörünü kontrol edelim ve yoksa oluşturalım.
    report_folder_path = general_evaluation_folder_path / 'Abone Değerlendirme'
    report_folder_path.mkdir(parents=True, exist_ok=True)


    ay_bazli_yillik_degerlendirme = ay_bazli_yillik_degerlendirme.sort_values(by='Arıza Yaşanan Gün (Ay)',
                                                                              ascending=False)

    # Excel dosyasının kaydedileceği tam yol
    final_output_path = report_folder_path / 'Abone Aylık Değerlendirme Raporu.xlsx'

    ay_bazli_yillik_degerlendirme.to_excel(final_output_path, index=False)

    # Ayarlanacak boyutlar: Başlık satırı yüksekliği 143 piksel, diğer satırlar 66 piksel, tüm sütunların genişliği 10 piksel.
    # Excel'de piksel yerine nokta ve karakter genişliği kullanıldığı için, bu değerler yaklaşık olarak dönüştürülecektir.

    # Pikseli Excel'e uygun birime dönüştürme yaklaşık değerleri
    # Satır yüksekliği doğrudan piksel olarak ayarlanabilir (yaklaşık olarak)
    # Sütun genişliği için Excel genişliği, piksel değerine göre ayarlanmalıdır. Excel genişliği ve piksel arasında doğrudan bir oran yoktur,
    # ancak genel olarak 1 karakter genişliği yaklaşık 7 piksele eşdeğerdir. Bu yüzden, 10 piksel yaklaşık 1.43 karakter genişliğine eşittir.

    # Başlık satırı ve diğer satırlar için yükseklik ayarları
    header_row_height = 143  # Excel'de doğrudan piksel olarak ayarlanabilir
    other_row_height = 66  # Excel'de doğrudan piksel olarak ayarlanabilir

    # Sütun genişliği için pikseli karakter genişliğine çevirme
    column_width_in_pixels = 66
    # Excel'de genişlik karakter cinsindendir, bu yüzden dönüşüm yapılır
    column_width_in_chars = column_width_in_pixels / 7  # Yaklaşık bir dönüşüm

    wb = openpyxl.load_workbook(final_output_path)
    ws = wb.active

    # Biçimlendirme ayarları
    header_fill = PatternFill(start_color='60b516', end_color='60b516', fill_type='solid')  # Lacivert arka plan
    white_font = Font(bold=True, color='FFFFFF')  # Başlık için beyaz, kalın yazı tipi
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                         bottom=Side(style='thin'))  # İnce kenarlık
    centered_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)  # Metni ortala ve kaydır

    # Başlık satırının yüksekliğini ayarla
    ws.row_dimensions[1].height = 143

    # Diğer tüm satırlar için yüksekliği ayarla
    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].height = 35

    # Tüm sütunlar için genişliği ayarla
    for col in ws.iter_cols(min_col=1, max_col=ws.max_column):
        ws.column_dimensions[col[0].column_letter].width = 73 / 7  # Piksel değerini Excel genişliğine çevirme

    # Tüm hücrelere biçimlendirme uygula
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
            cell.alignment = centered_alignment
            if cell.row == 1:  # Başlık satırı için ek biçimlendirme
                cell.fill = header_fill
                cell.font = white_font

    # Yuvarlanacak başlıkların listesi
    headers_to_round = [
        "En Yüksek Çözüm Süresi Saat (Gün)",
    ]

    # Sütun harflerini ve başlıklarını bul
    columns_to_round = {}
    for idx, col in enumerate(ws.iter_cols(1, ws.max_column, 1, 1, True), start=1):
        if col[0] in headers_to_round:
            columns_to_round[get_column_letter(idx)] = col[0]

    # Her belirtilen sütun için, sayısal hücreleri yuvarla ve formatı '0.0' olarak ayarla
    for col_letter, header in columns_to_round.items():
        for cell in ws[col_letter][1:]:  # Üstbilgiyi atla, 2. satırdan başla
            if isinstance(cell.value, (int, float)):
                # Sayısal değeri yuvarla ve tek ondalık basamakla formatla
                cell.value = round(cell.value, 1)
                cell.number_format = '0.0'

    # Save the workbook with all the changes
    wb.save(final_output_path)

    print_islem_bitir('Abone Değerlendirme', 'Aylık Değerlendirme')
    bio_ses('BS05.mp3')

    ################################## Abone Yıllık Değerlendirme İşlemleri ###############################################

    print_islem_baslat('Abone Değerlendirme', 'Yıllık Değerlendirme')

    # Veri setini yükleme
    data = pd.read_excel(final_output_path)

    # Sayısal metrikler listesi
    numerical_metrics = [

        'Arıza Yaşanan Gün (Ay)',

        'Ortalama Arıza Adedi (Gün)',
        'En Yüksek Arıza Adedi (Gün)',
        'En Yüksek Arıza Günü (Gün)',

        'Ortalama Tekrar Yüzdesi (Gün)',
        'En Yüksek Çözüm Süresi Saat (Gün)',
        'En Yüksek Çözüm Süresi Günü (Gün)',

        'Ortalama Arıza Süresi Saat (Gün)',

        'Ortalama Yoğunluk Süresi Saati (Gün)',

        'Arıza Tekrar Adet (Gün)',

    ]

    # Kategorik metrikler listesi
    categorical_metrics = [

        'Müsait Olduğu Zaman (Saat)',
        'En Sık Tekrar Eden Takım (Gün)',

        'En Sık Arıza Nedeni (Gün)',
        'En Sık Arıza Açıklaması (Gün)',
        'En Sık Çözüm Açıklaması (Gün)',
        'En Sık Çözüm Nedenleri (Gün)',

        'İş Ortağı',
        'Bölge',
        'FN'

    ]
    # 'Yıl' ve 'Ay' bazında sayısal metrikler için ortalama değerleri hesaplayalım ve sonuçları tam sayı olarak ayarlayalım.
    average_values_by_year_month = data.groupby(["Yıl", 'Hizmet No'])[numerical_metrics].mean().reset_index()
    average_values_by_year_month[numerical_metrics] = average_values_by_year_month[numerical_metrics].round(0).astype(
        int)

    # Kategorik metrikler için mod değerlerini hesaplayalım.
    # Kategorik metrikler için 'safe_mode' fonksiyonunu kullanarak hata yönetimini iyileştireceğiz.
    def safe_mode(series):
        if series.empty:
            return None
        mode_result = series.mode()
        return mode_result[0] if not mode_result.empty else None

    # 'Yıl' ve 'Ay' bazında kategorik metrikler için mod değerlerini hesaplayalım.
    mode_values_by_year_month = data.groupby(["Yıl", 'Hizmet No']).agg(
        {metric: safe_mode for metric in categorical_metrics}).reset_index()

    # İki DataFrame'i "Yıl" ve "Ay" sütunlarına göre birleştir
    # Bu durumda her iki DataFrame'de de ortak olan "Yıl" ve "Ay" sütunları birleştirilmiş DataFrame'de bir kez gösterilir
    yillik_bazli_genel_degerlendirme = pd.merge(average_values_by_year_month, mode_values_by_year_month,
                                                on=["Yıl", 'Hizmet No'])

    # "Genel Değerlendirme" klasörünü kontrol edelim ve yoksa oluşturalım.
    general_evaluation_folder_path = Path('Yıllık Değerlendirme')
    general_evaluation_folder_path.mkdir(parents=True, exist_ok=True)

    # "Genel Değerlendirme" klasörü içinde "Rapor" klasörünü kontrol edelim ve yoksa oluşturalım.
    report_folder_path = general_evaluation_folder_path / 'Abone Değerlendirme'
    report_folder_path.mkdir(parents=True, exist_ok=True)


    yillik_bazli_genel_degerlendirme = yillik_bazli_genel_degerlendirme.sort_values(by='Arıza Yaşanan Gün (Ay)',
                                                                                    ascending=False)

    # Excel dosyasının kaydedileceği tam yol
    final_output_path = report_folder_path / 'Abone Yıllık Değerlendirme.xlsx'

    yillik_bazli_genel_degerlendirme.to_excel(final_output_path, index=False)

    # Ayarlanacak boyutlar: Başlık satırı yüksekliği 143 piksel, diğer satırlar 66 piksel, tüm sütunların genişliği 10 piksel.
    # Excel'de piksel yerine nokta ve karakter genişliği kullanıldığı için, bu değerler yaklaşık olarak dönüştürülecektir.

    # Pikseli Excel'e uygun birime dönüştürme yaklaşık değerleri
    # Satır yüksekliği doğrudan piksel olarak ayarlanabilir (yaklaşık olarak)
    # Sütun genişliği için Excel genişliği, piksel değerine göre ayarlanmalıdır. Excel genişliği ve piksel arasında doğrudan bir oran yoktur,
    # ancak genel olarak 1 karakter genişliği yaklaşık 7 piksele eşdeğerdir. Bu yüzden, 10 piksel yaklaşık 1.43 karakter genişliğine eşittir.

    # Başlık satırı ve diğer satırlar için yükseklik ayarları
    header_row_height = 143  # Excel'de doğrudan piksel olarak ayarlanabilir
    other_row_height = 66  # Excel'de doğrudan piksel olarak ayarlanabilir

    # Sütun genişliği için pikseli karakter genişliğine çevirme
    column_width_in_pixels = 66
    # Excel'de genişlik karakter cinsindendir, bu yüzden dönüşüm yapılır
    column_width_in_chars = column_width_in_pixels / 7  # Yaklaşık bir dönüşüm

    wb = openpyxl.load_workbook(final_output_path)
    ws = wb.active

    # Biçimlendirme ayarları
    header_fill = PatternFill(start_color='60b516', end_color='60b516', fill_type='solid')  # Lacivert arka plan
    white_font = Font(bold=True, color='FFFFFF')  # Başlık için beyaz, kalın yazı tipi
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                         bottom=Side(style='thin'))  # İnce kenarlık
    centered_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)  # Metni ortala ve kaydır

    # Başlık satırının yüksekliğini ayarla
    ws.row_dimensions[1].height = 143

    # Diğer tüm satırlar için yüksekliği ayarla
    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].height = 35

    # Tüm sütunlar için genişliği ayarla
    for col in ws.iter_cols(min_col=1, max_col=ws.max_column):
        ws.column_dimensions[col[0].column_letter].width = 73 / 7  # Piksel değerini Excel genişliğine çevirme

    # Tüm hücrelere biçimlendirme uygula
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
            cell.alignment = centered_alignment
            if cell.row == 1:  # Başlık satırı için ek biçimlendirme
                cell.fill = header_fill
                cell.font = white_font

    # Yuvarlanacak başlıkların listesi
    headers_to_round = [
        "En Yüksek Çözüm Süresi Saat (Gün)",
    ]

    # Sütun harflerini ve başlıklarını bul
    columns_to_round = {}
    for idx, col in enumerate(ws.iter_cols(1, ws.max_column, 1, 1, True), start=1):
        if col[0] in headers_to_round:
            columns_to_round[get_column_letter(idx)] = col[0]

    # Her belirtilen sütun için, sayısal hücreleri yuvarla ve formatı '0.0' olarak ayarla
    for col_letter, header in columns_to_round.items():
        for cell in ws[col_letter][1:]:  # Üstbilgiyi atla, 2. satırdan başla
            if isinstance(cell.value, (int, float)):
                # Sayısal değeri yuvarla ve tek ondalık basamakla formatla
                cell.value = round(cell.value, 1)
                cell.number_format = '0.0'

    # Save the workbook with all the changes
    wb.save(final_output_path)

    print_islem_bitir('Abone Değerlendirme', 'Yıllık Değerlendirme')
    bio_ses('BS05.mp3')

    ################################## Abone Genel Değerlendirme İşlemleri ###############################################

    print_islem_baslat('Abone Değerlendirme', 'Genel Değerlendirme')

    data = pd.read_excel(final_output_path)

    # Sayısal metrikler listesi
    numerical_metrics = [

        'Arıza Yaşanan Gün (Ay)',

        'Ortalama Arıza Adedi (Gün)',
        'En Yüksek Arıza Adedi (Gün)',
        'En Yüksek Arıza Günü (Gün)',

        'Ortalama Tekrar Yüzdesi (Gün)',
        'En Yüksek Çözüm Süresi Saat (Gün)',
        'En Yüksek Çözüm Süresi Günü (Gün)',

        'Ortalama Arıza Süresi Saat (Gün)',

        'Ortalama Yoğunluk Süresi Saati (Gün)',

        'Arıza Tekrar Adet (Gün)'

    ]

    # Kategorik metrikler listesi
    categorical_metrics = [

        'Müsait Olduğu Zaman (Saat)',
        'En Sık Tekrar Eden Takım (Gün)',

        'En Sık Arıza Nedeni (Gün)',
        'En Sık Arıza Açıklaması (Gün)',
        'En Sık Çözüm Açıklaması (Gün)',
        'En Sık Çözüm Nedenleri (Gün)',

        'İş Ortağı',
        'Bölge',
        'FN'

    ]

    # 'Yıl' ve 'Ay' bazında sayısal metrikler için ortalama değerleri hesaplayalım ve sonuçları tam sayı olarak ayarlayalım.
    average_values_by_year_month = data.groupby(['Hizmet No'])[numerical_metrics].mean().reset_index()
    average_values_by_year_month[numerical_metrics] = average_values_by_year_month[numerical_metrics].round(0).astype(
        int)

    # Kategorik metrikler için mod değerlerini hesaplayalım.
    # Kategorik metrikler için 'safe_mode' fonksiyonunu kullanarak hata yönetimini iyileştireceğiz.
    def safe_mode(series):
        if series.empty:
            return None
        mode_result = series.mode()
        return mode_result[0] if not mode_result.empty else None

    # 'Yıl' ve 'Ay' bazında kategorik metrikler için mod değerlerini hesaplayalım.
    mode_values_by_year_month = data.groupby(['Hizmet No']).agg(
        {metric: safe_mode for metric in categorical_metrics}).reset_index()

    # İki DataFrame'i "Yıl" ve "Ay" sütunlarına göre birleştir
    # Bu durumda her iki DataFrame'de de ortak olan "Yıl" ve "Ay" sütunları birleştirilmiş DataFrame'de bir kez gösterilir
    genel_degerlendirme = pd.merge(average_values_by_year_month, mode_values_by_year_month, on=['Hizmet No'])

    # "Genel Değerlendirme" klasörünü kontrol edelim ve yoksa oluşturalım.
    general_evaluation_folder_path = Path('Genel Değerlendirme')
    general_evaluation_folder_path.mkdir(parents=True, exist_ok=True)

    # "Genel Değerlendirme" klasörü içinde "Rapor" klasörünü kontrol edelim ve yoksa oluşturalım.
    report_folder_path = general_evaluation_folder_path / 'Abone Değerlendirme'
    report_folder_path.mkdir(parents=True, exist_ok=True)


    genel_degerlendirme = genel_degerlendirme.sort_values(by='Arıza Yaşanan Gün (Ay)', ascending=False)

    # Excel dosyasının kaydedileceği tam yol
    final_output_path = report_folder_path / 'Abone Genel Değerlendirme Raporu.xlsx'

    genel_degerlendirme.to_excel(final_output_path, index=False)

    # Ayarlanacak boyutlar: Başlık satırı yüksekliği 143 piksel, diğer satırlar 66 piksel, tüm sütunların genişliği 10 piksel.
    # Excel'de piksel yerine nokta ve karakter genişliği kullanıldığı için, bu değerler yaklaşık olarak dönüştürülecektir.

    # Pikseli Excel'e uygun birime dönüştürme yaklaşık değerleri
    # Satır yüksekliği doğrudan piksel olarak ayarlanabilir (yaklaşık olarak)
    # Sütun genişliği için Excel genişliği, piksel değerine göre ayarlanmalıdır. Excel genişliği ve piksel arasında doğrudan bir oran yoktur,
    # ancak genel olarak 1 karakter genişliği yaklaşık 7 piksele eşdeğerdir. Bu yüzden, 10 piksel yaklaşık 1.43 karakter genişliğine eşittir.

    # Başlık satırı ve diğer satırlar için yükseklik ayarları
    header_row_height = 143  # Excel'de doğrudan piksel olarak ayarlanabilir
    other_row_height = 66  # Excel'de doğrudan piksel olarak ayarlanabilir

    # Sütun genişliği için pikseli karakter genişliğine çevirme
    column_width_in_pixels = 66
    # Excel'de genişlik karakter cinsindendir, bu yüzden dönüşüm yapılır
    column_width_in_chars = column_width_in_pixels / 7  # Yaklaşık bir dönüşüm

    wb = openpyxl.load_workbook(final_output_path)
    ws = wb.active

    # Biçimlendirme ayarları
    header_fill = PatternFill(start_color='60b516', end_color='60b516', fill_type='solid')  # Lacivert arka plan
    white_font = Font(bold=True, color='FFFFFF')  # Başlık için beyaz, kalın yazı tipi
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                         bottom=Side(style='thin'))  # İnce kenarlık
    centered_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)  # Metni ortala ve kaydır

    # Başlık satırının yüksekliğini ayarla
    ws.row_dimensions[1].height = 143

    # Diğer tüm satırlar için yüksekliği ayarla
    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].height = 35

    # Tüm sütunlar için genişliği ayarla
    for col in ws.iter_cols(min_col=1, max_col=ws.max_column):
        ws.column_dimensions[col[0].column_letter].width = 73 / 7  # Piksel değerini Excel genişliğine çevirme

    # Tüm hücrelere biçimlendirme uygula
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
            cell.alignment = centered_alignment
            if cell.row == 1:  # Başlık satırı için ek biçimlendirme
                cell.fill = header_fill
                cell.font = white_font

    # Yuvarlanacak başlıkların listesi
    headers_to_round = [
        "En Yüksek Çözüm Süresi Saat (Gün)",
    ]

    # Sütun harflerini ve başlıklarını bul
    columns_to_round = {}
    for idx, col in enumerate(ws.iter_cols(1, ws.max_column, 1, 1, True), start=1):
        if col[0] in headers_to_round:
            columns_to_round[get_column_letter(idx)] = col[0]

    # Her belirtilen sütun için, sayısal hücreleri yuvarla ve formatı '0.0' olarak ayarla
    for col_letter, header in columns_to_round.items():
        for cell in ws[col_letter][1:]:  # Üstbilgiyi atla, 2. satırdan başla
            if isinstance(cell.value, (int, float)):
                # Sayısal değeri yuvarla ve tek ondalık basamakla formatla
                cell.value = round(cell.value, 1)
                cell.number_format = '0.0'

    # Save the workbook with all the changes
    wb.save(final_output_path)

    print_islem_bitir('Abone Değerlendirme', 'Genel Değerlendirme')

    ################################## Abone Tüm İşlemleri Tamamlandı ###############################################

    print_islem_bitir('Abone Değerlendirme', 'Tüm Değerlendirme')
    bio_ses('BS05.mp3')

def Bölge_Degerlendirme(data_original, BioT):
    ################################## Bölge Veri Optimizasyon İşlemleri ###############################################

    print('')

    print_islem_baslat('Bölge Değerlendirme', 'Veri Optimizasyon')

    data = data_original

    # Yeni ayrıştırma fonksiyonunu uygulama
    data[['İş Ortağı', 'Bölge', 'FN', 'Anfi', 'Tap-Off', 'Kol']] = pd.DataFrame(
        data['Devre No'].apply(ayir_genis).tolist(), index=data.index)

    # 'Yıl', 'Ay', ve 'Gün' sütunlarını tekrar ekleyelim
    data['Yıl'] = data['Çözüm Tarihi'].dt.year
    data['Ay'] = data['Çözüm Tarihi'].dt.month
    data['Gün'] = data['Çözüm Tarihi'].dt.day

    # Günlük arıza sayılarını hesaplama
    data['FN Arıza Adet (Gün)'] = data.groupby(['Yıl', 'Ay', 'Gün', 'İş Ortağı', 'Bölge', 'FN'])['Çağrı No'].transform(
        'size')
    data['Anfi Arıza Adet (Gün)'] = data.groupby(['Yıl', 'Ay', 'Gün', 'İş Ortağı', 'Bölge', 'FN', 'Anfi'])[
        'Çağrı No'].transform('size')
    data['Tap-Off Arıza Adet (Gün)'] = \
    data.groupby(['Yıl', 'Ay', 'Gün', 'İş Ortağı', 'Bölge', 'FN', 'Anfi', 'Tap-Off'])['Çağrı No'].transform('size')
    data['Kol Arıza Adet (Gün)'] = \
    data.groupby(['Yıl', 'Ay', 'Gün', 'İş Ortağı', 'Bölge', 'FN', 'Anfi', 'Tap-Off', 'Kol'])['Çağrı No'].transform(
        'size')

    # Her ekip ve her gün için başlangıç ve bitiş zamanlarını bulma
    data['Baslangic_Zamani'] = data.groupby(['Bölge', 'FN', data['Kayıt Tarihi'].dt.date])['Kayıt Tarihi'].transform(
        'min')
    data['Bitis_Zamani'] = data.groupby(['Bölge', 'FN', data['Çözüm Tarihi'].dt.date])['Çözüm Tarihi'].transform('max')

    # 'Kayıt Tarihi' ve 'Çözüm Tarihi' sütunlarını datetime türüne dönüştürme
    data['Kayıt Tarihi'] = pd.to_datetime(data['Kayıt Tarihi'], format='%d/%m/%Y %H:%M:%S')
    data['Çözüm Tarihi'] = pd.to_datetime(data['Çözüm Tarihi'], format='%d/%m/%Y %H:%M:%S')

    # Arıza süresini saat cinsinden hesaplama
    data['Arıza Süresi Saat (Gün)'] = (data['Çözüm Tarihi'] - data['Kayıt Tarihi']).dt.total_seconds() / 3600

    # Çalışma saat aralığını hesaplayalım ve yeni bir sütun olarak ekleyelim
    data['Arıza Zamanı Saat (Gün)'] = data.apply(lambda x:
                                                 f"{x['Baslangic_Zamani'].strftime('%H:%M') if not pd.isnull(x['Baslangic_Zamani']) else 'NaN'} - {x['Bitis_Zamani'].strftime('%H:%M') if not pd.isnull(x['Bitis_Zamani']) else 'NaN'}",
                                                 axis=1)

    # Her abone ve her gün için toplam arıza süresini hesaplama
    data['Kayıt Tarihi (Gün)'] = data['Kayıt Tarihi'].dt.date
    toplam_ariza_suresi = data.groupby(['Bölge', 'FN', 'Kayıt Tarihi (Gün)'])[
        'Arıza Süresi Saat (Gün)'].sum().reset_index(name='Toplam Arıza Süresi Saat (Gün)')

    def tek_karar_ver(row):
        # Arıza sayılarını al
        fn = row['FN Arıza Adet (Gün)']
        anfi = row['Anfi Arıza Adet (Gün)']
        tap_off = row['Tap-Off Arıza Adet (Gün)']
        kol = row['Kol Arıza Adet (Gün)']

        # FN ve Kol arasındaki farkı kontrol et
        if abs(fn - kol) <= 1:
            return "Kol Arızası!"
        # FN ve Tap-Off arasındaki farkı kontrol et
        elif abs(fn - tap_off) <= 1:
            return "Tap-Off Arızası!"
        # FN ve Anfi arasındaki farkı kontrol et
        elif abs(fn - anfi) <= 1:
            return "Anfi Arızası!"
        else:
            return "FN Arızası!"

    # Veri setinize bu fonksiyonu uygulayın
    data['Karar (Gün)'] = data.apply(tek_karar_ver, axis=1)

    def karara_gore_bolge_numarasi(row):
        # Karar (Gün) sütunundaki değere göre Bölge Numarası (Gün) sütununu doldur
        if row['Karar (Gün)'] == 'Anfi Arızası!':
            return f"{row['İş Ortağı']}.{row['Bölge']}.{row['FN']}.{row['Anfi']}"
        elif row['Karar (Gün)'] == 'FN Arızası!':
            return f"{row['İş Ortağı']}.{row['Bölge']}.{row['FN']}"
        elif row['Karar (Gün)'] == 'Tap-Off Arızası!':
            return f"{row['İş Ortağı']}.{row['Bölge']}.{row['FN']}.{row['Anfi']}.{row['Tap-Off']}"
        elif row['Karar (Gün)'] == 'Kol Arızası!':
            return f"{row['İş Ortağı']}.{row['Bölge']}.{row['FN']}.{row['Anfi']}.{row['Tap-Off']}.{row['Kol']}"
        else:
            return "Tespit edilemedi veya diğer"

    # Yeni Bölge Numarası (Gün) sütunu için fonksiyonu uygula
    data['Bölge No (Gün)'] = data.apply(karara_gore_bolge_numarasi, axis=1)

    fn_en_sik_ariza_nedenleri = data.groupby(['Bölge', 'FN'])['Açıklama'].agg(
        lambda x: x.mode().iloc[0] if not x.mode().empty else np.nan).reset_index(name='FN En Sık Arıza Nedeni (Gün)')
    fn_en_sik_ariza_cozumleri = data.groupby(['Bölge', 'FN'])['Çözüm Açıklaması'].agg(
        lambda x: x.mode().iloc[0] if not x.mode().empty else np.nan).reset_index(name='FN En Sık Arıza Çözümü (Gün)')

    anfi_en_sik_ariza_nedenleri = data.groupby(['Bölge', 'FN', 'Anfi'])['Açıklama'].agg(
        lambda x: x.mode().iloc[0] if not x.mode().empty else np.nan).reset_index(name='Anfi En Sık Arıza Nedeni (Gün)')
    anfi_en_sik_ariza_cozumleri = data.groupby(['Bölge', 'FN', 'Anfi'])['Çözüm Açıklaması'].agg(
        lambda x: x.mode().iloc[0] if not x.mode().empty else np.nan).reset_index(name='Anfi En Sık Arıza Çözümü (Gün)')

    tapoff_en_sik_ariza_nedenleri = data.groupby(['Bölge', 'FN', 'Anfi', 'Tap-Off'])['Açıklama'].agg(
        lambda x: x.mode().iloc[0] if not x.mode().empty else np.nan).reset_index(
        name='Tap-off En Sık Arıza Nedeni (Gün)')
    tapoff_en_sik_ariza_cozumleri = data.groupby(['Bölge', 'FN', 'Anfi', 'Tap-Off'])['Çözüm Açıklaması'].agg(
        lambda x: x.mode().iloc[0] if not x.mode().empty else np.nan).reset_index(
        name='Tap-off En Sık Arıza Çözümü (Gün)')

    kol_en_sik_ariza_nedenleri = data.groupby(['Bölge', 'FN', 'Anfi', 'Tap-Off', 'Kol'])['Açıklama'].agg(
        lambda x: x.mode().iloc[0] if not x.mode().empty else np.nan).reset_index(name='Kol En Sık Arıza Nedeni (Gün)')
    kol_en_sik_ariza_cozumleri = data.groupby(['Bölge', 'FN', 'Anfi', 'Tap-Off', 'Kol'])['Çözüm Açıklaması'].agg(
        lambda x: x.mode().iloc[0] if not x.mode().empty else np.nan).reset_index(name='Kol En Sık Arıza Çözümü (Gün)')

    # Bölge ve FN Kombinasyonundan Gelen Arıza Sayısı için birleştirme
    data = data.merge(fn_en_sik_ariza_nedenleri, on=['Bölge', 'FN'], how='left')
    data = data.merge(fn_en_sik_ariza_cozumleri, on=['Bölge', 'FN'], how='left')

    # Bölge, FN ve Anfi Kombinasyonundan Gelen Arıza Sayısı için birleştirme
    data = data.merge(anfi_en_sik_ariza_nedenleri, on=['Bölge', 'FN', 'Anfi'], how='left')
    data = data.merge(anfi_en_sik_ariza_cozumleri, on=['Bölge', 'FN', 'Anfi'], how='left')

    # Bölge, FN, Anfi ve Tap-Off Kombinasyonundan Gelen Arıza Sayısı için birleştirme
    data = data.merge(tapoff_en_sik_ariza_nedenleri, on=['Bölge', 'FN', 'Anfi', 'Tap-Off'], how='left')
    data = data.merge(tapoff_en_sik_ariza_cozumleri, on=['Bölge', 'FN', 'Anfi', 'Tap-Off'], how='left')

    # Bölge, FN, Anfi, Tap-Off ve Kol Kombinasyonundan Gelen Arıza Sayısı için birleştirme
    data = data.merge(kol_en_sik_ariza_nedenleri, on=['Bölge', 'FN', 'Anfi', 'Tap-Off', 'Kol'], how='left')
    data = data.merge(kol_en_sik_ariza_cozumleri, on=['Bölge', 'FN', 'Anfi', 'Tap-Off', 'Kol'], how='left')

    # 'Bölge' ve 'FN' kombinasyonuna göre gruplandırıp, her grubun en sık tekrar eden 'Hizmet No'sunu bulma
    data['En Sık Tekrar Eden Hizmet No (Gün)'] = data.groupby(['Bölge', 'FN'])['Hizmet No'].transform(
        lambda x: x.mode()[0])

    print_islem_bitir('Bölge Değerlendirme', 'Veri Optimizasyon')
    bio_ses('BS05.mp3')

    ################################## Bölge Günlük Detaylı Değerlendirme İşlemleri ###############################################

    print_islem_baslat('Bölge Değerlendirme', 'Günlük Detaylı Değerlendirme')

    # Sadece hesaplanan ve eklenen sütunları içeren yeni bir DataFrame oluşturalım
    data = data[['Yıl', 'Ay', 'Gün', 'İş Ortağı', 'Bölge', 'FN', 'Anfi', 'Tap-Off', 'Kol',
                 'FN Arıza Adet (Gün)', 'Anfi Arıza Adet (Gün)', 'Tap-Off Arıza Adet (Gün)', 'Kol Arıza Adet (Gün)',
                 'Karar (Gün)', 'Bölge No (Gün)', 'FN En Sık Arıza Nedeni (Gün)', 'FN En Sık Arıza Çözümü (Gün)',
                 'Anfi En Sık Arıza Nedeni (Gün)', 'Anfi En Sık Arıza Çözümü (Gün)',
                 'Tap-off En Sık Arıza Nedeni (Gün)', 'Tap-off En Sık Arıza Çözümü (Gün)',
                 'Kol En Sık Arıza Nedeni (Gün)', 'Kol En Sık Arıza Çözümü (Gün)', 'En Sık Tekrar Eden Hizmet No (Gün)',
                 'Arıza Süresi Saat (Gün)', 'Arıza Zamanı Saat (Gün)']].drop_duplicates()

    # "Genel Değerlendirme" klasörünü kontrol edelim ve yoksa oluşturalım.
    general_evaluation_folder_path = Path('Günlük Değerlendirme')
    general_evaluation_folder_path.mkdir(parents=True, exist_ok=True)

    # "Genel Değerlendirme" klasörü içinde "Rapor" klasörünü kontrol edelim ve yoksa oluşturalım.
    report_folder_path = general_evaluation_folder_path / 'Bölge Değerlendirme'
    report_folder_path.mkdir(parents=True, exist_ok=True)

    # Günlük_Arıza sütununa göre veriyi büyükten küçüğe sıralayın
    data = data.sort_values(by='FN Arıza Adet (Gün)', ascending=False)

    # Excel dosyasının kaydedileceği tam yol
    final_output_path = report_folder_path / 'Bölge Günlük Detaylı Değerlendirme Raporu.xlsx'

    data.to_excel(final_output_path, index=False)

    # Ayarlanacak boyutlar: Başlık satırı yüksekliği 143 piksel, diğer satırlar 66 piksel, tüm sütunların genişliği 10 piksel.
    # Excel'de piksel yerine nokta ve karakter genişliği kullanıldığı için, bu değerler yaklaşık olarak dönüştürülecektir.

    # Pikseli Excel'e uygun birime dönüştürme yaklaşık değerleri
    # Satır yüksekliği doğrudan piksel olarak ayarlanabilir (yaklaşık olarak)
    # Sütun genişliği için Excel genişliği, piksel değerine göre ayarlanmalıdır. Excel genişliği ve piksel arasında doğrudan bir oran yoktur,
    # ancak genel olarak 1 karakter genişliği yaklaşık 7 piksele eşdeğerdir. Bu yüzden, 10 piksel yaklaşık 1.43 karakter genişliğine eşittir.

    # Başlık satırı ve diğer satırlar için yükseklik ayarları
    header_row_height = 143  # Excel'de doğrudan piksel olarak ayarlanabilir
    other_row_height = 66  # Excel'de doğrudan piksel olarak ayarlanabilir

    # Sütun genişliği için pikseli karakter genişliğine çevirme
    column_width_in_pixels = 66
    # Excel'de genişlik karakter cinsindendir, bu yüzden dönüşüm yapılır
    column_width_in_chars = column_width_in_pixels / 7  # Yaklaşık bir dönüşüm

    wb = openpyxl.load_workbook(final_output_path)
    ws = wb.active

    # Biçimlendirme ayarları
    header_fill = PatternFill(start_color='FFE67300', end_color='FFE67300', fill_type='solid')  # Lacivert arka plan
    white_font = Font(bold=True, color='FFFFFF')  # Başlık için beyaz, kalın yazı tipi
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                         bottom=Side(style='thin'))  # İnce kenarlık
    centered_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)  # Metni ortala ve kaydır

    # Başlık satırının yüksekliğini ayarla
    ws.row_dimensions[1].height = 143

    # Diğer tüm satırlar için yüksekliği ayarla
    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].height = 35

    # Tüm sütunlar için genişliği ayarla
    for col in ws.iter_cols(min_col=1, max_col=ws.max_column):
        ws.column_dimensions[col[0].column_letter].width = 73 / 7  # Piksel değerini Excel genişliğine çevirme

    # Tüm hücrelere biçimlendirme uygula
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
            cell.alignment = centered_alignment
            if cell.row == 1:  # Başlık satırı için ek biçimlendirme
                cell.fill = header_fill
                cell.font = white_font

    # Save the workbook with all the changes
    wb.save(final_output_path)

    ################################## Bölge Günlük Değerlendirme İşlemleri ###############################################

    print_islem_baslat('Bölge Değerlendirme', 'Günlük Değerlendirme')

    data = pd.read_excel(final_output_path)


    def en_sik_tekrar(series):
        modes = series.mode()
        if modes.empty:
            return None  # ya da uygun bir değer, örneğin np.nan
        else:
            return modes[0]

    # Yıl, Ay, Gün sütunlarını Nullable Integer türüne dönüştür
    data['Yıl'] = data['Yıl'].astype('Int64')
    data['Ay'] = data['Ay'].astype('Int64')
    data['Gün'] = data['Gün'].astype('Int64')

    # Öncelikle, DataFrame'i gruplandırma ve ortalama alma
    bolge_daily_grouped = data.groupby(['Yıl', 'Ay', 'Gün', 'İş Ortağı', 'Bölge', 'FN']).agg(
        FN_Arıza_Adet=('FN Arıza Adet (Gün)', 'mean'),
        Anfi_Arıza_Adet=('Anfi Arıza Adet (Gün)', 'mean'),
        Tap_off_Arıza_Adet=('Tap-Off Arıza Adet (Gün)', 'mean'),
        Kol_Arıza_Adet=('Kol Arıza Adet (Gün)', 'mean'),
        Karar=('Karar (Gün)', en_sik_tekrar),
        Bölge_No=('Bölge No (Gün)', en_sik_tekrar),
        FN_En_Sık_Arıza_Nedeni=('FN En Sık Arıza Nedeni (Gün)', en_sik_tekrar),
        FN_En_Sık_Arıza_Çözümü=('FN En Sık Arıza Çözümü (Gün)', en_sik_tekrar),
        Anfi_En_Sık_Arıza_Nedeni=('Anfi En Sık Arıza Nedeni (Gün)', en_sik_tekrar),
        Anfi_En_Sık_Arıza_Çözümü=('Anfi En Sık Arıza Çözümü (Gün)', en_sik_tekrar),
        Tap_off_En_Sık_Arıza_Nedeni=('Tap-off En Sık Arıza Nedeni (Gün)', en_sik_tekrar),
        Tap_off_En_Sık_Arıza_Çözümü=('Tap-off En Sık Arıza Çözümü (Gün)', en_sik_tekrar),
        Kol_En_Sık_Arıza_Nedeni=('Kol En Sık Arıza Nedeni (Gün)', en_sik_tekrar),
        Kol_En_Sık_Arıza_Çözümü=('Kol En Sık Arıza Çözümü (Gün)', en_sik_tekrar),
        En_Sık_Tekrar_Eden_Hizmet_No=('En Sık Tekrar Eden Hizmet No (Gün)', en_sik_tekrar),
        Arıza_süresi_saat=('Arıza Süresi Saat (Gün)', 'mean'),
        Arıza_zamanı_saat=('Arıza Zamanı Saat (Gün)', en_sik_tekrar),

    ).reset_index()

    # Sütun isimlerini yeniden adlandırma
    bolge_daily_grouped = bolge_daily_grouped.rename(columns={
        'FN_Arıza_Adet': 'FN Arıza Adet (Gün)',
        'Anfi_Arıza_Adet': 'Anfi Arıza Adet (Gün)',
        'Tap_off_Arıza_Adet': 'Tap-Off Arıza Adet (Gün)',
        'Kol_Arıza_Adet': 'Kol Arıza Adet (Gün)',
        'Karar': 'Karar (Gün)',
        'Bölge_No': 'Bölge No (Gün)',
        'FN_En_Sık_Arıza_Nedeni': 'FN En Sık Arıza Nedeni (Gün)',
        'FN_En_Sık_Arıza_Çözümü': 'FN En Sık Arıza Çözümü (Gün)',
        'Anfi_En_Sık_Arıza_Nedeni': 'Anfi En Sık Arıza Nedeni (Gün)',
        'Anfi_En_Sık_Arıza_Çözümü': 'Anfi En Sık Arıza Çözümü (Gün)',
        'Tap_off_En_Sık_Arıza_Nedeni': 'Tap-off En Sık Arıza Nedeni (Gün)',
        'Tap_off_En_Sık_Arıza_Çözümü': 'Tap-off En Sık Arıza Çözümü (Gün)',
        'Kol_En_Sık_Arıza_Nedeni': 'Kol En Sık Arıza Nedeni (Gün)',
        'Kol_En_Sık_Arıza_Çözümü': 'Kol En Sık Arıza Çözümü (Gün)',
        'En_Sık_Tekrar_Eden_Hizmet_No': 'En Sık Tekrar Eden Hizmet No (Gün)',
        'Arıza_süresi_saat': 'En Sık Arıza Süresi Saat (Gün)',
        'Arıza_zamanı_saat': 'En Sık Arıza Zamanı Saat (Gün)'
    })

    # Sadece hesaplanan ve eklenen sütunları içeren yeni bir DataFrame oluşturalım
    bolge_daily_grouped = bolge_daily_grouped[['Yıl', 'Ay', 'Gün', 'İş Ortağı', 'Bölge', 'FN',
                                               'FN Arıza Adet (Gün)', 'Anfi Arıza Adet (Gün)',
                                               'Tap-Off Arıza Adet (Gün)', 'Kol Arıza Adet (Gün)', 'Karar (Gün)',
                                               'Bölge No (Gün)', 'FN En Sık Arıza Nedeni (Gün)',
                                               'FN En Sık Arıza Çözümü (Gün)', 'Anfi En Sık Arıza Nedeni (Gün)',
                                               'Anfi En Sık Arıza Çözümü (Gün)', 'Tap-off En Sık Arıza Nedeni (Gün)',
                                               'Tap-off En Sık Arıza Çözümü (Gün)', 'Kol En Sık Arıza Nedeni (Gün)',
                                               'Kol En Sık Arıza Çözümü (Gün)', 'En Sık Tekrar Eden Hizmet No (Gün)',
                                               'En Sık Arıza Süresi Saat (Gün)',
                                               'En Sık Arıza Zamanı Saat (Gün)']].drop_duplicates()

    # "Genel Değerlendirme" klasörünü kontrol edelim ve yoksa oluşturalım.
    general_evaluation_folder_path = Path('Günlük Değerlendirme')
    general_evaluation_folder_path.mkdir(parents=True, exist_ok=True)

    # "Genel Değerlendirme" klasörü içinde "Rapor" klasörünü kontrol edelim ve yoksa oluşturalım.
    report_folder_path = general_evaluation_folder_path / 'Bölge Değerlendirme'
    report_folder_path.mkdir(parents=True, exist_ok=True)

    # "Genel Değerlendirme" klasörü içinde "Rapor" klasörünü kontrol edelim ve yoksa oluşturalım.
    takım_dosya = report_folder_path / 'Bölge Değerlendirme Raporu'
    takım_dosya.mkdir(parents=True, exist_ok=True)

    # Günlük_Arıza sütununa göre veriyi büyükten küçüğe sıralayın
    bolge_daily_grouped = bolge_daily_grouped.sort_values(by='FN Arıza Adet (Gün)', ascending=False)

    # Excel dosyasının kaydedileceği tam yol
    final_output_path = takım_dosya / 'Bölge Günlük Değerlendirme Raporu.xlsx'

    bolge_daily_grouped.to_excel(final_output_path, index=False)

    # Ayarlanacak boyutlar: Başlık satırı yüksekliği 143 piksel, diğer satırlar 66 piksel, tüm sütunların genişliği 10 piksel.
    # Excel'de piksel yerine nokta ve karakter genişliği kullanıldığı için, bu değerler yaklaşık olarak dönüştürülecektir.

    # Pikseli Excel'e uygun birime dönüştürme yaklaşık değerleri
    # Satır yüksekliği doğrudan piksel olarak ayarlanabilir (yaklaşık olarak)
    # Sütun genişliği için Excel genişliği, piksel değerine göre ayarlanmalıdır. Excel genişliği ve piksel arasında doğrudan bir oran yoktur,
    # ancak genel olarak 1 karakter genişliği yaklaşık 7 piksele eşdeğerdir. Bu yüzden, 10 piksel yaklaşık 1.43 karakter genişliğine eşittir.

    # Başlık satırı ve diğer satırlar için yükseklik ayarları
    header_row_height = 143  # Excel'de doğrudan piksel olarak ayarlanabilir
    other_row_height = 66  # Excel'de doğrudan piksel olarak ayarlanabilir

    # Sütun genişliği için pikseli karakter genişliğine çevirme
    column_width_in_pixels = 66
    # Excel'de genişlik karakter cinsindendir, bu yüzden dönüşüm yapılır
    column_width_in_chars = column_width_in_pixels / 7  # Yaklaşık bir dönüşüm

    wb = openpyxl.load_workbook(final_output_path)
    ws = wb.active

    # Biçimlendirme ayarları
    header_fill = PatternFill(start_color='FFE67300', end_color='FFE67300', fill_type='solid')  # Lacivert arka plan
    white_font = Font(bold=True, color='FFFFFF')  # Başlık için beyaz, kalın yazı tipi
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                         bottom=Side(style='thin'))  # İnce kenarlık
    centered_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)  # Metni ortala ve kaydır

    # Başlık satırının yüksekliğini ayarla
    ws.row_dimensions[1].height = 143

    # Diğer tüm satırlar için yüksekliği ayarla
    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].height = 35

    # Tüm sütunlar için genişliği ayarla
    for col in ws.iter_cols(min_col=1, max_col=ws.max_column):
        ws.column_dimensions[col[0].column_letter].width = 73 / 7  # Piksel değerini Excel genişliğine çevirme

    # Tüm hücrelere biçimlendirme uygula
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
            cell.alignment = centered_alignment
            if cell.row == 1:  # Başlık satırı için ek biçimlendirme
                cell.fill = header_fill
                cell.font = white_font

    # Save the workbook with all the changes
    wb.save(final_output_path)

    print_islem_bitir('Bölge Değerlendirme', 'Günlük Değerlendirme')
    bio_ses('BS05.mp3')

    ################################## Bölge Aylık Detaylı Değerlendirme İşlemleri ###############################################

    print_islem_baslat('Bölge Değerlendirme', 'Aylık Değerlendirme')

    data = pd.read_excel(final_output_path)

    # Sayısal metrikler listesi
    numerical_metrics = [

        'FN Arıza Adet (Gün)',
        'Anfi Arıza Adet (Gün)',
        'Tap-Off Arıza Adet (Gün)',
        'Kol Arıza Adet (Gün)',
        'En Sık Arıza Süresi Saat (Gün)'

    ]

    # Kategorik metrikler listesi
    categorical_metrics = [

        'Karar (Gün)',
        'Bölge No (Gün)',
        'FN En Sık Arıza Nedeni (Gün)',
        'FN En Sık Arıza Çözümü (Gün)',
        'Anfi En Sık Arıza Nedeni (Gün)',
        'Anfi En Sık Arıza Çözümü (Gün)',
        'Tap-off En Sık Arıza Nedeni (Gün)',
        'Tap-off En Sık Arıza Çözümü (Gün)',
        'Kol En Sık Arıza Nedeni (Gün)',
        'Kol En Sık Arıza Çözümü (Gün)',
        'En Sık Tekrar Eden Hizmet No (Gün)',
        'En Sık Arıza Zamanı Saat (Gün)'

    ]

    # 'Yıl' ve 'Ay' bazında sayısal metrikler için ortalama değerleri hesaplayalım ve sonuçları tam sayı olarak ayarlayalım.
    average_values_by_year_month = data.groupby(['Yıl', 'Ay', 'İş Ortağı', 'Bölge', 'FN'])[
        numerical_metrics].sum().reset_index()
    average_values_by_year_month[numerical_metrics] = average_values_by_year_month[numerical_metrics].round(0).astype(
        int)

    # Kategorik metrikler için mod değerlerini hesaplayalım.
    # Kategorik metrikler için 'safe_mode' fonksiyonunu kullanarak hata yönetimini iyileştireceğiz.
    def safe_mode(series):
        if series.empty:
            return None
        mode_result = series.mode()
        return mode_result[0] if not mode_result.empty else None

    # Filtreleme ve gruplama işlemini 'Yıl', 'Ay', 'İş Ortağı', 'Bölge', 'FN' sütunları üzerinden yapalım
    grouped_data = data.groupby(['Yıl', 'Ay', 'İş Ortağı', 'Bölge', 'FN'])

    # Sayısal metrikler için ortalama değerleri hesaplayalım
    average_values = grouped_data[numerical_metrics].mean().round(0).astype(int).reset_index()

    # Kategorik metrikler için mod değerlerini hesaplayalım
    mode_values = grouped_data[categorical_metrics].agg(
        lambda x: x.mode().iloc[0] if not x.mode().empty else None).reset_index()

    # İki DataFrame'i birleştirme
    bolge_daily_grouped = pd.merge(average_values, mode_values, on=['Yıl', 'Ay', 'İş Ortağı', 'Bölge', 'FN'])

    # Sadece hesaplanan ve eklenen sütunları içeren yeni bir DataFrame oluşturalım
    bolge_daily_grouped = bolge_daily_grouped[['Yıl', 'Ay', 'İş Ortağı', 'Bölge', 'FN',
                                               'FN Arıza Adet (Gün)', 'Anfi Arıza Adet (Gün)',
                                               'Tap-Off Arıza Adet (Gün)', 'Kol Arıza Adet (Gün)', 'Karar (Gün)',
                                               'Bölge No (Gün)', 'FN En Sık Arıza Nedeni (Gün)',
                                               'FN En Sık Arıza Çözümü (Gün)', 'Anfi En Sık Arıza Nedeni (Gün)',
                                               'Anfi En Sık Arıza Çözümü (Gün)', 'Tap-off En Sık Arıza Nedeni (Gün)',
                                               'Tap-off En Sık Arıza Çözümü (Gün)', 'Kol En Sık Arıza Nedeni (Gün)',
                                               'Kol En Sık Arıza Çözümü (Gün)', 'En Sık Tekrar Eden Hizmet No (Gün)',
                                               'En Sık Arıza Süresi Saat (Gün)',
                                               'En Sık Arıza Zamanı Saat (Gün)']].drop_duplicates()

    # "Genel Değerlendirme" klasörünü kontrol edelim ve yoksa oluşturalım.
    general_evaluation_folder_path = Path('Aylık Değerlendirme')
    general_evaluation_folder_path.mkdir(parents=True, exist_ok=True)

    # "Genel Değerlendirme" klasörü içinde "Rapor" klasörünü kontrol edelim ve yoksa oluşturalım.
    report_folder_path = general_evaluation_folder_path / 'Bölge Değerlendirme'
    report_folder_path.mkdir(parents=True, exist_ok=True)

    # "Genel Değerlendirme" klasörü içinde "Rapor" klasörünü kontrol edelim ve yoksa oluşturalım.
    takım_dosya = report_folder_path / 'Bölge Değerlendirme Raporu'
    takım_dosya.mkdir(parents=True, exist_ok=True)

    # Günlük_Arıza sütununa göre veriyi büyükten küçüğe sıralayın
    bolge_daily_grouped = bolge_daily_grouped.sort_values(by='FN Arıza Adet (Gün)', ascending=False)

    # Excel dosyasının kaydedileceği tam yol
    final_output_path = takım_dosya / 'Bölge Aylık Değerlendirme Raporu.xlsx'

    bolge_daily_grouped.to_excel(final_output_path, index=False)

    # Ayarlanacak boyutlar: Başlık satırı yüksekliği 143 piksel, diğer satırlar 66 piksel, tüm sütunların genişliği 10 piksel.
    # Excel'de piksel yerine nokta ve karakter genişliği kullanıldığı için, bu değerler yaklaşık olarak dönüştürülecektir.

    # Pikseli Excel'e uygun birime dönüştürme yaklaşık değerleri
    # Satır yüksekliği doğrudan piksel olarak ayarlanabilir (yaklaşık olarak)
    # Sütun genişliği için Excel genişliği, piksel değerine göre ayarlanmalıdır. Excel genişliği ve piksel arasında doğrudan bir oran yoktur,
    # ancak genel olarak 1 karakter genişliği yaklaşık 7 piksele eşdeğerdir. Bu yüzden, 10 piksel yaklaşık 1.43 karakter genişliğine eşittir.

    # Başlık satırı ve diğer satırlar için yükseklik ayarları
    header_row_height = 143  # Excel'de doğrudan piksel olarak ayarlanabilir
    other_row_height = 66  # Excel'de doğrudan piksel olarak ayarlanabilir

    # Sütun genişliği için pikseli karakter genişliğine çevirme
    column_width_in_pixels = 66
    # Excel'de genişlik karakter cinsindendir, bu yüzden dönüşüm yapılır
    column_width_in_chars = column_width_in_pixels / 7  # Yaklaşık bir dönüşüm

    wb = openpyxl.load_workbook(final_output_path)
    ws = wb.active

    # Biçimlendirme ayarları
    header_fill = PatternFill(start_color='FFE67300', end_color='FFE67300', fill_type='solid')  # Lacivert arka plan
    white_font = Font(bold=True, color='FFFFFF')  # Başlık için beyaz, kalın yazı tipi
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                         bottom=Side(style='thin'))  # İnce kenarlık
    centered_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)  # Metni ortala ve kaydır

    # Başlık satırının yüksekliğini ayarla
    ws.row_dimensions[1].height = 143

    # Diğer tüm satırlar için yüksekliği ayarla
    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].height = 35

    # Tüm sütunlar için genişliği ayarla
    for col in ws.iter_cols(min_col=1, max_col=ws.max_column):
        ws.column_dimensions[col[0].column_letter].width = 73 / 7  # Piksel değerini Excel genişliğine çevirme

    # Tüm hücrelere biçimlendirme uygula
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
            cell.alignment = centered_alignment
            if cell.row == 1:  # Başlık satırı için ek biçimlendirme
                cell.fill = header_fill
                cell.font = white_font

    # Save the workbook with all the changes
    wb.save(final_output_path)

    print_islem_bitir('Bölge Değerlendirme', 'Aylık Değerlendirme')

    ################################## Bölge Yıllık İşlemleri ###############################################

    print_islem_baslat('Bölge Değerlendirme', 'Yıllık Değerlendirme')

    data = pd.read_excel(final_output_path)

    # Sayısal metrikler listesi
    numerical_metrics = [

        'FN Arıza Adet (Gün)',
        'Anfi Arıza Adet (Gün)',
        'Tap-Off Arıza Adet (Gün)',
        'Kol Arıza Adet (Gün)',
        'En Sık Arıza Süresi Saat (Gün)'

    ]

    # Kategorik metrikler listesi
    categorical_metrics = [

        'Karar (Gün)',
        'Bölge No (Gün)',
        'FN En Sık Arıza Nedeni (Gün)',
        'FN En Sık Arıza Çözümü (Gün)',
        'Anfi En Sık Arıza Nedeni (Gün)',
        'Anfi En Sık Arıza Çözümü (Gün)',
        'Tap-off En Sık Arıza Nedeni (Gün)',
        'Tap-off En Sık Arıza Çözümü (Gün)',
        'Kol En Sık Arıza Nedeni (Gün)',
        'Kol En Sık Arıza Çözümü (Gün)',
        'En Sık Tekrar Eden Hizmet No (Gün)',
        'En Sık Arıza Zamanı Saat (Gün)'

    ]

    # 'Yıl' ve 'Ay' bazında sayısal metrikler için ortalama değerleri hesaplayalım ve sonuçları tam sayı olarak ayarlayalım.
    average_values_by_year_month = data.groupby(['Yıl', 'İş Ortağı', 'Bölge', 'FN'])[
        numerical_metrics].sum().reset_index()
    average_values_by_year_month[numerical_metrics] = average_values_by_year_month[numerical_metrics].round(0).astype(
        int)

    # Kategorik metrikler için mod değerlerini hesaplayalım.
    # Kategorik metrikler için 'safe_mode' fonksiyonunu kullanarak hata yönetimini iyileştireceğiz.
    def safe_mode(series):
        if series.empty:
            return None
        mode_result = series.mode()
        return mode_result[0] if not mode_result.empty else None

    # Filtreleme ve gruplama işlemini 'Yıl', 'Ay', 'İş Ortağı', 'Bölge', 'FN' sütunları üzerinden yapalım
    grouped_data = data.groupby(['Yıl', 'İş Ortağı', 'Bölge', 'FN'])

    # Sayısal metrikler için ortalama değerleri hesaplayalım
    average_values = grouped_data[numerical_metrics].mean().round(0).astype(int).reset_index()

    # Kategorik metrikler için mod değerlerini hesaplayalım
    mode_values = grouped_data[categorical_metrics].agg(
        lambda x: x.mode().iloc[0] if not x.mode().empty else None).reset_index()

    # İki DataFrame'i birleştirme
    bolge_daily_grouped = pd.merge(average_values, mode_values, on=['Yıl', 'İş Ortağı', 'Bölge', 'FN'])

    # Sadece hesaplanan ve eklenen sütunları içeren yeni bir DataFrame oluşturalım
    bolge_daily_grouped = bolge_daily_grouped[['Yıl', 'İş Ortağı', 'Bölge', 'FN',
                                               'FN Arıza Adet (Gün)', 'Anfi Arıza Adet (Gün)',
                                               'Tap-Off Arıza Adet (Gün)', 'Kol Arıza Adet (Gün)', 'Karar (Gün)',
                                               'Bölge No (Gün)', 'FN En Sık Arıza Nedeni (Gün)',
                                               'FN En Sık Arıza Çözümü (Gün)', 'Anfi En Sık Arıza Nedeni (Gün)',
                                               'Anfi En Sık Arıza Çözümü (Gün)', 'Tap-off En Sık Arıza Nedeni (Gün)',
                                               'Tap-off En Sık Arıza Çözümü (Gün)', 'Kol En Sık Arıza Nedeni (Gün)',
                                               'Kol En Sık Arıza Çözümü (Gün)', 'En Sık Tekrar Eden Hizmet No (Gün)',
                                               'En Sık Arıza Süresi Saat (Gün)',
                                               'En Sık Arıza Zamanı Saat (Gün)']].drop_duplicates()

    # "Genel Değerlendirme" klasörünü kontrol edelim ve yoksa oluşturalım.
    general_evaluation_folder_path = Path('Yıllık Değerlendirme')
    general_evaluation_folder_path.mkdir(parents=True, exist_ok=True)

    # "Genel Değerlendirme" klasörü içinde "Rapor" klasörünü kontrol edelim ve yoksa oluşturalım.
    report_folder_path = general_evaluation_folder_path / 'Bölge Değerlendirme'
    report_folder_path.mkdir(parents=True, exist_ok=True)

    # "Genel Değerlendirme" klasörü içinde "Rapor" klasörünü kontrol edelim ve yoksa oluşturalım.
    takım_dosya = report_folder_path / 'Bölge Değerlendirme Raporu'
    takım_dosya.mkdir(parents=True, exist_ok=True)

    # Günlük_Arıza sütununa göre veriyi büyükten küçüğe sıralayın
    bolge_daily_grouped = bolge_daily_grouped.sort_values(by='FN Arıza Adet (Gün)', ascending=False)

    # Excel dosyasının kaydedileceği tam yol
    final_output_path = takım_dosya / 'Bölge Yıllık Değerlendirme Raporu.xlsx'

    bolge_daily_grouped.to_excel(final_output_path, index=False)

    # Ayarlanacak boyutlar: Başlık satırı yüksekliği 143 piksel, diğer satırlar 66 piksel, tüm sütunların genişliği 10 piksel.
    # Excel'de piksel yerine nokta ve karakter genişliği kullanıldığı için, bu değerler yaklaşık olarak dönüştürülecektir.

    # Pikseli Excel'e uygun birime dönüştürme yaklaşık değerleri
    # Satır yüksekliği doğrudan piksel olarak ayarlanabilir (yaklaşık olarak)
    # Sütun genişliği için Excel genişliği, piksel değerine göre ayarlanmalıdır. Excel genişliği ve piksel arasında doğrudan bir oran yoktur,
    # ancak genel olarak 1 karakter genişliği yaklaşık 7 piksele eşdeğerdir. Bu yüzden, 10 piksel yaklaşık 1.43 karakter genişliğine eşittir.

    # Başlık satırı ve diğer satırlar için yükseklik ayarları
    header_row_height = 143  # Excel'de doğrudan piksel olarak ayarlanabilir
    other_row_height = 66  # Excel'de doğrudan piksel olarak ayarlanabilir

    # Sütun genişliği için pikseli karakter genişliğine çevirme
    column_width_in_pixels = 66
    # Excel'de genişlik karakter cinsindendir, bu yüzden dönüşüm yapılır
    column_width_in_chars = column_width_in_pixels / 7  # Yaklaşık bir dönüşüm

    wb = openpyxl.load_workbook(final_output_path)
    ws = wb.active

    # Biçimlendirme ayarları
    header_fill = PatternFill(start_color='FFE67300', end_color='FFE67300', fill_type='solid')  # Lacivert arka plan
    white_font = Font(bold=True, color='FFFFFF')  # Başlık için beyaz, kalın yazı tipi
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                         bottom=Side(style='thin'))  # İnce kenarlık
    centered_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)  # Metni ortala ve kaydır

    # Başlık satırının yüksekliğini ayarla
    ws.row_dimensions[1].height = 143

    # Diğer tüm satırlar için yüksekliği ayarla
    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].height = 35

    # Tüm sütunlar için genişliği ayarla
    for col in ws.iter_cols(min_col=1, max_col=ws.max_column):
        ws.column_dimensions[col[0].column_letter].width = 73 / 7  # Piksel değerini Excel genişliğine çevirme

    # Tüm hücrelere biçimlendirme uygula
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
            cell.alignment = centered_alignment
            if cell.row == 1:  # Başlık satırı için ek biçimlendirme
                cell.fill = header_fill
                cell.font = white_font

    # Yuvarlanacak başlıkların listesi
    headers_to_round = [
        "Arıza Süresi Saat (Gün)"
    ]

    # Sütun harflerini ve başlıklarını bul
    columns_to_round = {}
    for idx, col in enumerate(ws.iter_cols(1, ws.max_column, 1, 1, True), start=1):
        if col[0] in headers_to_round:
            columns_to_round[get_column_letter(idx)] = col[0]

    # Her belirtilen sütun için, sayısal hücreleri yuvarla ve formatı '0.0' olarak ayarla
    for col_letter, header in columns_to_round.items():
        for cell in ws[col_letter][1:]:  # Üstbilgiyi atla, 2. satırdan başla
            if isinstance(cell.value, (int, float)):
                # Sayısal değeri yuvarla ve tek ondalık basamakla formatla
                cell.value = round(cell.value, 1)
                cell.number_format = '0.0'

    # Save the workbook with all the changes
    wb.save(final_output_path)

    print_islem_bitir('Bölge Değerlendirme', 'Yıllık Değerlendirme')
    bio_ses('BS05.mp3')

    ################################## Bölge Genel Değerlendirme İşlemleri ###############################################

    print_islem_baslat('Bölge Değerlendirme', 'Genel Değerlendirme')

    data = pd.read_excel(final_output_path)

    # Sayısal metrikler listesi
    numerical_metrics = [

        'FN Arıza Adet (Gün)',
        'Anfi Arıza Adet (Gün)',
        'Tap-Off Arıza Adet (Gün)',
        'Kol Arıza Adet (Gün)',
        'En Sık Arıza Süresi Saat (Gün)'

    ]

    # Kategorik metrikler listesi
    categorical_metrics = [

        'Karar (Gün)',
        'Bölge No (Gün)',
        'FN En Sık Arıza Nedeni (Gün)',
        'FN En Sık Arıza Çözümü (Gün)',
        'Anfi En Sık Arıza Nedeni (Gün)',
        'Anfi En Sık Arıza Çözümü (Gün)',
        'Tap-off En Sık Arıza Nedeni (Gün)',
        'Tap-off En Sık Arıza Çözümü (Gün)',
        'Kol En Sık Arıza Nedeni (Gün)',
        'Kol En Sık Arıza Çözümü (Gün)',
        'En Sık Tekrar Eden Hizmet No (Gün)',
        'En Sık Arıza Zamanı Saat (Gün)'

    ]

    # 'Yıl' ve 'Ay' bazında sayısal metrikler için ortalama değerleri hesaplayalım ve sonuçları tam sayı olarak ayarlayalım.
    average_values_by_year_month = data.groupby(['İş Ortağı', 'Bölge', 'FN'])[numerical_metrics].sum().reset_index()
    average_values_by_year_month[numerical_metrics] = average_values_by_year_month[numerical_metrics].round(0).astype(
        int)

    # Kategorik metrikler için mod değerlerini hesaplayalım.
    # Kategorik metrikler için 'safe_mode' fonksiyonunu kullanarak hata yönetimini iyileştireceğiz.
    def safe_mode(series):
        if series.empty:
            return None
        mode_result = series.mode()
        return mode_result[0] if not mode_result.empty else None

    # Filtreleme ve gruplama işlemini 'Yıl', 'Ay', 'İş Ortağı', 'Bölge', 'FN' sütunları üzerinden yapalım
    grouped_data = data.groupby(['İş Ortağı', 'Bölge', 'FN'])

    # Sayısal metrikler için ortalama değerleri hesaplayalım
    average_values = grouped_data[numerical_metrics].mean().round(0).astype(int).reset_index()

    # Kategorik metrikler için mod değerlerini hesaplayalım
    mode_values = grouped_data[categorical_metrics].agg(
        lambda x: x.mode().iloc[0] if not x.mode().empty else None).reset_index()

    # İki DataFrame'i birleştirme
    bolge_daily_grouped = pd.merge(average_values, mode_values, on=['İş Ortağı', 'Bölge', 'FN'])

    # Sadece hesaplanan ve eklenen sütunları içeren yeni bir DataFrame oluşturalım
    bolge_daily_grouped = bolge_daily_grouped[['İş Ortağı', 'Bölge', 'FN',
                                               'FN Arıza Adet (Gün)', 'Anfi Arıza Adet (Gün)',
                                               'Tap-Off Arıza Adet (Gün)', 'Kol Arıza Adet (Gün)', 'Karar (Gün)',
                                               'Bölge No (Gün)', 'FN En Sık Arıza Nedeni (Gün)',
                                               'FN En Sık Arıza Çözümü (Gün)', 'Anfi En Sık Arıza Nedeni (Gün)',
                                               'Anfi En Sık Arıza Çözümü (Gün)', 'Tap-off En Sık Arıza Nedeni (Gün)',
                                               'Tap-off En Sık Arıza Çözümü (Gün)', 'Kol En Sık Arıza Nedeni (Gün)',
                                               'Kol En Sık Arıza Çözümü (Gün)', 'En Sık Tekrar Eden Hizmet No (Gün)',
                                               'En Sık Arıza Süresi Saat (Gün)',
                                               'En Sık Arıza Zamanı Saat (Gün)']].drop_duplicates()

    # "Genel Değerlendirme" klasörünü kontrol edelim ve yoksa oluşturalım.
    general_evaluation_folder_path = Path('Genel Değerlendirme')
    general_evaluation_folder_path.mkdir(parents=True, exist_ok=True)

    # "Genel Değerlendirme" klasörü içinde "Rapor" klasörünü kontrol edelim ve yoksa oluşturalım.
    report_folder_path = general_evaluation_folder_path / 'Bölge Değerlendirme'
    report_folder_path.mkdir(parents=True, exist_ok=True)

    # "Genel Değerlendirme" klasörü içinde "Rapor" klasörünü kontrol edelim ve yoksa oluşturalım.
    takım_dosya = report_folder_path / 'Bölge Değerlendirme Raporu'
    takım_dosya.mkdir(parents=True, exist_ok=True)

    # Günlük_Arıza sütununa göre veriyi büyükten küçüğe sıralayın
    bolge_daily_grouped = bolge_daily_grouped.sort_values(by='FN Arıza Adet (Gün)', ascending=False)

    # Excel dosyasının kaydedileceği tam yol
    final_output_path = takım_dosya / 'Bölge Genel Değerlendirme Raporu.xlsx'

    bolge_daily_grouped.to_excel(final_output_path, index=False)

    # Ayarlanacak boyutlar: Başlık satırı yüksekliği 143 piksel, diğer satırlar 66 piksel, tüm sütunların genişliği 10 piksel.
    # Excel'de piksel yerine nokta ve karakter genişliği kullanıldığı için, bu değerler yaklaşık olarak dönüştürülecektir.

    # Pikseli Excel'e uygun birime dönüştürme yaklaşık değerleri
    # Satır yüksekliği doğrudan piksel olarak ayarlanabilir (yaklaşık olarak)
    # Sütun genişliği için Excel genişliği, piksel değerine göre ayarlanmalıdır. Excel genişliği ve piksel arasında doğrudan bir oran yoktur,
    # ancak genel olarak 1 karakter genişliği yaklaşık 7 piksele eşdeğerdir. Bu yüzden, 10 piksel yaklaşık 1.43 karakter genişliğine eşittir.

    # Başlık satırı ve diğer satırlar için yükseklik ayarları
    header_row_height = 143  # Excel'de doğrudan piksel olarak ayarlanabilir
    other_row_height = 66  # Excel'de doğrudan piksel olarak ayarlanabilir

    # Sütun genişliği için pikseli karakter genişliğine çevirme
    column_width_in_pixels = 66
    # Excel'de genişlik karakter cinsindendir, bu yüzden dönüşüm yapılır
    column_width_in_chars = column_width_in_pixels / 7  # Yaklaşık bir dönüşüm

    wb = openpyxl.load_workbook(final_output_path)
    ws = wb.active

    # Biçimlendirme ayarları
    header_fill = PatternFill(start_color='FFE67300', end_color='FFE67300', fill_type='solid')  # Lacivert arka plan
    white_font = Font(bold=True, color='FFFFFF')  # Başlık için beyaz, kalın yazı tipi
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                         bottom=Side(style='thin'))  # İnce kenarlık
    centered_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)  # Metni ortala ve kaydır

    # Başlık satırının yüksekliğini ayarla
    ws.row_dimensions[1].height = 143

    # Diğer tüm satırlar için yüksekliği ayarla
    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].height = 35

    # Tüm sütunlar için genişliği ayarla
    for col in ws.iter_cols(min_col=1, max_col=ws.max_column):
        ws.column_dimensions[col[0].column_letter].width = 73 / 7  # Piksel değerini Excel genişliğine çevirme

    # Tüm hücrelere biçimlendirme uygula
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
            cell.alignment = centered_alignment
            if cell.row == 1:  # Başlık satırı için ek biçimlendirme
                cell.fill = header_fill
                cell.font = white_font

    # Yuvarlanacak başlıkların listesi
    headers_to_round = [
        "Arıza Süresi Saat (Gün)"
    ]

    # Sütun harflerini ve başlıklarını bul
    columns_to_round = {}
    for idx, col in enumerate(ws.iter_cols(1, ws.max_column, 1, 1, True), start=1):
        if col[0] in headers_to_round:
            columns_to_round[get_column_letter(idx)] = col[0]

    # Her belirtilen sütun için, sayısal hücreleri yuvarla ve formatı '0.0' olarak ayarla
    for col_letter, header in columns_to_round.items():
        for cell in ws[col_letter][1:]:  # Üstbilgiyi atla, 2. satırdan başla
            if isinstance(cell.value, (int, float)):
                # Sayısal değeri yuvarla ve tek ondalık basamakla formatla
                cell.value = round(cell.value, 1)
                cell.number_format = '0.0'

    # Save the workbook with all the changes
    wb.save(final_output_path)

    print_islem_bitir('Bölge Değerlendirme', 'Genel Değerlendirme')

    ################################## Bölge İşlemleri Tamamlandı ###############################################

    print_islem_bitir('Bölge Değerlendirme', 'Tüm Değerlendirme')
    bio_ses('BS05.mp3')

def Genel_Durum_Degerlendirme(data_original, BioT):
    ################################## Genel Durum Değerlendirme Veri Optimizasyon İşlemleri ###############################################
    print('')

    print_islem_baslat('Genel Durum Değerlendirme', 'Veri Optimizasyon')

    data=data_original


    # Veri setini yükle
    data = pd.read_excel('Veri Girişi/METROBIL.xlsx')

    # Tarih/saat sütunlarını datetime türüne dönüştürme
    data['Kayıt Tarihi'] = pd.to_datetime(data['Kayıt Tarihi'])
    data['Çözüm Tarihi'] = pd.to_datetime(data['Çözüm Tarihi'])

    # Yıl, Ay ve Gün sütunlarını ekleyin
    data['Yıl'] = data['Kayıt Tarihi'].dt.year
    data['Ay'] = data['Kayıt Tarihi'].dt.month
    data['Gün'] = data['Kayıt Tarihi'].dt.day

    # Onay ve Ret durumlarını sayısal değerlere dönüştürme
    data['Onay'] = (data['Durum'] == 'Onaylandı').astype(int)
    data['Ret'] = (data['Ret Geçmişi'] == 'Var').astype(int)

    # Çözüm süresini saat cinsinden hesaplama
    data['Çözüm Süresi Saat'] = (data['Çözüm Tarihi'] - data['Kayıt Tarihi']).dt.total_seconds() / 3600

    # Devre No sütunundaki NaN veya boş değerleri işleme ve metin olarak ele alınması için düzeltme
    data['Devre No'] = data['Devre No'].fillna('').astype(str)

    # Ayrıştırma işlemi
    data[['İş Ortağı', 'Bölge', 'FN', 'Anfi', 'Tap-Off', 'Kol']] = pd.DataFrame(data['Devre No'].apply(lambda x: x.split('.')[:6] if len(x.split('.')) >= 6 else [None] * 6).tolist(), index=data.index)

    # Günlük istatistiklerin hesaplanması
    gunluk_istatistikler = data.groupby(['Yıl', 'Ay', 'Gün']).agg(
        Günlük_Toplam_Arıza_Sayısı=('Çağrı No', 'count'),
        Günlük_Toplam_Onaylanan_Arıza_Sayısı=('Onay', 'sum'),
        Günlük_Toplam_Ret_Edilen_Arıza_Sayısı=('Ret', 'sum'),
        Günlük_Ortalama_Çözüm_Süresi=('Çözüm Süresi Saat', 'mean')
    ).reset_index()

    # 'Hizmet No' ve 'Kayıt Tarihi'ne göre sıralama
    data_sorted = data.sort_values(by=['Hizmet No', 'Kayıt Tarihi'])

    # Her bir arıza kaydı için bir önceki kayıtla arasındaki gün farkını hesaplama
    data_sorted['Önceki Kayıt Farkı'] = data_sorted.groupby('Hizmet No')['Kayıt Tarihi'].diff().dt.days

    # Belirlenen gün sayısı veya daha az olan farkları işaretleyelim.
    data_sorted['Tekrar'] = data_sorted['Önceki Kayıt Farkı'] <= BioT

    # Günlük bazda, 'Tekrar' True olan kayıtların sayısını hesaplama
    gunluk_tekrar_eden_ariza_sayisi = data_sorted.groupby(['Yıl', 'Ay', 'Gün'])['Tekrar'].sum().reset_index(name='Günlük_Tekrar_Eden_Toplam_Arıza_Sayısı')

    # En çok arıza çözümleyen takımı hesaplama
    gunluk_en_cok_ariza_cozumleyen_takim = data.groupby(['Yıl', 'Ay', 'Gün', 'Takım']).size().reset_index(name='Çözümlenen_Arıza_Sayısı').sort_values(by=['Yıl', 'Ay', 'Gün', 'Çözümlenen_Arıza_Sayısı'], ascending=[True, True, True, False]).drop_duplicates(subset=['Yıl', 'Ay', 'Gün']).rename(columns={'Takım': 'Günlük_En_Çok_Arıza_Çözümleyen_Takım', 'Çözümlenen_Arıza_Sayısı': 'Günlük_En_Çok_Arıza_Çözümleyen_Takımın_Çözümlediği_Arıza_Sayısı'})

    # Bölge-FN kombinasyonu için istatistikler
    data['İşOrtağı_Bölge_FN_Kombinasyonu'] = data['İş Ortağı'] + '.' + data['Bölge'] + '.' + data['FN']
    gunluk_en_cok_ariza_gelen_bolge_fn = data.groupby(['Yıl', 'Ay', 'Gün', 'İşOrtağı_Bölge_FN_Kombinasyonu']).size().reset_index(name='Arıza_Sayısı').sort_values(by=['Yıl', 'Ay', 'Gün', 'Arıza_Sayısı'], ascending=[True, True, True, False]).drop_duplicates(subset=['Yıl', 'Ay', 'Gün']).rename(columns={'İşOrtağı_Bölge_FN_Kombinasyonu': 'Günlük_En_Çok_Arıza_Gelen_Bölge_FN_Kombinasyonu', 'Arıza_Sayısı': 'Günlük_En_Çok_Arıza_Gelen_Bölge_FN_Kombinasyonu_Arıza_Sayısı'})

    # Sonuçların birleştirilmesi ve sütun isimlerinin güncellenmesi
    sonuc_raporu_gunluk = pd.merge(gunluk_istatistikler, gunluk_tekrar_eden_ariza_sayisi, on=['Yıl', 'Ay', 'Gün'])
    sonuc_raporu_gunluk = pd.merge(sonuc_raporu_gunluk, gunluk_en_cok_ariza_cozumleyen_takim, on=['Yıl', 'Ay', 'Gün'])
    sonuc_raporu_gunluk = pd.merge(sonuc_raporu_gunluk, gunluk_en_cok_ariza_gelen_bolge_fn, on=['Yıl', 'Ay', 'Gün'])

    print_islem_bitir('Genel Durum Değerlendirme', 'Veri Optimizasyon')


    ################################## Genel Durum Günlük Değerlendirme Veri Optimizasyon İşlemleri ###############################################

    print_islem_baslat('Genel Durum Değerlendirme', 'Günlük Değerlendirme')


    # Aylık en çok arıza bırakan abone ve bu abonenin bıraktığı arıza sayısı hesaplama (Günlük için de benzer işlemler)
    gunluk_en_cok_ariza_birakan_hizmet_no = data.groupby(['Yıl', 'Ay', 'Gün', 'Hizmet No']).size().reset_index(name='Bıraktığı_Arıza_Sayısı').sort_values(by=['Yıl', 'Ay', 'Gün', 'Bıraktığı_Arıza_Sayısı'], ascending=[True, True, True, False])

    # Her gün için en çok arıza bırakan aboneyi belirle
    gunluk_en_cok_ariza_birakan_hizmet_no = gunluk_en_cok_ariza_birakan_hizmet_no.drop_duplicates(subset=['Yıl', 'Ay', 'Gün'])

    # Bu hesaplamaların sonuçlarını ana sonuç raporuna ekleyelim
    sonuc_raporu_gunluk = pd.merge(sonuc_raporu_gunluk, gunluk_en_cok_ariza_birakan_hizmet_no, on=['Yıl', 'Ay', 'Gün'], how='left').rename(columns={'Hizmet No': 'Günlük_En_Çok_Arıza_Bırakan_Abone_Hizmet_No', 'Bıraktığı_Arıza_Sayısı': 'Günlük_En_Çok_Arıza_Bırakan_Abonenin_Arıza_Sayısı'})

    # Günlük arıza bırakan benzersiz abone sayısını hesaplama
    gunluk_arıza_bırakan_abone_sayısı = data.groupby(['Yıl', 'Ay', 'Gün'])['Hizmet No'].nunique().reset_index(name='Günlük_Arıza_Bırakan_Abone_Sayısı')

    # 'Tekrar' bilgisini kullanarak tekrar eden arızaları filtreleyelim.
    tekrar_edilen_ariza_birakan_hizmet_no_gunluk = data_sorted[data_sorted['Tekrar'] == True].groupby(['Yıl', 'Ay', 'Gün', 'Hizmet No']).size().reset_index(name='Tekrar_Eden_Arıza_Sayısı').sort_values(by=['Yıl', 'Ay', 'Gün', 'Tekrar_Eden_Arıza_Sayısı'], ascending=[True, True, True, False])

    # Her gün için en çok tekrar eden arıza bırakan aboneyi belirleyelim.
    gunluk_en_cok_tekrar_edilen_ariza_birakan_hizmet_no = tekrar_edilen_ariza_birakan_hizmet_no_gunluk.drop_duplicates(subset=['Yıl', 'Ay', 'Gün'])

    # Bu hesaplamaların sonuçlarını ana sonuç raporuna ekleyelim.
    sonuc_raporu_gunluk = pd.merge(sonuc_raporu_gunluk, gunluk_en_cok_tekrar_edilen_ariza_birakan_hizmet_no.rename(columns={'Hizmet No': 'Günlük_En_Çok_Tekrar_Eden_Arıza_Bırakan_Abone_Hizmet_No', 'Tekrar_Eden_Arıza_Sayısı': 'Günlük_En_Çok_Tekrar_Eden_Arıza_Bırakan_Abonenin_Tekrar_Eden_Arıza_Sayısı'}), on=['Yıl', 'Ay', 'Gün'], how='left')

    # Günlük arıza bırakan benzersiz abone sayısını hesaplama
    gunluk_arıza_bırakan_abone_sayısı = data.groupby(['Yıl', 'Ay', 'Gün'])['Hizmet No'].nunique().reset_index(name='Günlük_Arıza_Bırakan_Abone_Sayısı')

    # Bu hesaplamaların sonuçlarını ana sonuç raporuna ekleyelim
    sonuc_raporu_gunluk = pd.merge(sonuc_raporu_gunluk, gunluk_arıza_bırakan_abone_sayısı, on=['Yıl', 'Ay', 'Gün'], how='left')


    sonuc_raporu_gunluk = sonuc_raporu_gunluk.rename(columns={
        'Günlük_Toplam_Arıza_Sayısı': 'Günlük Toplam Arıza Sayısı',
        'Günlük_Toplam_Onaylanan_Arıza_Sayısı': 'Günlük Toplam Onaylanan Arıza Sayısı',
        'Günlük_Toplam_Ret_Edilen_Arıza_Sayısı': 'Günlük Toplam Ret Edilen Arıza Sayısı',
        'Günlük_Ortalama_Çözüm_Süresi': 'Günlük Ortalama Çözüm Süresi (Saat)',
        'Günlük_Tekrar_Eden_Toplam_Arıza_Sayısı': 'Günlük Toplam Tekrar Eden Arıza Sayısı',
        'Günlük_En_Çok_Arıza_Çözümleyen_Takım': 'Günlük En Çok Arıza Çözümleyen Takım',
        'Günlük_En_Çok_Arıza_Çözümleyen_Takımın_Çözümlediği_Arıza_Sayısı': 'Günlük En Çok Arıza Çözümleyen Takımın Çözümlediği Arıza Sayısı',
        'Günlük_En_Çok_Arıza_Gelen_Bölge_FN_Kombinasyonu': 'Günlük En Çok Arıza Gelen Bölge FN',
        'Günlük_En_Çok_Arıza_Gelen_Bölge_FN_Kombinasyonu_Arıza_Sayısı': 'Günlük En Çok Arıza Gelen Bölge FN Arıza Sayısı',
        'Günlük_En_Çok_Tekrar_Eden_Arıza_Çözümleyen_Takım': 'Günlük En Çok Tekrar Eden Arıza Çözümleyen Takım',
        'Günlük_En_Çok_Tekrar_Eden_Arıza_Çözümleyen_Takımın_Çözümlediği_Tekrar_Eden_Arıza_Sayısı': 'Günlük En Çok Tekrar Eden Arıza Çözümleyen Takımın Çözümlediği Tekrar Eden Arıza Sayısı',
        'Günlük_En_Çok_Arıza_Bırakan_Abone_Hizmet_No': 'Günlük En Çok Arıza Bırakan Abone Hizmet No',
        'Günlük_En_Çok_Arıza_Bırakan_Abonenin_Arıza_Sayısı': 'Günlük En Çok Arıza Bırakan Abonenin Arıza Sayısı',
        'Günlük_En_Çok_Tekrar_Eden_Arıza_Bırakan_Abone_Hizmet_No': 'Günlük En Çok Tekrar Eden Arıza Bırakan Abone Hizmet No',
        'Günlük_En_Çok_Tekrar_Eden_Arıza_Bırakan_Abonenin_Tekrar_Eden_Arıza_Sayısı': 'Günlük En Çok Tekrar Eden Arıza Bırakan Abonenin Tekrar Eden Arıza Sayısı',
        'Günlük_Arıza_Bırakan_Abone_Sayısı': 'Günlük Arıza Bırakan Abone Sayısı'
    })

    # "Genel Değerlendirme" klasörünü kontrol edelim ve yoksa oluşturalım.
    general_evaluation_folder_path = Path('Günlük Değerlendirme')
    general_evaluation_folder_path.mkdir(parents=True, exist_ok=True)

    # "Genel Değerlendirme" klasörü içinde "Rapor" klasörünü kontrol edelim ve yoksa oluşturalım.
    report_folder_path = general_evaluation_folder_path / 'Genel Durum Değerlendirme'
    report_folder_path.mkdir(parents=True, exist_ok=True)

    # Günlük_Arıza sütununa göre veriyi büyükten küçüğe sıralayın
    #sonuc_raporu_gunluk = sonuc_raporu_gunluk.sort_values(by='Günlük Toplam Arıza Sayısı', ascending=False)

    # Excel dosyasının kaydedileceği tam yol
    final_output_path = report_folder_path / 'Genel Durum Günlük Değerlendirme Raporu.xlsx'

    sonuc_raporu_gunluk.to_excel(final_output_path, index=False)


    # Ayarlanacak boyutlar: Başlık satırı yüksekliği 143 piksel, diğer satırlar 66 piksel, tüm sütunların genişliği 10 piksel.
    # Excel'de piksel yerine nokta ve karakter genişliği kullanıldığı için, bu değerler yaklaşık olarak dönüştürülecektir.

    # Pikseli Excel'e uygun birime dönüştürme yaklaşık değerleri
    # Satır yüksekliği doğrudan piksel olarak ayarlanabilir (yaklaşık olarak)
    # Sütun genişliği için Excel genişliği, piksel değerine göre ayarlanmalıdır. Excel genişliği ve piksel arasında doğrudan bir oran yoktur,
    # ancak genel olarak 1 karakter genişliği yaklaşık 7 piksele eşdeğerdir. Bu yüzden, 10 piksel yaklaşık 1.43 karakter genişliğine eşittir.

    # Başlık satırı ve diğer satırlar için yükseklik ayarları
    header_row_height = 132  # Excel'de doğrudan piksel olarak ayarlanabilir
    other_row_height = 125  # Excel'de doğrudan piksel olarak ayarlanabilir

    # Sütun genişliği için pikseli karakter genişliğine çevirme
    column_width_in_pixels = 125
    # Excel'de genişlik karakter cinsindendir, bu yüzden dönüşüm yapılır
    column_width_in_chars = column_width_in_pixels / 7  # Yaklaşık bir dönüşüm

    # Dosyayı yükleyin
    wb = openpyxl.load_workbook(final_output_path)
    ws = wb.active

    # Biçimlendirme ayarları
    header_fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')  # Koyu lacivert arka plan
    white_font = Font(bold=True, color='FFFFFF')  # Başlık için beyaz, kalın yazı tipi
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))  # İnce kenarlık
    centered_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)  # Metni ortala


    # Başlık satırının yüksekliğini ayarla
    ws.row_dimensions[1].height = header_row_height

    # Diğer tüm satırlar için yüksekliği ayarla
    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].height = other_row_height

    # Tüm sütunlar için genişliği ayarla
    for col in ws.iter_cols(min_col=1, max_col=ws.max_column, min_row=1, max_row=1):
        ws.column_dimensions[get_column_letter(col[0].column)].width = column_width_in_chars

    # Tüm hücrelere biçimlendirme uygula
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
            cell.alignment = centered_alignment
            if cell.row == 1:  # Başlık satırı için ek biçimlendirme
                cell.fill = header_fill
                cell.font = white_font

    # Değerleri yuvarlama ve biçimlendirme
    # Yuvarlanacak sütunlar için doğru indeksleri bul
    headers_to_round = ['Günlük Ortalama Çözüm Süresi (Saat)']  # Yuvarlanacak başlıkların listesi
    columns_to_round = []  # Yuvarlanacak sütun indekslerini saklar

    for col in ws.iter_cols(min_row=1, max_row=1, min_col=1, max_col=ws.max_column, values_only=True):
        if col[0] in headers_to_round:
            columns_to_round.append(col[0])

    # Her belirtilen sütun için, sayısal hücreleri yuvarla ve formatı '0.0' olarak ayarla
    for idx, header in enumerate(ws[1], start=1):
        if header.value in columns_to_round:
            for cell in ws[get_column_letter(idx)][2:]:  # Üstbilgiyi atla, 2. satırdan başla
                if isinstance(cell.value, (int, float)):
                    cell.value = round(cell.value, 1)  # Sayısal değeri yuvarla
                    cell.number_format = '0.0'  # Biçimi '0.0' olarak ayarla

    # Değişiklikleri kaydet
    wb.save(final_output_path)

    ###################################################### Grafik Durum Değerlendirme PDF ############################################

    try:

        print_islem_baslat('Genel Durum Değerlendirme', 'Günlük Değerlendirme Grafik PDF')

        # Font ayarlarını güncelleyelim
        matplotlib.rcParams['font.family'] = 'DejaVu Sans'
        matplotlib.rcParams['pdf.fonttype'] = 42

        # Stil ayarını güncelleyelim
        plt.style.use('seaborn-v0_8-darkgrid')

        # Raporlama klasörlerinin oluşturulması
        günlük_degerlendirme_folder_path = Path('Günlük Değerlendirme')
        günlük_degerlendirme_folder_path.mkdir(parents=True, exist_ok=True)

        report_folder_path = günlük_degerlendirme_folder_path / 'Genel Durum Değerlendirme'
        report_folder_path.mkdir(parents=True, exist_ok=True)

        # PDF olarak kaydetme yolu
        final_pdf_path = report_folder_path / 'Günlük Genel Durum Değerlendirme Grafiği.pdf'

        try:
            # Verileri tam sayıya dönüştürme
            sonuc_raporu_gunluk['Yıl'] = sonuc_raporu_gunluk['Yıl'].astype(int)
            sonuc_raporu_gunluk['Ay'] = sonuc_raporu_gunluk['Ay'].astype(int)
            sonuc_raporu_gunluk['Gün'] = sonuc_raporu_gunluk['Gün'].astype(int)
        except:
            pass

        sonuc_raporu_gunluk['Tarih_str'] = sonuc_raporu_gunluk['Yıl'].astype(str) + '-' + sonuc_raporu_gunluk[
            'Gün'].astype(str) + '-' + sonuc_raporu_gunluk['Ay'].astype(str)

        # Tarih sütununu oluşturma
        sonuc_raporu_gunluk['Tarih'] = pd.to_datetime(sonuc_raporu_gunluk['Tarih_str'], format='%Y-%d-%m')


        with PdfPages(final_pdf_path) as pdf:
            for column in sonuc_raporu_gunluk.columns:
                # 'Tarih' ve 'Tarih_str' sütunlarını hariç tut
                if column not in ['Yıl', 'Ay', 'Gün', 'Tarih', 'Tarih_str']:
                    fig, ax = plt.subplots(figsize=(19, 13))
                    # 'Tarih' sütununu kullanarak plot yapın
                    ax.plot(sonuc_raporu_gunluk['Tarih'], sonuc_raporu_gunluk[column],
                            marker='o', linestyle='-', color='navy', linewidth=2, markersize=8)
                    ax.set_title(column, fontsize=16, fontweight='bold', color='navy')
                    ax.set_xlabel('Yıl-Ay-Gün', fontsize=14, fontweight='bold')
                    ax.set_ylabel(column, fontsize=14, fontweight='bold')

                    # Tarih formatını ayarlama ve tarihleri dikey yazdırma
                    ax.xaxis.set_major_locator(AutoDateLocator())
                    ax.xaxis.set_major_formatter(DateFormatter('%Y-%m-%d'))
                    ax.tick_params(axis='x', rotation=90, labelsize=10)  # Tarihleri dikey yazdır
                    ax.tick_params(axis='y', labelsize=12)

                    # Her grafiğin sağ üstüne not ekleyelim
                    text = 'Copyright 2024\nCreated by Biosphere.\nDeveloped by Furkan ARINCI'
                    plt.text(0.95, 0.95, text, fontsize=8, verticalalignment='top', horizontalalignment='right',
                             transform=ax.transAxes, color='gray', alpha=0.7)

                    pdf.savefig(fig, dpi=600)  # Grafikleri PDF'e kaydet, DPI ayarını burada yap
                    plt.close(fig)

        # İşlem tamamlandı
        print_islem_bitir('Genel Durum Değerlendirme', 'Günlük Değerlendirme Grafik PDF')

    except Exception as e:
        print(f"Hata: {e}")

    print_islem_bitir('Genel Durum Değerlendirme', 'Günlük Değerlendirme')
    bio_ses('BS05.mp3')

    ################################## Genel Durum Aylık Değerlendirme  İşlemleri ###############################################


    print_islem_baslat('Genel Durum Değerlendirme', 'Aylık Değerlendirme')

    data=data_original


    # Veri setini yükle
    data = pd.read_excel('Veri Girişi/METROBIL.xlsx')

    # Tarih/saat sütunlarını datetime türüne dönüştürme
    data['Kayıt Tarihi'] = pd.to_datetime(data['Kayıt Tarihi'])
    data['Çözüm Tarihi'] = pd.to_datetime(data['Çözüm Tarihi'])

    # Yıl ve Ay sütunlarını ekleyin
    data['Yıl'] = data['Kayıt Tarihi'].dt.year
    data['Ay'] = data['Kayıt Tarihi'].dt.month

    # Onay ve Ret durumlarını sayısal değerlere dönüştürme
    data['Onay'] = (data['Durum'] == 'Onaylandı').astype(int)
    data['Ret'] = (data['Ret Geçmişi'] == 'Var').astype(int)

    # Çözüm süresini saat cinsinden hesaplama
    data['Çözüm Süresi Saat'] = (data['Çözüm Tarihi'] - data['Kayıt Tarihi']).dt.total_seconds() / 3600

    # Devre No sütunundaki NaN veya boş değerleri işleme ve metin olarak ele alınması için düzeltme
    data['Devre No'] = data['Devre No'].fillna('').astype(str)

    # Ayrıştırma işlemi
    data[['İş Ortağı', 'Bölge', 'FN', 'Anfi', 'Tap-Off', 'Kol']] = pd.DataFrame(data['Devre No'].apply(lambda x: x.split('.')[:6] if len(x.split('.')) >= 6 else [None] * 6).tolist(), index=data.index)

    # Aylık istatistiklerin hesaplanması
    aylık_istatistikler = data.groupby(['Yıl', 'Ay']).agg(
        Aylık_Toplam_Arıza_Sayısı=('Çağrı No', 'count'),
        Aylık_Toplam_Onaylanan_Arıza_Sayısı=('Onay', 'sum'),
        Aylık_Toplam_Ret_Edilen_Arıza_Sayısı=('Ret', 'sum'),
        Aylık_Ortalama_Çözüm_Süresi=('Çözüm Süresi Saat', 'mean')
    ).reset_index()


    # İlk olarak, veri setini 'Hizmet No' ve 'Kayıt Tarihi'ne göre sıralayalım.
    data_sorted = data.sort_values(by=['Hizmet No', 'Kayıt Tarihi'])

    # Her bir arıza kaydı için bir önceki kayıtla arasındaki gün farkını hesaplayalım.
    data_sorted['Önceki Kayıt Farkı'] = data_sorted.groupby('Hizmet No')['Kayıt Tarihi'].diff().dt.days

    # 5 gün veya daha az olan farkları işaretleyelim.
    data_sorted['x Gün İçinde Tekrar'] = data_sorted['Önceki Kayıt Farkı'] <= BioT

    # Şimdi, 'Yıl' ve 'Ay' bazında, '5 Gün İçinde Tekrar' True olan kayıtların sayısını hesaplayalım.
    aylık_tekrar_eden_ariza_sayisi = data_sorted.groupby(['Yıl', 'Ay'])['x Gün İçinde Tekrar'].sum().reset_index(name='Aylık_Tekrar_Eden_Toplam_Arıza_Sayısı')


    # En çok arıza çözümleyen takımı hesaplama
    en_cok_ariza_cozumleyen_takim = data.groupby(['Yıl', 'Ay', 'Takım']).size().reset_index(name='Çözümlenen_Arıza_Sayısı').sort_values(by=['Yıl', 'Ay', 'Çözümlenen_Arıza_Sayısı'], ascending=[True, True, False]).drop_duplicates(subset=['Yıl', 'Ay']).rename(columns={'Takım': 'Aylık_En_Çok_Arıza_Çözümleyen_Takım', 'Çözümlenen_Arıza_Sayısı': 'Aylık_En_Çok_Arıza_Çözümleyen_Takımın_Çözümlediği_Arıza_Sayısı'})

    # Bölge-FN kombinasyonu için istatistikler
    data['İşOrtağı_Bölge_FN_Kombinasyonu'] = data['İş Ortağı'] + '.' + data['Bölge'] + '.' + data['FN']
    en_cok_ariza_gelen_bolge_fn = data.groupby(['Yıl', 'Ay', 'İşOrtağı_Bölge_FN_Kombinasyonu']).size().reset_index(name='Arıza_Sayısı').sort_values(by=['Yıl', 'Ay', 'Arıza_Sayısı'], ascending=[True, True, False]).drop_duplicates(subset=['Yıl', 'Ay']).rename(columns={'İşOrtağı_Bölge_FN_Kombinasyonu': 'Aylık_En_Çok_Arıza_Gelen_Bölge_FN_Kombinasyonu', 'Arıza_Sayısı': 'Aylık_En_Çok_Arıza_Gelen_Bölge_FN_Kombinasyonu_Arıza_Sayısı'})

    # Sonuçların birleştirilmesi ve sütun isimlerinin güncellenmesi
    sonuc_raporu = pd.merge(aylık_istatistikler, aylık_tekrar_eden_ariza_sayisi, on=['Yıl', 'Ay'])
    sonuc_raporu = pd.merge(sonuc_raporu, en_cok_ariza_cozumleyen_takim, on=['Yıl', 'Ay'])
    sonuc_raporu = pd.merge(sonuc_raporu, en_cok_ariza_gelen_bolge_fn, on=['Yıl', 'Ay'])

    # Bölge-FN kombinasyonu için istatistikler
    data['İşOrtağı_Bölge_FN_Kombinasyonu'] = data['İş Ortağı'] + '.' + data['Bölge'] + '.' + data['FN']
    aylık_en_cok_ariza_gelen_bolge_fn = data.groupby(['Yıl', 'Ay', 'İşOrtağı_Bölge_FN_Kombinasyonu']).size().reset_index(name='Arıza_Sayısı').sort_values(by=['Yıl', 'Ay', 'Arıza_Sayısı'], ascending=[True, True, False]).drop_duplicates(subset=['Yıl', 'Ay']).rename(columns={'İşOrtağı_Bölge_FN_Kombinasyonu': 'Aylık_En_Çok_Arıza_Gelen_Bölge_FN_Kombinasyonu', 'Arıza_Sayısı': 'Aylık_En_Çok_Arıza_Gelen_Bölge_FN_Kombinasyonu_Arıza_Sayısı'})

    # Sonuçların birleştirilmesi
    sonuc_raporu = pd.merge(aylık_istatistikler, aylık_tekrar_eden_ariza_sayisi, on=['Yıl', 'Ay'])
    sonuc_raporu = pd.merge(sonuc_raporu, en_cok_ariza_cozumleyen_takim, on=['Yıl', 'Ay'])
    sonuc_raporu = pd.merge(sonuc_raporu, aylık_en_cok_ariza_gelen_bolge_fn, on=['Yıl', 'Ay'])

    # Aylık en çok arıza bırakan abone ve bu abonenin bıraktığı arıza sayısı hesaplama
    aylık_en_cok_ariza_birakan_hizmet_no = data.groupby(['Yıl', 'Ay', 'Hizmet No']).size().reset_index(name='Bıraktığı_Arıza_Sayısı').sort_values(by=['Yıl', 'Ay', 'Bıraktığı_Arıza_Sayısı'], ascending=[True, True, False])

    # Her ay için en çok arıza bırakan aboneyi belirle
    aylık_en_cok_ariza_birakan_hizmet_no = aylık_en_cok_ariza_birakan_hizmet_no.drop_duplicates(subset=['Yıl', 'Ay'])

    # Bu hesaplamaların sonuçlarını ana sonuç raporuna ekleyelim
    sonuc_raporu = pd.merge(sonuc_raporu, aylık_en_cok_ariza_birakan_hizmet_no, on=['Yıl', 'Ay'], how='left').rename(columns={'Hizmet No': 'Aylık_En_Çok_Arıza_Bırakan_Abone_Hizmet_No', 'Bıraktığı_Arıza_Sayısı': 'Aylık_En_Çok_Arıza_Bırakan_Abonenin_Arıza_Sayısı'})

    # Öncelikle '5 Gün İçinde Tekrar' bilgisini kullanarak tekrar eden arızaları filtreleyelim.
    tekrar_edilen_ariza_birakan_hizmet_no = data_sorted[data_sorted['x Gün İçinde Tekrar'] == True].groupby(['Yıl', 'Ay', 'Hizmet No']).size().reset_index(name='Tekrar_Eden_Arıza_Sayısı').sort_values(by=['Yıl', 'Ay', 'Tekrar_Eden_Arıza_Sayısı'], ascending=[True, True, False])

    # Her ay için en çok tekrar eden arıza bırakan aboneyi belirleyelim.
    aylık_en_cok_tekrar_edilen_ariza_birakan_hizmet_no = tekrar_edilen_ariza_birakan_hizmet_no.drop_duplicates(subset=['Yıl', 'Ay'])

    # Bu hesaplamaların sonuçlarını ana sonuç raporuna ekleyelim.
    sonuc_raporu = pd.merge(sonuc_raporu, aylık_en_cok_tekrar_edilen_ariza_birakan_hizmet_no.rename(columns={'Hizmet No': 'Aylık_En_Çok_Tekrar_Eden_Arıza_Bırakan_Abone_Hizmet_No', 'Tekrar_Eden_Arıza_Sayısı': 'Aylık_En_Çok_Tekrar_Eden_Arıza_Bırakan_Abonenin_Tekrar_Eden_Arıza_Sayısı'}), on=['Yıl', 'Ay'], how='left')

    # Aylık arıza bırakan benzersiz abone sayısını hesaplama
    aylık_arıza_bırakan_abone_sayısı = data.groupby(['Yıl', 'Ay'])['Hizmet No'].nunique().reset_index(name='Aylık_Arıza_Bırakan_Abone_Sayısı')

    # Bu hesaplamaların sonuçlarını ana sonuç raporuna ekleyelim
    sonuc_raporu = pd.merge(sonuc_raporu, aylık_arıza_bırakan_abone_sayısı, on=['Yıl', 'Ay'], how='left')

    # Başlık isimlerini güncelleme
    sonuc_raporu = sonuc_raporu.rename(columns={
        'Aylık_Toplam_Arıza_Sayısı': 'Aylık Toplam Arıza Sayısı',
        'Aylık_Toplam_Onaylanan_Arıza_Sayısı': 'Aylık Toplam Onaylanan Arıza Sayısı',
        'Aylık_Toplam_Ret_Edilen_Arıza_Sayısı': 'Aylık Toplam Ret Edilen Arıza Sayısı',
        'Aylık_Ortalama_Çözüm_Süresi': 'Aylık Ortalama Çözüm Süresi (Saat)',
        'Aylık_Tekrar_Eden_Toplam_Arıza_Sayısı': 'Aylık Toplam Tekrar Eden Arıza Sayısı',
        'Aylık_En_Çok_Arıza_Çözümleyen_Takım': 'Aylık En Çok Arıza Çözümleyen Takım',
        'Aylık_En_Çok_Arıza_Çözümleyen_Takımın_Çözümlediği_Arıza_Sayısı': 'Aylık En Çok Arıza Çözümleyen Takımın Çözümlediği Arıza Sayısı',
        'Aylık_En_Çok_Arıza_Gelen_Bölge_FN_Kombinasyonu': 'Aylık En Çok Arıza Gelen Bölge FN',
        'Aylık_En_Çok_Arıza_Gelen_Bölge_FN_Kombinasyonu_Arıza_Sayısı': 'Aylık En Çok Arıza Gelen Bölge FN Arıza Sayısı',
        'Aylık_En_Çok_Tekrar_Eden_Arıza_Çözümleyen_Takım': 'Aylık En Çok Tekrar Eden Arıza Çözümleyen Takım',
        'Aylık_En_Çok_Tekrar_Eden_Arıza_Çözümleyen_Takımın_Çözümlediği_Tekrar_Eden_Arıza_Sayısı': 'Aylık En Çok Tekrar Eden Arıza Çözümleyen Takımın Çözümlediği Tekrar Eden Arıza Sayısı',
        'Aylık_En_Çok_Arıza_Bırakan_Abone_Hizmet_No': 'Aylık En Çok Arıza Bırakan Abone Hizmet No',
        'Aylık_En_Çok_Arıza_Bırakan_Abonenin_Arıza_Sayısı': 'Aylık En Çok Arıza Bırakan Abonenin Arıza Sayısı',
        'Aylık_En_Çok_Tekrar_Eden_Arıza_Bırakan_Abone_Hizmet_No': 'Aylık En Çok Tekrar Eden Arıza Bırakan Abone Hizmet No',
        'Aylık_En_Çok_Tekrar_Eden_Arıza_Bırakan_Abonenin_Tekrar_Eden_Arıza_Sayısı': 'Aylık En Çok Tekrar Eden Arıza Bırakan Abonenin Tekrar Eden Arıza Sayısı',
        'Aylık_Arıza_Bırakan_Abone_Sayısı' : 'Aylık Arıza Bırakan Abone Sayısı'
    })


    # Günlük_Arıza sütununa göre veriyi büyükten küçüğe sıralayın
    #sonuc_raporu = sonuc_raporu.sort_values(by='Aylık Toplam Arıza Sayısı', ascending=False)

    # Raporlama klasörlerinin oluşturulması
    aylık_değerlendirme_folder_path = Path('Aylık Değerlendirme')
    aylık_değerlendirme_folder_path.mkdir(parents=True, exist_ok=True)

    report_folder_path = aylık_değerlendirme_folder_path / 'Genel Durum Değerlendirme'
    report_folder_path.mkdir(parents=True, exist_ok=True)

    # Raporun Excel dosyası olarak kaydedilmesi
    final_output_path = report_folder_path / 'Aylık Genel Durum Değerlendirme Raporu.xlsx'
    sonuc_raporu.to_excel(final_output_path, index=False)


    # Ayarlanacak boyutlar: Başlık satırı yüksekliği 143 piksel, diğer satırlar 66 piksel, tüm sütunların genişliği 10 piksel.
    # Excel'de piksel yerine nokta ve karakter genişliği kullanıldığı için, bu değerler yaklaşık olarak dönüştürülecektir.

    # Pikseli Excel'e uygun birime dönüştürme yaklaşık değerleri
    # Satır yüksekliği doğrudan piksel olarak ayarlanabilir (yaklaşık olarak)
    # Sütun genişliği için Excel genişliği, piksel değerine göre ayarlanmalıdır. Excel genişliği ve piksel arasında doğrudan bir oran yoktur,
    # ancak genel olarak 1 karakter genişliği yaklaşık 7 piksele eşdeğerdir. Bu yüzden, 10 piksel yaklaşık 1.43 karakter genişliğine eşittir.

    # Başlık satırı ve diğer satırlar için yükseklik ayarları
    header_row_height = 132  # Excel'de doğrudan piksel olarak ayarlanabilir
    other_row_height = 125  # Excel'de doğrudan piksel olarak ayarlanabilir

    # Sütun genişliği için pikseli karakter genişliğine çevirme
    column_width_in_pixels = 125
    # Excel'de genişlik karakter cinsindendir, bu yüzden dönüşüm yapılır
    column_width_in_chars = column_width_in_pixels / 7  # Yaklaşık bir dönüşüm

    # Dosyayı yükleyin
    wb = openpyxl.load_workbook(final_output_path)
    ws = wb.active

    # Biçimlendirme ayarları
    header_fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')  # Koyu lacivert arka plan
    white_font = Font(bold=True, color='FFFFFF')  # Başlık için beyaz, kalın yazı tipi
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))  # İnce kenarlık
    centered_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)  # Metni ortala


    # Başlık satırının yüksekliğini ayarla
    ws.row_dimensions[1].height = header_row_height

    # Diğer tüm satırlar için yüksekliği ayarla
    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].height = other_row_height

    # Tüm sütunlar için genişliği ayarla
    for col in ws.iter_cols(min_col=1, max_col=ws.max_column, min_row=1, max_row=1):
        ws.column_dimensions[get_column_letter(col[0].column)].width = column_width_in_chars

    # Tüm hücrelere biçimlendirme uygula
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
            cell.alignment = centered_alignment
            if cell.row == 1:  # Başlık satırı için ek biçimlendirme
                cell.fill = header_fill
                cell.font = white_font

    # Değerleri yuvarlama ve biçimlendirme
    # Yuvarlanacak sütunlar için doğru indeksleri bul
    headers_to_round = ['Aylık Ortalama Çözüm Süresi (Saat)']  # Yuvarlanacak başlıkların listesi
    columns_to_round = []  # Yuvarlanacak sütun indekslerini saklar

    for col in ws.iter_cols(min_row=1, max_row=1, min_col=1, max_col=ws.max_column, values_only=True):
        if col[0] in headers_to_round:
            columns_to_round.append(col[0])

    # Her belirtilen sütun için, sayısal hücreleri yuvarla ve formatı '0.0' olarak ayarla
    for idx, header in enumerate(ws[1], start=1):
        if header.value in columns_to_round:
            for cell in ws[get_column_letter(idx)][2:]:  # Üstbilgiyi atla, 2. satırdan başla
                if isinstance(cell.value, (int, float)):
                    cell.value = round(cell.value, 1)  # Sayısal değeri yuvarla
                    cell.number_format = '0.0'  # Biçimi '0.0' olarak ayarla

    # Değişiklikleri kaydet
    wb.save(final_output_path)



    ###################################################### Grafik Durum Değerlendirme PDF ############################################

    try:

        print_islem_baslat('Genel Durum Değerlendirme', 'Aylık Değerlendirme Grafik PDF')

        # Font ayarlarını güncelleyelim
        matplotlib.rcParams['font.family'] = 'DejaVu Sans'
        matplotlib.rcParams['pdf.fonttype'] = 42

        # Stil ayarını güncelleyelim
        plt.style.use('seaborn-v0_8-darkgrid')

        # Raporlama klasörlerinin oluşturulması
        aylik_degerlendirme_folder_path = Path('Aylık Değerlendirme')
        aylik_degerlendirme_folder_path.mkdir(parents=True, exist_ok=True)

        report_folder_path = aylik_degerlendirme_folder_path / 'Genel Durum Değerlendirme'
        report_folder_path.mkdir(parents=True, exist_ok=True)

        # PDF olarak kaydetme yolu
        final_pdf_path = report_folder_path / 'Aylık Genel Durum Değerlendirme Grafiği.pdf'

        try:
            # Verileri tam sayıya dönüştürme
            sonuc_raporu['Yıl'] = sonuc_raporu['Yıl'].astype(int)
            sonuc_raporu['Ay'] = sonuc_raporu['Ay'].astype(int)
        except:
            pass

        sonuc_raporu['Tarih_str'] = sonuc_raporu['Yıl'].astype(str) + '-' + sonuc_raporu['Ay'].astype(str)

        # Tarih sütununu oluşturma
        sonuc_raporu['Tarih'] = pd.to_datetime(sonuc_raporu['Tarih_str'], format='%Y-%m')

        with PdfPages(final_pdf_path) as pdf:
            for column in sonuc_raporu.columns:
                # 'Tarih' ve 'Tarih_str' sütunlarını hariç tut
                if column not in ['Yıl', 'Ay', 'Tarih', 'Tarih_str']:
                    fig, ax = plt.subplots(figsize=(19, 13))
                    # 'Tarih' sütununu kullanarak plot yapın
                    ax.plot(sonuc_raporu['Tarih'], sonuc_raporu[column],
                            marker='o', linestyle='-', color='navy', linewidth=2, markersize=8)
                    ax.set_title(column, fontsize=16, fontweight='bold', color='navy')
                    ax.set_xlabel('Yıl-Ay', fontsize=14, fontweight='bold')
                    ax.set_ylabel(column, fontsize=14, fontweight='bold')

                    # Tarih formatını ayarlama ve tarihleri dikey yazdırma
                    ax.xaxis.set_major_locator(AutoDateLocator())
                    ax.xaxis.set_major_formatter(DateFormatter('%Y-%m'))
                    ax.tick_params(axis='x', rotation=90, labelsize=10)  # Tarihleri dikey yazdır
                    ax.tick_params(axis='y', labelsize=12)

                    # Her grafiğin sağ üstüne not ekleyelim
                    text = 'Copyright 2024\nCreated by Biosphere.\nDeveloped by Furkan ARINCI'
                    plt.text(0.95, 0.95, text, fontsize=8, verticalalignment='top', horizontalalignment='right',
                             transform=ax.transAxes, color='gray', alpha=0.7)

                    pdf.savefig(fig, dpi=600)  # Grafikleri PDF'e kaydet, DPI ayarını burada yap
                    plt.close(fig)

        # İşlem tamamlandı
        print_islem_bitir('Genel Durum Değerlendirme', 'Aylık Değerlendirme Grafik PDF')

    except Exception as e:
        print(f"Hata: {e}")

    bio_ses('BS05.mp3')

    ################################## Genel Durum Yıllık Değerlendirme İşlemleri ###############################################

    print_islem_baslat('Genel Durum Değerlendirme', 'Yıllık Değerlendirme')

    data=data_original

    # Veri setini yükle
    data = pd.read_excel('Veri Girişi/METROBIL.xlsx')

    # Tarih/saat sütunlarını datetime türüne dönüştürme
    data['Kayıt Tarihi'] = pd.to_datetime(data['Kayıt Tarihi'])
    data['Çözüm Tarihi'] = pd.to_datetime(data['Çözüm Tarihi'])

    # Yıl sütununu ekleyin
    data['Yıl'] = data['Kayıt Tarihi'].dt.year

    # Onay ve Ret durumlarını sayısal değerlere dönüştürme
    data['Onay'] = (data['Durum'] == 'Onaylandı').astype(int)
    data['Ret'] = (data['Ret Geçmişi'] == 'Var').astype(int)

    # Çözüm süresini saat cinsinden hesaplama
    data['Çözüm Süresi Saat'] = (data['Çözüm Tarihi'] - data['Kayıt Tarihi']).dt.total_seconds() / 3600

    # Devre No sütunundaki NaN veya boş değerleri işleme ve metin olarak ele alınması için düzeltme
    data['Devre No'] = data['Devre No'].fillna('').astype(str)

    # Ayrıştırma işlemi
    data[['İş Ortağı', 'Bölge', 'FN', 'Anfi', 'Tap-Off', 'Kol']] = pd.DataFrame(data['Devre No'].apply(lambda x: x.split('.')[:6] if len(x.split('.')) >= 6 else [None] * 6).tolist(), index=data.index)

    # Yıllık istatistiklerin hesaplanması
    yıllık_istatistikler = data.groupby('Yıl').agg(
        Yıllık_Toplam_Arıza_Sayısı=('Çağrı No', 'count'),
        Yıllık_Toplam_Onaylanan_Arıza_Sayısı=('Onay', 'sum'),
        Yıllık_Toplam_Ret_Edilen_Arıza_Sayısı=('Ret', 'sum'),
        Yıllık_Ortalama_Çözüm_Süresi=('Çözüm Süresi Saat', 'mean')
    ).reset_index()

    # Veri setini 'Hizmet No' ve 'Kayıt Tarihi'ne göre sıralayalım ve her bir arıza kaydı için bir önceki kayıtla arasındaki gün farkını hesaplayalım.
    data_sorted = data.sort_values(by=['Hizmet No', 'Kayıt Tarihi'])
    data_sorted['Önceki Kayıt Farkı'] = data_sorted.groupby('Hizmet No')['Kayıt Tarihi'].diff().dt.days

    # Belirlenen gün sayısından daha az olan farkları işaretleyelim.
    data_sorted['Yıllık Tekrar'] = data_sorted['Önceki Kayıt Farkı'] <= BioT

    # 'Yıl' bazında, tekrar eden kayıtların sayısını hesaplayalım.
    yıllık_tekrar_eden_ariza_sayisi = data_sorted.groupby('Yıl')['Yıllık Tekrar'].sum().reset_index(name='Yıllık_Tekrar_Eden_Toplam_Arıza_Sayısı')

    # En çok arıza çözümleyen takımı hesaplama
    en_cok_ariza_cozumleyen_takim = data.groupby(['Yıl', 'Takım']).size().reset_index(name='Çözümlenen_Arıza_Sayısı').sort_values(by=['Yıl', 'Çözümlenen_Arıza_Sayısı'], ascending=[True, False]).drop_duplicates(subset='Yıl').rename(columns={'Takım': 'Yıllık_En_Çok_Arıza_Çözümleyen_Takım', 'Çözümlenen_Arıza_Sayısı': 'Yıllık_En_Çok_Arıza_Çözümleyen_Takımın_Çözümlediği_Arıza_Sayısı'})

    # Bölge-FN kombinasyonu için istatistikler
    data['İşOrtağı_Bölge_FN_Kombinasyonu'] = data['İş Ortağı'] + '.' + data['Bölge'] + '.' + data['FN']
    en_cok_ariza_gelen_bolge_fn = data.groupby(['Yıl', 'İşOrtağı_Bölge_FN_Kombinasyonu']).size().reset_index(name='Arıza_Sayısı').sort_values(by=['Yıl', 'Arıza_Sayısı'], ascending=[True, False]).drop_duplicates(subset='Yıl').rename(columns={'İşOrtağı_Bölge_FN_Kombinasyonu': 'Yıllık_En_Çok_Arıza_Gelen_Bölge_FN_Kombinasyonu', 'Arıza_Sayısı': 'Yıllık_En_Çok_Arıza_Gelen_Bölge_FN_Kombinasyonu_Arıza_Sayısı'})

    # Sonuçların birleştirilmesi
    sonuc_raporu = pd.merge(yıllık_istatistikler, yıllık_tekrar_eden_ariza_sayisi, on='Yıl')
    sonuc_raporu = pd.merge(sonuc_raporu, en_cok_ariza_cozumleyen_takim, on='Yıl')
    sonuc_raporu = pd.merge(sonuc_raporu, en_cok_ariza_gelen_bolge_fn, on='Yıl')

    # Yıllık en çok arıza bırakan abone ve bu abonenin bıraktığı arıza sayısı hesaplama
    yıllık_en_cok_ariza_birakan_hizmet_no = data.groupby(['Yıl', 'Hizmet No']).size().reset_index(name='Bıraktığı_Arıza_Sayısı').sort_values(by=['Yıl', 'Bıraktığı_Arıza_Sayısı'], ascending=[True, False]).drop_duplicates(subset='Yıl').rename(columns={'Hizmet No': 'Yıllık_En_Çok_Arıza_Bırakan_Abone_Hizmet_No', 'Bıraktığı_Arıza_Sayısı': 'Yıllık_En_Çok_Arıza_Bırakan_Abonenin_Arıza_Sayısı'})

    # Sonuç raporuna ekleme
    sonuc_raporu = pd.merge(sonuc_raporu, yıllık_en_cok_ariza_birakan_hizmet_no, on='Yıl', how='left')

    # Yıllık arıza bırakan benzersiz abone sayısını hesaplama
    yıllık_arıza_bırakan_abone_sayısı = data.groupby('Yıl')['Hizmet No'].nunique().reset_index(name='Yıllık_Arıza_Bırakan_Abone_Sayısı')

    # Sonuç raporuna ekleme
    sonuc_raporu = pd.merge(sonuc_raporu, yıllık_arıza_bırakan_abone_sayısı, on='Yıl', how='left')



    # Başlık isimlerini güncelleme
    sonuc_raporu = sonuc_raporu.rename(columns={
        'Yıllık_Toplam_Arıza_Sayısı': 'Yıllık Toplam Arıza Sayısı',
        'Yıllık_Toplam_Onaylanan_Arıza_Sayısı': 'Yıllık Toplam Onaylanan Arıza Sayısı',
        'Yıllık_Toplam_Ret_Edilen_Arıza_Sayısı': 'Yıllık Toplam Ret Edilen Arıza Sayısı',
        'Yıllık_Ortalama_Çözüm_Süresi': 'Yıllık Ortalama Çözüm Süresi (Saat)',
        'Yıllık_Tekrar_Eden_Toplam_Arıza_Sayısı': 'Yıllık Toplam Tekrar Eden Arıza Sayısı',
        'Yıllık_En_Çok_Arıza_Çözümleyen_Takım': 'Yıllık En Çok Arıza Çözümleyen Takım',
        'Yıllık_En_Çok_Arıza_Çözümleyen_Takımın_Çözümlediği_Arıza_Sayısı': 'Yıllık En Çok Arıza Çözümleyen Takımın Çözümlediği Arıza Sayısı',
        'Yıllık_En_Çok_Arıza_Gelen_Bölge_FN_Kombinasyonu': 'Yıllık En Çok Arıza Gelen Bölge FN',
        'Yıllık_En_Çok_Arıza_Gelen_Bölge_FN_Kombinasyonu_Arıza_Sayısı': 'Yıllık En Çok Arıza Gelen Bölge FN Arıza Sayısı',
        'Yıllık_En_Çok_Arıza_Bırakan_Abone_Hizmet_No' : 'Yıllık En Çok Arıza Bırakan Abone Hizmet No',
        'Yıllık_En_Çok_Arıza_Bırakan_Abonenin_Arıza_Sayısı' : 'Yıllık En Çok Arıza Bırakan Abonenin Arıza Sayısı',
        'Yıllık_En_Çok_Tekrar_Eden_Arıza_Bırakan_Abone_Hizmet_No': 'Yıllık En Çok Tekrar Eden Arıza Bırakan Abone Hizmet No',
        'Yıllık_En_Çok_Tekrar_Eden_Arıza_Bırakan_Abonenin_Tekrar_Eden_Arıza_Sayısı': 'Yıllık En Çok Tekrar Eden Arıza Bırakan Abonenin Tekrar Eden Arıza Sayısı',
        'Yıllık_Arıza_Bırakan_Abone_Sayısı': 'Yıllık Arıza Bırakan Abone Sayısı'
    })

    # Raporlama klasörlerinin oluşturulması
    yillik_degerlendirme_folder_path = Path('Yıllık Değerlendirme')
    yillik_degerlendirme_folder_path.mkdir(parents=True, exist_ok=True)

    report_folder_path = yillik_degerlendirme_folder_path / 'Genel Durum Değerlendirme'
    report_folder_path.mkdir(parents=True, exist_ok=True)

    # Raporun Excel dosyası olarak kaydedilmesi
    final_output_path = report_folder_path / 'Yıllık Genel Durum Değerlendirme Raporu.xlsx'
    sonuc_raporu.to_excel(final_output_path, index=False)


    # Ayarlanacak boyutlar: Başlık satırı yüksekliği 143 piksel, diğer satırlar 66 piksel, tüm sütunların genişliği 10 piksel.
    # Excel'de piksel yerine nokta ve karakter genişliği kullanıldığı için, bu değerler yaklaşık olarak dönüştürülecektir.

    # Pikseli Excel'e uygun birime dönüştürme yaklaşık değerleri
    # Satır yüksekliği doğrudan piksel olarak ayarlanabilir (yaklaşık olarak)
    # Sütun genişliği için Excel genişliği, piksel değerine göre ayarlanmalıdır. Excel genişliği ve piksel arasında doğrudan bir oran yoktur,
    # ancak genel olarak 1 karakter genişliği yaklaşık 7 piksele eşdeğerdir. Bu yüzden, 10 piksel yaklaşık 1.43 karakter genişliğine eşittir.

    # Başlık satırı ve diğer satırlar için yükseklik ayarları
    header_row_height = 132  # Excel'de doğrudan piksel olarak ayarlanabilir
    other_row_height = 125  # Excel'de doğrudan piksel olarak ayarlanabilir

    # Sütun genişliği için pikseli karakter genişliğine çevirme
    column_width_in_pixels = 125
    # Excel'de genişlik karakter cinsindendir, bu yüzden dönüşüm yapılır
    column_width_in_chars = column_width_in_pixels / 7  # Yaklaşık bir dönüşüm

    # Dosyayı yükleyin
    wb = openpyxl.load_workbook(final_output_path)
    ws = wb.active

    # Biçimlendirme ayarları
    header_fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')  # Koyu lacivert arka plan
    white_font = Font(bold=True, color='FFFFFF')  # Başlık için beyaz, kalın yazı tipi
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))  # İnce kenarlık
    centered_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)  # Metni ortala


    # Başlık satırının yüksekliğini ayarla
    ws.row_dimensions[1].height = header_row_height

    # Diğer tüm satırlar için yüksekliği ayarla
    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].height = other_row_height

    # Tüm sütunlar için genişliği ayarla
    for col in ws.iter_cols(min_col=1, max_col=ws.max_column, min_row=1, max_row=1):
        ws.column_dimensions[get_column_letter(col[0].column)].width = column_width_in_chars

    # Tüm hücrelere biçimlendirme uygula
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
            cell.alignment = centered_alignment
            if cell.row == 1:  # Başlık satırı için ek biçimlendirme
                cell.fill = header_fill
                cell.font = white_font

    # Değerleri yuvarlama ve biçimlendirme
    # Yuvarlanacak sütunlar için doğru indeksleri bul
    headers_to_round = ['Yıllık Ortalama Çözüm Süresi (Saat)']  # Yuvarlanacak başlıkların listesi
    columns_to_round = []  # Yuvarlanacak sütun indekslerini saklar

    for col in ws.iter_cols(min_row=1, max_row=1, min_col=1, max_col=ws.max_column, values_only=True):
        if col[0] in headers_to_round:
            columns_to_round.append(col[0])

    # Her belirtilen sütun için, sayısal hücreleri yuvarla ve formatı '0.0' olarak ayarla
    for idx, header in enumerate(ws[1], start=1):
        if header.value in columns_to_round:
            for cell in ws[get_column_letter(idx)][2:]:  # Üstbilgiyi atla, 2. satırdan başla
                if isinstance(cell.value, (int, float)):
                    cell.value = round(cell.value, 1)  # Sayısal değeri yuvarla
                    cell.number_format = '0.0'  # Biçimi '0.0' olarak ayarla

    # Değişiklikleri kaydet
    wb.save(final_output_path)

    try:

        ###################################################### Grafik Durum Değerlendirme PDF ############################################

        print_islem_baslat('Genel Durum Değerlendirme', 'Yıllık Değerlendirme Grafik PDF')

        # Font ayarlarını güncelleyelim
        matplotlib.rcParams['font.family'] = 'DejaVu Sans'
        matplotlib.rcParams['pdf.fonttype'] = 42

        # Stil ayarını güncelleyelim
        plt.style.use('seaborn-v0_8-darkgrid')

        # Raporlama klasörlerinin oluşturulması
        yillik_degerlendirme_folder_path = Path('Yıllık Değerlendirme')
        yillik_degerlendirme_folder_path.mkdir(parents=True, exist_ok=True)

        report_folder_path = yillik_degerlendirme_folder_path / 'Genel Durum Değerlendirme'
        report_folder_path.mkdir(parents=True, exist_ok=True)

        # PDF olarak kaydetme yolu
        final_pdf_path = report_folder_path / 'Yıllık Genel Durum Değerlendirme Grafiği.pdf'
        try:
            # Verileri tam sayıya dönüştürme
            sonuc_raporu['Yıl'] = sonuc_raporu['Yıl'].astype(int)

        except:
            pass

        sonuc_raporu['Tarih_str'] = sonuc_raporu['Yıl'].astype(str)

        # 'Yıl' sütununu doğrudan tarih sütunu olarak kullanma
        sonuc_raporu['Tarih'] = pd.to_datetime(sonuc_raporu['Yıl'], format='%Y')

        with PdfPages(final_pdf_path) as pdf:
            for column in sonuc_raporu.columns:
                # 'Tarih' ve 'Tarih_str' sütunlarını hariç tut
                if column not in ['Yıl','Tarih', 'Tarih_str']:
                    fig, ax = plt.subplots(figsize=(19, 13))
                    # 'Tarih' sütununu kullanarak plot yapın
                    ax.plot(sonuc_raporu['Tarih'], sonuc_raporu[column],
                            marker='o', linestyle='-', color='navy', linewidth=2, markersize=8)
                    ax.set_title(column, fontsize=16, fontweight='bold', color='navy')
                    ax.set_xlabel('Yıl', fontsize=14, fontweight='bold')
                    ax.set_ylabel(column, fontsize=14, fontweight='bold')

                    # Tarih formatını ayarlama ve tarihleri dikey yazdırma
                    ax.xaxis.set_major_locator(AutoDateLocator())
                    ax.xaxis.set_major_formatter(DateFormatter('%Y'))
                    ax.tick_params(axis='x', rotation=90, labelsize=10)  # Tarihleri dikey yazdır
                    ax.tick_params(axis='y', labelsize=12)

                    # Her grafiğin sağ üstüne not ekleyelim
                    text = 'Copyright 2024\nCreated by Biosphere.\nDeveloped by Furkan ARINCI'
                    plt.text(0.95, 0.95, text, fontsize=8, verticalalignment='top', horizontalalignment='right',
                             transform=ax.transAxes, color='gray', alpha=0.7)

                    pdf.savefig(fig, dpi=600)  # Grafikleri PDF'e kaydet, DPI ayarını burada yap
                    plt.close(fig)

        # İşlem tamamlandı
        print_islem_bitir('Genel Durum Değerlendirme', 'Yıllık Değerlendirme Grafik PDF')

    except Exception as e:
        print(f"Hata: {e}")

    print_islem_bitir('Genel Durum Değerlendirme', 'Yıllık Değerlendirme')
    print_islem_bitir('Genel Durum Değerlendirme', 'Tüm Değerlendirme')
    bio_ses('BS05.mp3')



print("[. . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . .]")
print("[. . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . .]")
print("[. . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . .]")
print("[.                                                                           .]")
print("[.                           B I O S P H E R E                               .]")
print("[.                       Developed by Furkan ARINCI                          .]")
print("[.                           Copyright © 2024                                .]")
print("[.                                                                           .]")
print("[. . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . .]")
print("[.                             Başlıyoruz...                                 .]")
print("[. . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . .]")
print("[. . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . . .]")
bio_ses('BS01.mp3')

if os.path.exists("Bio_Control.xlsx"):
    try:
        print("[ Hata tespit edildi ]")
        # Dosya varsa sil
        os.remove("BioControl.xlsx")
        print("[ Hata ayıklandı ]")
        print("[ Yeniden başlatıyorum. Lütfen uygulamayı manuel kapatmayınız. Lütfen Bekleyiniz ]")
        bio_ses('BS02.mp3')
        time.sleep(1000)
        sys.exit()
    except:
        pass



# Kontrol mekanizması kullanımı
try:
    BioV, BioCodes, BioT, BioD = bio_control_et()
except Exception as e:
    bio_ses('BS02.mp3')
    print(f"Hata: {e}")
    # Hata durumunda varsayılan değerler
    BioV = "Biosphere S1-JP"
    BioCode = "Girilmez"
    BioT = 5
    BioD = "OFF"

bio_ses('BS03.mp3')
if BioD == 'ON':
    print("[ Biosphere ile Bağlantı Kuruldu! ]")
elif BioD == 'OFF':
    # bio_ses("B01.wav")  # Ses dosyası çalma işlevi varsayımıyla
    print("[ Biosphere Meşgul! ]")
    print("[ Kullanım için hazır değiliz! ]")
    time.sleep(1000)
    sys.exit()

while True:
    # Dosyanın içeriğini kontrol et
    with open('BioControl.txt', 'r') as file:
        content = file.read().strip()

    # Eğer içerik "Girilmez" ise, kullanıcıdan BioCode iste ve dosyayı güncelle
    if content == "Girilmez":
        bio_ses('BS02.mp3')
        main_bio_code = input("[ BioCode'u giriniz: ")
        with open('BioControl.txt', 'w') as file:
            file.write(main_bio_code)
    else:
        # Eğer içerik "Girilmez" değilse, içeriği main_bio_code değişkenine ata
        main_bio_code = content

    # Excel'den alınan BioCode ile kaydedilen BioCode'u karşılaştır
    if main_bio_code not in BioCodes:
        bio_ses('BS02.mp3')
        print("[ BioCode Eşleşmesi Başarısız! BioCode'unuzun süresi dolmuş olabilir. Geliştiricinizden yeni bir BioCode talep ediniz ]")
        main_bio_code = input("[ BioCode'u giriniz: ")
        with open('BioControl.txt', 'w') as file:
            file.write(main_bio_code)
        continue
    else:
        print("[ BioCode Eşleşme Başarılı! ]")
        break

bio_ses('BS04.mp3')
print(f"[ BioCode: {main_bio_code}, BioT: {BioT}, BioD: {BioD} ]")
data_original=Orjinal_Veri_Optimizasyon()

while True:
    print('')
    print("[ 1 ] Tüm Kriterleri Değerlendir \n[ 2 ] Genel Durum Değerlendirmesi \n[ 3 ] Takım Değerlendirmesi \n[ 4 ] Abone Değerlendirmesi \n[ 5 ] Bölge Değerlendirmesi \n[ 6 ] Çıkış")

    secim = input("[ Lütfen bir seçenek seçin: ")

    if secim == "1":
        Takım_Degerlendirme(data_original, BioT)
        Abone_Degerlendirme(data_original, BioT)
        Bölge_Degerlendirme(data_original, BioT)
        Genel_Durum_Degerlendirme(data_original, BioT)
    if secim == "2":
        Genel_Durum_Degerlendirme(data_original, BioT)
    elif secim == "3":
        Takım_Degerlendirme(data_original, BioT)
        Genel_Durum_Degerlendirme(data_original, BioT)
    elif secim == "4":
        Abone_Degerlendirme(data_original, BioT)
        Genel_Durum_Degerlendirme(data_original, BioT)
    elif secim == "5":
        Bölge_Degerlendirme(data_original, BioT)
        Genel_Durum_Degerlendirme(data_original, BioT)
    elif secim == "6":
        print(" [ Biospehere'den Çıkılıyor ]")
        time.sleep(1000)
        sys.exit()
        break
    else:
        print("[ Geçersiz seçenek! Lütfen tekrar deneyin ]")


############       Copyright 2024        #############
############ Developed By Furkan ARINCI  #############
############       BioDNA Biosphere S1-JP ############
###########     Biosphere                  ###########
# 2024.2.25 - 03:00 AM
# -*- coding: utf-8 -*-